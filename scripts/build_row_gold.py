#!/usr/bin/env python3
"""Build independent row-gold from XLSX BOQ files.

Each XLSX enquiry's filled BOQ is transcribed into BoqRow-compatible JSON.
This gold is INDEPENDENT of the prediction pipeline — it must NOT import
src.pipeline, XLSXRowPipeline, or BOQAssembler.

Output: data/real_rfqs/gold/rows/<enquiry_id>.rowgold.json
"""

from __future__ import annotations

import json
import re as _re
import sys
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from src.rules.units import normalize_unit as _normalize_unit

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.domain.xlsx_column_mapper import XLSXColumnMapper

ENQUIRY_DIR = Path("data/real_rfqs/swa_enquiries")
OUT_DIR = Path("data/real_rfqs/gold/rows")
OUT_DIR.mkdir(parents=True, exist_ok=True)

ENQUIRIES: dict[str, tuple[str, dict]] = {
    "02_isro_vssc": (
        "02_isro_vssc/VSSC_BOQ_with_qty.xlsx",
        {"material": 1, "quantity": 3, "unit": 2},
    ),
    "03_zydus_matoda_osd": (
        "03_zydus_matoda_osd/Zydus_Matoda_Insulation_Enquiry.xlsx",
        {"material": 1, "quantity": 3, "unit": 2},
    ),
    "05_zydus_animal_pharmez": (
        "05_zydus_animal_pharmez/Copy of Insulation Enquiry-Zydus Animal Health Expansion Project - Pharmez-Ahmedabad_.xlsx",
        {"material": 1, "unit": 2},
    ),
    "08_sael": (
        "08_sael/Copy of Insulation Enquiry - SAEL.xlsx",
        {"material": 1, "quantity": 3, "unit": 2},
    ),
}

ACTION_KEYWORDS = {
    "supply",
    "install",
    "supply and install",
    "supply & install",
    "furnish and install",
    "furnish & install",
    "furnish",
    "erect",
    "fix",
    "provide",
    "provide and fix",
    "sitc",
    "s.i.t.c",
    "laying",
    "fixing",
    "application",
    "commissioning",
}

_SECTION_HEADER_RE = _re.compile(r"^[A-Z]$|^Option\s*\d+$|^Note[\s:]", _re.IGNORECASE)

_SPEC_PHRASES = (
    "application to be",
    "above mentioned specifications",
    "third party test certificates",
    "nabl accredited labs",
    "contracting agency",
    "manufacturer representative",
    "thermal conductivity",
    "fire performance",
    "water absorption",
    "test to be conducted",
    "to be submitted",
    "to be borne by",
    "to be approved",
    "specifications to be",
    "certified by",
    "installation training",
    "site visit",
    "good workmanship",
    "well ventilated",
    "as per astm",
    "as per en",
    "as per is",
    "compliance to",
    "deviations to be filled",
    "shall be backed up",
    "random lab test",
    "periodic monthly site visit",
    "to be conducted by",
    "class 'o'",
    "class o",
    "elastomeric nitrile",
    "chemically crosslinked",
    "factory backed",
    "as specified for following",
    "low fire propagation",
    "ul94",
    "fm approved",
    "bird screen",
    "aerodynamic performance",
    "shall be provided",
    "no separate measurement",
)

_CATEGORY_WORDS = ("manifold", "ducting", "insulation", "piping", "civil", "structure", "deck")


def _is_section_header(val: str, unit: str = "", qty: Decimal = Decimal("0")) -> bool:
    """Return True if val is a section header / category row, not a billable BOQ item."""
    if not val:
        return False
    v = val.strip()

    # Single letter sections (A, B, C, D)
    if len(v) <= 1 and v.isalpha():
        return True

    # Known patterns (Option 1, Note:, etc.)
    if _SECTION_HEADER_RE.match(v):
        return True

    # All-caps text without quantity (e.g., "THERMAL INSULATION")
    if v.isupper() and len(v) < 80 and qty == 0:
        return True

    # Category names without action keywords, sizes, or qty/unit
    if qty == 0 and not unit:
        has_category = any(w in v.lower() for w in _CATEGORY_WORDS)
        has_action = any(v.lower().startswith(w) for w in ACTION_KEYWORDS)
        has_size = bool(_re.search(r"\d+\s*(mm|cm|m|inch|ft)", v, _re.IGNORECASE))
        if has_category and not has_action and not has_size and len(v) < 60:
            return True

    # Descriptive header text: category word + "will be" / "shall be" / "to be"
    if qty == 0 and not unit:
        header_starts = (
            "manifold will be",
            "drain piping",
            "colour shall be",
            "supply, installation, testing and commissioning",
            "supply, installation, testing & commissioning",
            "supply and installation of thermal insulation",
            "optional",
        )
        if any(v.lower().startswith(w) for w in header_starts):
            return True

    # Very short text that doesn't look like a material size/code
    if len(v) < 5 and not _re.search(r"\d\s*(mm|cm|m|inch|ft|od|id|dia|kg|nos|set)", v, _re.IGNORECASE):
        return True

    # Catch-all: no quantity and no unit on long text = section header / spec paragraph.
    # Short dimension codes (e.g. "15MM", "25mm OD") are kept as real items.
    return bool(qty == 0 and not unit and len(v) > 20)


def _is_spec_paragraph(val: str, unit: str = "", qty: Decimal = Decimal("0")) -> bool:
    """Return True if val is a specification paragraph, not a billable BOQ item.

    Rows with a real unit AND quantity > 0 are always kept — they are billable
    line items even if the description contains spec language.
    """
    # Billable items with qty+unit are never spec paragraphs
    if qty > 0 and unit:
        return False

    if len(val) > 1200:
        return True
    if 150 < len(val) <= 1200:
        phrase_count = sum(1 for p in _SPEC_PHRASES if p in val.lower())
        # Strong spec signal: many spec phrases and no real quantity/unit
        if phrase_count >= 3 and qty == 0 and not unit:
            return True
        # Very strong spec signal
        if phrase_count >= 5 and qty == 0 and not unit:
            return True
    return False


@dataclass
class RowGoldEntry:
    item_no: int
    material: str
    quantity: Decimal
    unit: str
    action: str = "supply"
    grade: str = ""
    dimensions: list[str] = field(default_factory=list)
    standard: list[str] = field(default_factory=list)
    location: str = ""
    source_file: str = ""
    source_sheet: str = ""
    source_row: int = 0
    human_verified: bool = False
    notes: str = ""

    def to_dict(self) -> dict:
        return {
            "item_no": self.item_no,
            "material": self.material,
            "quantity": str(self.quantity),
            "unit": self.unit,
            "action": self.action,
            "grade": self.grade,
            "dimensions": self.dimensions,
            "standard": self.standard,
            "location": self.location,
            "source_file": self.source_file,
            "source_sheet": self.source_sheet,
            "source_row": self.source_row,
            "human_verified": self.human_verified,
            "notes": self.notes,
        }


def _select_boq_sheet(wb: openpyxl.Workbook) -> tuple[str, openpyxl.worksheet.worksheet.Worksheet]:
    for name in wb.sheetnames:
        if "boq" in name.lower():
            return name, wb[name]
    best_name, best_ws, best_rows = "", None, 0
    for name in wb.sheetnames:
        ws = wb[name]
        row_count = sum(1 for row in ws.iter_rows(values_only=True) if any(c for c in row))
        if row_count > best_rows:
            best_rows, best_name, best_ws = row_count, name, ws
    if best_ws is None:
        raise ValueError(f"No non-empty sheets in workbook: {wb.sheetnames}")
    return best_name, best_ws


def _extract_action(description: str) -> str:
    desc_lower = description.lower().strip()
    for kw in sorted(ACTION_KEYWORDS, key=len, reverse=True):
        if desc_lower.startswith(kw):
            return kw
    return "supply"


def _parse_decimal(val: str) -> Decimal:
    if not val or str(val).strip() in ("", "None", "-"):
        return Decimal("0")
    try:
        return Decimal(str(val).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _vssc_needs_merge(row: list[str], mat_col: int, qty_col: int | None, unit_col: int | None) -> bool:
    if mat_col >= len(row):
        return False
    mat = str(row[mat_col]).strip() if row[mat_col] else ""
    qty = str(row[qty_col]).strip() if qty_col is not None and qty_col < len(row) and row[qty_col] else ""
    unit = str(row[unit_col]).strip() if unit_col is not None and unit_col < len(row) and row[unit_col] else ""
    return bool(mat) and not qty and not unit


def build_row_gold(enquiry_id: str, xlsx_path: Path, col_overrides: dict) -> list[RowGoldEntry]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    sheet_name, ws = _select_boq_sheet(wb)

    raw_headers: list[str | None] = []
    raw_rows: list[list[str | None]] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            raw_headers = list(row)
        else:
            raw_rows.append(list(row))

    headers = [str(h).strip() if h else "" for h in raw_headers]
    rows: list[list[str]] = []
    for row in raw_rows:
        rows.append([str(c).strip() if c else "" for c in row])

    mapper = XLSXColumnMapper()
    col_mapping = mapper.map_columns(headers, rows[:5])

    for key, val in col_overrides.items():
        if val is not None:
            attr = {"material": "material_col", "quantity": "quantity_col", "unit": "unit_col"}.get(key)
            if attr:
                setattr(col_mapping, attr, val)

    entries: list[RowGoldEntry] = []
    item_no = 0

    if enquiry_id == "02_isro_vssc":
        i = 0
        while i < len(rows):
            row = rows[i]
            if not any(c for c in row):
                i += 1
                continue
            mat_val = (
                str(row[col_mapping.material_col]).strip()
                if col_mapping.material_col is not None
                and col_mapping.material_col < len(row)
                and row[col_mapping.material_col]
                else ""
            )
            unit_val = (
                str(row[col_mapping.unit_col]).strip()
                if col_mapping.unit_col is not None and col_mapping.unit_col < len(row) and row[col_mapping.unit_col]
                else ""
            )
            qty_val = (
                str(row[col_mapping.quantity_col]).strip()
                if col_mapping.quantity_col is not None
                and col_mapping.quantity_col < len(row)
                and row[col_mapping.quantity_col]
                else ""
            )
            qty = _parse_decimal(qty_val)

            if _is_section_header(mat_val, unit_val, qty):
                i += 1
                continue
            if _is_spec_paragraph(mat_val, unit_val, qty):
                i += 1
                continue

            if _vssc_needs_merge(
                row, col_mapping.material_col, col_mapping.quantity_col, col_mapping.unit_col
            ) and i + 1 < len(rows):
                next_row = rows[i + 1]
                next_mat = (
                    str(next_row[col_mapping.material_col]).strip()
                    if col_mapping.material_col is not None
                    and col_mapping.material_col < len(next_row)
                    and next_row[col_mapping.material_col]
                    else ""
                )
                next_qty = (
                    str(next_row[col_mapping.quantity_col]).strip()
                    if col_mapping.quantity_col is not None
                    and col_mapping.quantity_col < len(next_row)
                    and next_row[col_mapping.quantity_col]
                    else ""
                )
                next_unit = (
                    str(next_row[col_mapping.unit_col]).strip()
                    if col_mapping.unit_col is not None
                    and col_mapping.unit_col < len(next_row)
                    and next_row[col_mapping.unit_col]
                    else ""
                )
                if next_mat and (next_qty or next_unit):
                    item_no += 1
                    entries.append(
                        RowGoldEntry(
                            item_no=item_no,
                            material=next_mat,
                            quantity=_parse_decimal(next_qty),
                            unit=_normalize_unit(next_unit),
                            action=_extract_action(next_mat),
                            source_file=xlsx_path.name,
                            source_sheet=sheet_name,
                            source_row=i + 2,
                        )
                    )
                    i += 2
                    continue
            if mat_val:
                item_no += 1
                entries.append(
                    RowGoldEntry(
                        item_no=item_no,
                        material=mat_val,
                        quantity=qty,
                        unit=_normalize_unit(unit_val) if unit_val else "no.",
                        action=_extract_action(mat_val),
                        source_file=xlsx_path.name,
                        source_sheet=sheet_name,
                        source_row=i + 2,
                    )
                )
            i += 1

    elif enquiry_id == "03_zydus_matoda_osd":
        # Enquiry 03 has a clean BOQ sheet with standard columns but merged
        # section headers (A/B/C/D, Option 2/3) that generic filtering misses.
        ws = wb["BOQ"]
        raw_rows = list(ws.iter_rows(values_only=True))
        header_row = [str(c).strip() if c else "" for c in raw_rows[0]]
        if "DESCRIPTION" not in header_row and "DESCRIPTION" not in [h.upper() for h in header_row]:
            raise RuntimeError(f"Unexpected header layout: {header_row}")
        mat_col = header_row.index("DESCRIPTION")
        unit_col = header_row.index("UNIT")
        qty_col = header_row.index("QUANTITY")

        def _is_section_header_label(val: str) -> bool:
            if not val:
                return True
            v = val.strip()
            if not v:
                return True
            if len(v) > 250:
                return True
            if v.lower() in {"a", "b", "c", "d", "e"}:
                return True
            return bool(v.lower().startswith("option "))

        def _has_valid_unit_and_qty(unit_val: str, qty_raw) -> bool:
            has_unit = bool(unit_val and unit_val.strip() and unit_val.strip() not in ("", "None", "nan"))
            if not has_unit:
                return False
            if qty_raw is None:
                return False
            s = str(qty_raw).strip()
            return s not in ("", "None", "nan", "-")

        for r_idx, row in enumerate(raw_rows[1:], start=2):
            cells = list(row)
            if not any(c for c in cells):
                continue
            if len(cells) <= mat_col:
                continue
            mat_val = str(cells[mat_col]).strip() if cells[mat_col] else ""
            unit_val = str(cells[unit_col]).strip() if unit_col < len(cells) and cells[unit_col] else ""
            qty_raw = cells[qty_col] if qty_col < len(cells) else None
            if not _has_valid_unit_and_qty(unit_val, qty_raw):
                continue
            if _is_section_header_label(mat_val):
                continue
            item_no += 1
            entries.append(
                RowGoldEntry(
                    item_no=item_no,
                    material=mat_val,
                    quantity=_parse_decimal(str(qty_raw)),
                    unit=_normalize_unit(unit_val) if unit_val else "no.",
                    action=_extract_action(mat_val),
                    source_file=xlsx_path.name,
                    source_sheet="BOQ",
                    source_row=r_idx,
                )
            )

    elif enquiry_id == "05_zydus_animal_pharmez":
        total_idx = next(
            (i for i, h in enumerate(headers) if h.lower().strip() == "total"), len(rows[0]) if rows else 0
        )
        for r_idx, row in enumerate(rows):
            if not any(c for c in row):
                continue
            mat_val = (
                str(row[col_mapping.material_col]).strip()
                if col_mapping.material_col is not None
                and col_mapping.material_col < len(row)
                and row[col_mapping.material_col]
                else ""
            )
            unit_val = (
                str(row[col_mapping.unit_col]).strip()
                if col_mapping.unit_col is not None and col_mapping.unit_col < len(row) and row[col_mapping.unit_col]
                else ""
            )
            qty = Decimal("0")
            if col_mapping.quantity_col is not None and col_mapping.quantity_col < len(row):
                qty = _parse_decimal(str(row[col_mapping.quantity_col]))

            if _is_section_header(mat_val, unit_val, qty):
                continue
            if _is_spec_paragraph(mat_val, unit_val, qty):
                continue
            if not mat_val:
                continue

            # Wide-matrix: sum system columns; if empty, fallback to TOTAL
            system_qty = sum(_parse_decimal(str(row[c])) for c in range(3, total_idx) if c < len(row))
            total_qty = _parse_decimal(str(row[total_idx])) if total_idx < len(row) else Decimal("0")
            best_qty = system_qty if system_qty > 0 else total_qty

            item_no += 1
            entries.append(
                RowGoldEntry(
                    item_no=item_no,
                    material=mat_val,
                    quantity=best_qty,
                    unit=_normalize_unit(unit_val) if unit_val else "no.",
                    action=_extract_action(mat_val),
                    source_file=xlsx_path.name,
                    source_sheet=sheet_name,
                    source_row=r_idx + 2,
                )
            )
    else:
        for r_idx, row in enumerate(rows):
            if not any(c for c in row):
                continue
            mat_val = (
                str(row[col_mapping.material_col]).strip()
                if col_mapping.material_col is not None
                and col_mapping.material_col < len(row)
                and row[col_mapping.material_col]
                else ""
            )
            unit_val = (
                str(row[col_mapping.unit_col]).strip()
                if col_mapping.unit_col is not None and col_mapping.unit_col < len(row) and row[col_mapping.unit_col]
                else ""
            )
            qty_val = (
                str(row[col_mapping.quantity_col]).strip()
                if col_mapping.quantity_col is not None
                and col_mapping.quantity_col < len(row)
                and row[col_mapping.quantity_col]
                else ""
            )
            qty = _parse_decimal(qty_val)

            if _is_section_header(mat_val, unit_val, qty):
                continue
            if _is_spec_paragraph(mat_val, unit_val, qty):
                continue
            if not mat_val:
                continue

            item_no += 1
            entries.append(
                RowGoldEntry(
                    item_no=item_no,
                    material=mat_val,
                    quantity=qty,
                    unit=_normalize_unit(unit_val) if unit_val else "no.",
                    action=_extract_action(mat_val),
                    source_file=xlsx_path.name,
                    source_sheet=sheet_name,
                    source_row=r_idx + 2,
                )
            )

    return entries


def main() -> None:
    print(f"Building row gold → {OUT_DIR}")
    for enquiry_id, (xlsx_rel, col_overrides) in ENQUIRIES.items():
        xlsx_path = ENQUIRY_DIR / xlsx_rel
        if not xlsx_path.exists():
            print(f"  [SKIP] {enquiry_id}: file not found {xlsx_path}")
            continue
        print(f"  Processing {enquiry_id}: {xlsx_path.name}")
        try:
            entries = build_row_gold(enquiry_id, xlsx_path, col_overrides)
        except Exception as e:
            print(f"  [ERROR] {enquiry_id}: {e}")
            import traceback

            traceback.print_exc()
            continue
        out_path = OUT_DIR / f"{enquiry_id}.rowgold.json"
        payload = {
            "doc_id": enquiry_id,
            "source_file": xlsx_path.name,
            "date": str(date.today()),
            "human_verified": False,
            "method": "independent-xlsx-transcription",
            "entries": [e.to_dict() for e in entries],
        }
        with open(out_path, "w") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        print(f"  → {out_path.name}: {len(entries)} rows")

    print("\nRow gold built. Please review entries in:")
    for enquiry_id in ENQUIRIES:
        p = OUT_DIR / f"{enquiry_id}.rowgold.json"
        print(f"  {p}")
    print("\nSet human_verified:true after review before using as gold.")


if __name__ == "__main__":
    main()
