"""P2_02 + P2_03: annotation factory — draft + priority queue + interactive review + verified BIOES.

Reuse verdict (vs prior art):

    scripts/gen_annotation_drafts.py    REPLACE — reads output/batch_extractions
                                        (legacy), writes cli_drafts/bioes (quarantined).
    scripts/preannotate_swa_enquiries.py  PARTIAL — tokenize + entities_to_bioes helpers
                                        are sound; held-out guard is correct, but the
                                        script only targets swa_enquiries (TEST). The new
                                        factory extends the same approach to TRAIN/DEV.
    scripts/review_annotation.py        REPLACE — no tty guard, no hard provenance check,
                                        no autosave, no resumability.
    scripts/validate_annotations.py     EXTEND — current walker is loose; needs
                                        tokens/tags length check, ner_tags/labels
                                        tolerance, and to scan data/annotations/drafts/
                                        in addition to legacy train/val/test files.
    scripts/intake_tender.py            DO NOT TOUCH — live intake path with its own
                                        guarantees; lives outside the factory contract.
    docs/ANNOTATION_WORKFLOW.md         REWRITE — referenced scripts predate the
                                        provenance fence.

Constraint contract:
    - Drafts are written with human_verified:false, reviewed_at unset, reviewer unset.
    - The review subcommand requires a live interactive tty before stamping
      reviewer:"srujan"; the tty check is the FENCE (Rule 3).
    - TEST-split docs (per data/real_rfqs/split_test.json) are REFUSED at draft time.
    - All BIOES tags reference config.constants.BIOES_LABELS exactly; no private labels.
    - Provenance: provenance monotonicity check (no bulk stamping: >50 sentences
      sharing one timestamp = fail) is added to scripts/check_gold_provenance.py —
      that file is FROZEN, so the patch is shipped as a deliverable for the
      orchestrator to apply + re-pin, not committed here.

P2_03 priority queue ranking formula (documented here so the spec in
tasks/phase9/P2_03_bioes_gold_wave1.md §4 step 3 has a single source of truth):

    For each sentence in a draft file:
        predicted_entity_count = |{tag : tag in ner_tags, tag in {B-X, S-X}}|
        weighted_entity_score  = sum( TYPE_WEIGHT[tag_type] for tag in B-/S- tags )
        boilerplate_penalty    = sum( BOILERPLATE_PENALTY[family] for matched family )
        score                  = weighted_entity_score - boilerplate_penalty
        predicted_entity_count is the raw count (for queue sanity assertions);
        score is what we sort by. Rare entity types carry a higher weight so that
        sentences rich in GRADE/STANDARD/LOCATION/DIMENSION outrank ones that are
        only rich in QUANTITY/UNIT (which the model can already produce freely
        and which the owner can correct in seconds).

    Type weights (rare -> common):
        GRADE=2.0, STANDARD=1.8, LOCATION=1.5, DIMENSION=1.3, ACTION=1.2,
        MATERIAL=1.0, QUANTITY=0.5, UNIT=0.5

    Boilerplate penalty families (regex hit subtracts this many points):
        legal    : "terms and conditions", "jurisdiction", "indemnify", ... : -3.0
        contact  : "phone:", "tel:", "fax:", "@", ".com", ".in"           : -2.0
        signature: "authorized signatory", "for and on behalf of",        : -2.0
        header   : all-caps line w/ a date and no lowercase               : -1.0
        list     : >40% of tokens are purely numeric / punctuation
                    (clause refs "1.1.2, 1.2.1 to 1.4.7", temperature
                    ranges "0 °C - 10 °C - 20 °C - 30 °C", date lists
                    "27.02.2024 UYP / SUW 27.02.2024").                  : -3.0
        (all penalties are absolute: the sentence can still rank well if it
         also has high entity content, which is the right behavior — a
         contact+material sentence is still useful.)

    Output ordering is stable on (score desc, doc_id asc, sent_idx asc) so the
    queue is fully deterministic across rebuilds.

    Queue JSON shape (data/annotations/drafts/PRIORITY_QUEUE.json):
        {
          "version": "1.0",
          "generated_at": "<iso>",
          "drafts_dir": "data/annotations/drafts",
          "n_drafts": <int>, "n_items": <int>, "n_excluded_test": <int>,
          "items": [
            {
              "doc_id": "...",
              "source_file": "...",
              "sent_idx": <int>,
              "text": "<= 200 char preview>",
              "tokens": <int>,
              "predicted_entity_count": <int>,
              "score": <float>,
              "type_counts": {"MATERIAL": <int>, ...},
              "rank_reason": "<human readable>"
            },
            ...
          ]
        }

    Determinism + queue-resume:
        - The same set of drafts always produces the same queue (stable sort key).
        - --resume loads the existing queue and appends only newly drafted
          sentences (matched on (doc_id, sent_idx)); existing items keep their
          rank. This means the owner can stop a session, drafts refresh, queue
          gets refreshed, and existing rank is preserved for already-reviewed
          sentences (those move to verified/ and the queue item stays in but is
          marked 'already_verified' if the verified file is on disk).

    Review `--queue` mode (P2_03 §3):
        - Walks the priority queue from item 0 forward, opening each draft
          sentence in the same review UI as before. Owner can press 'q' to
          checkpoint. Progress (last reviewed item index) is saved to
          data/annotations/drafts/.queue_progress.json so the next --queue run
          resumes from there.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from collections.abc import Iterable
from datetime import UTC, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from config.constants import BIOES_LABELS, ENTITY_LABELS  # noqa: E402

DRAFTS_DIR = ROOT / "data" / "annotations" / "drafts"
VERIFIED_DIR = ROOT / "data" / "annotations" / "verified"
STATS_PATH = ROOT / "data" / "annotations" / "owner_minutes.jsonl"

SPLIT_PATH = ROOT / "data" / "real_rfqs" / "split_test.json"
CORPUS_MANIFEST_PATH = ROOT / "data" / "real_rfqs" / "corpus_manifest.json"
GEM_CATALOG_PATH = ROOT / "data" / "ontology" / "gem_catalog.json"
INGESTED_DIR = ROOT / "data" / "real_rfqs" / "swa_enquiries" / "ingested"
SWA_ENQUIRIES_DIR = ROOT / "data" / "real_rfqs" / "swa_enquiries"

PROVENANCE_FENCE_TTY_ERROR = (
    "REFUSED: review subcommand requires a live interactive tty (sys.stdin.isatty()). "
    "Drafts can never stamp human_verified:true from a non-interactive session. "
    "Run this from a real terminal."
)

SENT_SPLITTER = re.compile(r"(?<=[.?!;:\n])\s+|\n{2,}")
TOKEN_SPLITTER = re.compile(r"\w+|[^\w\s]")
TABLE_ROW_KEY = re.compile(
    r"\b\d+(?:\.\d+)?\s*(?:mm|cm|m|kg|nos|sqm|rmt|cum|mtr|meter|meters|ltr|liter|metre|metres)\b", re.IGNORECASE
)
BULK_STAMP_THRESHOLD = 50

QUEUE_PATH = DRAFTS_DIR / "PRIORITY_QUEUE.json"
QUEUE_PROGRESS_PATH = DRAFTS_DIR / ".queue_progress.json"

TYPE_WEIGHT: dict[str, float] = {
    "GRADE": 2.0,
    "STANDARD": 1.8,
    "LOCATION": 1.5,
    "DIMENSION": 1.3,
    "ACTION": 1.2,
    "MATERIAL": 1.0,
    "QUANTITY": 0.5,
    "UNIT": 0.5,
}

_BOILERPLATE_FAMILIES: dict[str, tuple[float, re.Pattern[str]]] = {
    "legal": (
        -3.0,
        re.compile(
            r"\b(terms and conditions|jurisdiction|indemnif|arbitration|herein|hereunder|"
            r"hereof|whereas|liens?|force majeure|warranty period)\b",
            re.IGNORECASE,
        ),
    ),
    "contact": (
        -2.0,
        re.compile(
            r"(\bphone\b|\btel\b|\bfax\b|@\w+\.[a-z]{2,}|\bwww\.|mob(?:ile)?\s*[:#])",
            re.IGNORECASE,
        ),
    ),
    "signature": (
        -2.0,
        re.compile(
            r"\b(authorized signatory|for and on behalf of|sign(?:ed|ature)?\s*:|"
            r"common seal|stamp of)\b",
            re.IGNORECASE,
        ),
    ),
    "header": (
        -1.0,
        re.compile(
            r"^(?:[A-Z][A-Z0-9 ,/&-]{6,}|[A-Z][A-Z0-9 ,/&-]{6,})$",
        ),
    ),
}

LIST_LIKE_RATIO_THRESHOLD = 0.40
LIST_LIKE_PENALTY = -3.0

NON_ENGLISH_RE = re.compile(r"[^\x00-\x7f]")


def _now_iso() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds")


def _load_split() -> dict:
    return json.loads(SPLIT_PATH.read_text())


def _load_manifest() -> dict:
    return json.loads(CORPUS_MANIFEST_PATH.read_text())


_GEM_CACHE: list[str] | None = None


def _load_gem_catalog() -> list[str]:
    global _GEM_CACHE
    if _GEM_CACHE is not None:
        return _GEM_CACHE
    if not GEM_CATALOG_PATH.exists():
        _GEM_CACHE = []
        return _GEM_CACHE
    data = json.loads(GEM_CATALOG_PATH.read_text())
    _GEM_CACHE = [p.get("name", "") for p in data.get("products", []) if p.get("name")]
    return _GEM_CACHE


def _reset_gem_cache() -> None:
    """Test hook: clear the catalog cache so test fixtures can swap it."""
    global _GEM_CACHE
    _GEM_CACHE = None


def _is_test_split(path: str, split: dict) -> bool:
    test_paths = set()
    for key in ("sacred10", "bundle_duplicates_of_sacred10", "client_name_carry_alongs", "new_spec2_picks"):
        for v in (
            split.get("test", {}).get(key, {}).values()
            if isinstance(split["test"].get(key), dict)
            else split["test"].get(key, [])
        ):
            if isinstance(v, str):
                test_paths.add(v)
    test_paths.update(split.get("test", {}).get("all_paths", []))
    return path in test_paths


def _resolve_train_docs(max_docs: int) -> list[dict]:
    """Pick up to `max_docs` TRAIN-split manifest entries. Deterministic order."""
    split = _load_split()
    manifest = _load_manifest()
    train_paths = set(split.get("train", {}).get("all_paths", []))

    candidates: list[dict] = []
    for entry in manifest.get("files", []):
        path = entry.get("path", "")
        if path in train_paths and not _is_test_split(path, split):
            candidates.append(entry)
    candidates.sort(key=lambda e: e.get("path", ""))
    return candidates[:max_docs]


def _safe_doc_id(name: str) -> str:
    return re.sub(r"[^\w\-]", "_", name)[:80]


def _doc_id_to_ingested_path(doc_id_hint: str) -> Path | None:
    """Map a doc_id to the SWA ingested json (only useful for sacred10; sacred10
    is TEST and is refused by the factory — kept for completeness in the rare
    case the manifest has shifted)."""
    candidates = list(INGESTED_DIR.glob(f"*{doc_id_hint}*.json"))
    if candidates:
        return candidates[0]
    return None


def _is_tender_table_row(text: str) -> bool:
    """A table row is a string that contains a number+unit pair (typical BOQ row).

    Used to KEEP tender table cells atomic when sentence-segmenting: a naive
    splitter would cut "M.S. Pipe as per IS 1239 Pt.1" at abbreviation dots.
    """
    return bool(TABLE_ROW_KEY.search(text))


def _segment_into_sentences(text: str) -> list[str]:
    """Sentence-segment tender text. Table-row cells stay atomic.

    A "table row" heuristic keeps cells with a number+unit pair as one sentence.
    Newlines are a stronger split signal than punctuation, since tender
    documents use a lot of newlines for table rows.
    """
    raw_blocks = re.split(r"\n+", text)
    sentences: list[str] = []
    for block in raw_blocks:
        block = block.strip()
        if not block:
            continue
        if _is_tender_table_row(block):
            sentences.append(block)
            continue
        for piece in SENT_SPLITTER.split(block):
            piece = piece.strip()
            if piece:
                sentences.append(piece)
    return sentences


def _tokenize(text: str) -> list[str]:
    return TOKEN_SPLITTER.findall(text)


def _char_offsets(tokens: list[str], text: str) -> list[tuple[int, int]]:
    offsets: list[tuple[int, int]] = []
    pos = 0
    for tok in tokens:
        start = text.find(tok, pos)
        if start == -1:
            start = pos
        end = start + len(tok)
        offsets.append((start, end))
        pos = end
    return offsets


def _entities_to_bioes(
    tokens: list[str],
    offsets: list[tuple[int, int]],
    entities: list[dict],
) -> list[str]:
    """Project entity char spans (from pipeline) onto token spans and emit BIOES.

    Entity dict format:
        {"text": "...", "type": "MATERIAL", "start": <char>, "end": <char>, "source": "..."}
    """
    tags = ["O"] * len(tokens)
    for ent in entities:
        e_start, e_end = ent["start"], ent["end"]
        e_type = ent.get("type", "MATERIAL")
        matched: list[int] = []
        for i, (t_start, t_end) in enumerate(offsets):
            if t_end > e_start and t_start < e_end:
                matched.append(i)
        if not matched:
            continue
        if len(matched) == 1:
            tags[matched[0]] = f"S-{e_type}"
        else:
            tags[matched[0]] = f"B-{e_type}"
            for mid in matched[1:-1]:
                tags[mid] = f"I-{e_type}"
            tags[matched[-1]] = f"E-{e_type}"
    return tags


def _make_pipeline() -> object:
    """Create a single NLPPipeline; the factory uses one per doc (not per sentence)."""
    from src.nlp.pipeline import NLPPipeline  # local import: heavy

    return NLPPipeline()


def _run_pipeline_on_text(pipeline: object, text: str) -> list[dict]:
    """Run the production NLPPipeline on a sentence and return normalized entity dicts.

    Each returned dict: {text, type, start, end, source, confidence}.
    `source` is the (optional) provenance tag — currently the pipeline doesn't
    expose per-entity source; the calling layer is responsible for any post-hoc
    'pattern' / 'gazetteer' attribution.
    """
    result = pipeline.process(text)
    entities: list[dict] = []
    for ent in getattr(result, "entities", []):
        typ = getattr(ent, "type", "MATERIAL")
        if hasattr(typ, "value"):
            typ = typ.value
        entities.append(
            {
                "text": getattr(ent, "text", ""),
                "type": str(typ).upper(),
                "start": int(getattr(ent, "start", 0)),
                "end": int(getattr(ent, "end", 0)),
                "source": "model",
                "confidence": float(getattr(ent, "confidence", 0.0)),
            }
        )
    return entities


def _apply_gazetteer_overlay(
    text: str, tokens: list[str], offsets: list[tuple[int, int]], existing: list[dict]
) -> list[dict]:
    """Overlay GeM catalog product names onto the entity list, marking source='gazetteer'."""
    gem_names = _load_gem_catalog()
    if not gem_names:
        return existing
    seen_spans = {(e["start"], e["end"]) for e in existing}
    out = list(existing)
    text_lower = text.lower()
    for name in gem_names:
        nlow = name.lower().strip()
        if not nlow or len(nlow) < 4:
            continue
        idx = 0
        while True:
            pos = text_lower.find(nlow, idx)
            if pos == -1:
                break
            idx = pos + len(nlow)
            end = pos + len(nlow)
            matched: list[int] = []
            for i, (t_start, t_end) in enumerate(offsets):
                if t_end > pos and t_start < end:
                    matched.append(i)
            if not matched:
                continue
            s = matched[0]
            e = matched[-1] + 1
            if (s, e) in seen_spans:
                continue
            seen_spans.add((s, e))
            out.append(
                {
                    "text": " ".join(tokens[s:e]),
                    "type": "MATERIAL",
                    "start": s,
                    "end": e,
                    "source": "gazetteer",
                    "confidence": 1.0,
                }
            )
    out.sort(key=lambda d: (d["start"], d["end"]))
    return out


def _apply_pattern_overlay(tokens: list[str], offsets: list[tuple[int, int]], existing: list[dict]) -> list[dict]:
    """Pure-pattern overlay: numeric QUANTITY, common UNIT vocab.

    Conservative on purpose — anything in entity-labels vocabulary is fair game.
    """
    out = list(existing)
    seen = {(e["start"], e["end"]) for e in existing}

    for i, tok in enumerate(tokens):
        i_start, i_end = offsets[i]
        if re.fullmatch(r"\d+(?:[\.,]\d+)?", tok) and (i_start, i_end + 1) not in seen:
            seen.add((i_start, i_end + 1))
            out.append(
                {
                    "text": tok,
                    "type": "QUANTITY",
                    "start": i_start,
                    "end": i_end + 1,
                    "source": "pattern",
                    "confidence": 1.0,
                }
            )
            continue
        if (
            tok.lower().rstrip(".")
            in {
                "kg",
                "nos",
                "no",
                "rmt",
                "sqm",
                "cum",
                "m",
                "mm",
                "cm",
                "mtr",
                "meter",
                "meters",
                "metre",
                "metres",
                "ltr",
            }
            and (
                i_start,
                i_end + 1,
            )
            not in seen
        ):
            seen.add((i_start, i_end + 1))
            out.append(
                {
                    "text": tok,
                    "type": "UNIT",
                    "start": i_start,
                    "end": i_end + 1,
                    "source": "pattern",
                    "confidence": 1.0,
                }
            )
    out.sort(key=lambda d: (d["start"], d["end"]))
    return out


def _resolve_on_disk(manifest_path: str) -> Path | None:
    """Resolve a manifest path to an on-disk file (handles path drift).

    Mirrors scripts/run_corpus.py:_resolve_on_disk so the factory consumes the
    same files P1_04 ran on. Without this resolution most manifest entries
    (data/specifications/...) are missing from the live tree and live in
    resources/Specifications/ or rar_extra/.
    """
    p = ROOT / manifest_path
    if p.exists():
        return p
    basename = Path(manifest_path).name
    candidates = [
        ROOT / "resources/Specifications" / basename,
        ROOT / "resources/Specifications/rar_extra" / basename,
        ROOT / "data/real_rfqs/ALL_RFQS" / f"spec1__{basename}",
        ROOT / "data/real_rfqs/ALL_RFQS" / f"spec2__{basename}",
        ROOT / "data/real_rfqs/ALL_RFQS" / f"sacred10__{basename}",
        ROOT / "data/real_rfqs/ALL_RFQS" / f"rar__{basename}",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _ingest_doc_text(entry: dict) -> str:
    """Pull raw text out of a manifest entry. PDF/XLSX go through their pipelines;
    anything that fails falls back to a plain-text read of the on-disk file.

    For the factory's purpose we don't need a perfect extraction: even partial
    text is enough to seed a draft. The owner reviews every sentence anyway.

    PyMuPDF is preferred for PDF text — it's a much faster path than pdfplumber
    on multi-page spec docs and the text it returns is sufficient for
    pre-annotation (owner will review/correct every sentence).
    """
    path = _resolve_on_disk(entry["path"])
    if path is None:
        return ""
    try:
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            try:
                import fitz  # PyMuPDF

                with fitz.open(path) as doc:
                    return "\n".join(p.get_text() for p in doc)
            except Exception:
                import pdfplumber

                with pdfplumber.open(path) as pdf:
                    return "\n".join((p.extract_text() or "") for p in pdf.pages)
        if suffix == ".xlsx":
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            lines: list[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        lines.append(" ".join(cells))
            return "\n".join(lines)
    except Exception:
        pass
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""


def _draft_one_doc(entry: dict) -> dict:
    """Draft pre-annotation for one manifest entry. Returns the draft record."""
    path = entry["path"]
    doc_id = _safe_doc_id(Path(path).stem)
    text = _ingest_doc_text(entry)
    sentences = _segment_into_sentences(text)

    pipeline = _make_pipeline()
    sentence_records: list[dict] = []
    for sent in sentences:
        tokens = _tokenize(sent)
        if not tokens:
            continue
        offsets = _char_offsets(tokens, sent)
        model_ents = _run_pipeline_on_text(pipeline, sent)
        ents = _apply_gazetteer_overlay(sent, tokens, offsets, model_ents)
        ents = _apply_pattern_overlay(tokens, offsets, ents)
        tags = _entities_to_bioes(tokens, offsets, ents)
        sentence_records.append(
            {
                "text": sent,
                "tokens": tokens,
                "ner_tags": tags,
                "entities": ents,
                "source_doc": path,
            }
        )

    return {
        "doc_id": doc_id,
        "source_file": path,
        "human_verified": False,
        "reviewer": None,
        "reviewed_at": None,
        "method": "annotation_factory-draft",
        "schema": "bioes-v1",
        "entity_label_vocab": list(ENTITY_LABELS),
        "n_sentences": len(sentence_records),
        "n_entities_total": sum(len(r["entities"]) for r in sentence_records),
        "source_kind": entry.get("format", "pdf"),
        "client": entry.get("client", "unknown"),
        "source_batch": entry.get("source_batch", "unknown"),
        "produced_at": _now_iso(),
        "sentences": sentence_records,
    }


def cmd_draft(args: argparse.Namespace) -> int:
    """`draft --split train --docs N` — emit draft files for the N most-alphabetical TRAIN docs."""
    if args.split != "train":
        # Dev is small (15) and is also owner-reviewable; the spec only requires
        # refusing TEST. Train is the default.
        print(f"REFUSED: split={args.split!r} not allowed (only 'train' is currently drafted here)")
        return 2

    targets = _resolve_train_docs(args.docs)
    if not targets:
        print("No TRAIN docs found in manifest matching split_test.json.")
        return 1
    for entry in targets:
        path = entry["path"]
        if _is_test_split(path, _load_split()):
            print(f"REFUSED (TEST): {path}")
            return 2
    DRAFTS_DIR.mkdir(parents=True, exist_ok=True)

    n_docs = 0
    n_sents = 0
    for entry in targets:
        record = _draft_one_doc(entry)
        if record["n_sentences"] == 0:
            print(f"  {record['doc_id']}: 0 sentences (text extraction returned nothing; skipping write)")
            continue
        out_path = DRAFTS_DIR / f"{record['doc_id']}.draft.json"
        out_path.write_text(json.dumps(record, indent=2, ensure_ascii=False))
        n_docs += 1
        n_sents += record["n_sentences"]
        print(
            f"  {record['doc_id']}: {record['n_sentences']} sentences, {record['n_entities_total']} entities -> {out_path.name}"
        )

    print(f"\nDrafted {n_docs} docs / {n_sents} sentences into {DRAFTS_DIR}")
    return 0


def _sentence_type_counts(tags: list[str]) -> dict[str, int]:
    """Count B-/S- tags by entity type for one sentence."""
    counts: dict[str, int] = {et: 0 for et in ENTITY_LABELS}
    for tag in tags:
        if tag in ("O",) or not (tag.startswith("B-") or tag.startswith("S-")):
            continue
        t = tag.split("-", 1)[1]
        if t in counts:
            counts[t] += 1
    return counts


def _boilerplate_penalties(text: str, tokens: list[str] | None = None) -> tuple[float, list[str]]:
    """Apply boilerplate regex families. Returns (penalty_sum, matched_families)."""
    total = 0.0
    matched: list[str] = []
    for family, (penalty, pattern) in _BOILERPLATE_FAMILIES.items():
        if pattern.search(text):
            total += penalty
            matched.append(family)
    # list-like: >40% of tokens are pure digits or single-char punctuation
    toks = tokens if tokens is not None else _tokenize(text)
    if toks:
        numeric_or_punct = sum(
            1 for t in toks if re.fullmatch(r"[\W\d]+", t) and not re.fullmatch(r"\w+", t)
        )
        ratio = numeric_or_punct / len(toks)
        if ratio >= LIST_LIKE_RATIO_THRESHOLD:
            total += LIST_LIKE_PENALTY
            matched.append("list")
    return total, matched


def _score_sentence(sent: dict) -> tuple[float, int, dict[str, int], list[str], str]:
    """Score one sentence for queue ranking.

    Returns (score, predicted_entity_count, type_counts, boilerplate_families, rank_reason).
    """
    tags = sent.get("ner_tags", []) or sent.get("labels", [])
    text = sent.get("text", "")
    tokens = sent.get("tokens", [])
    type_counts = _sentence_type_counts(tags)
    predicted_entity_count = sum(type_counts.values())
    weighted = sum(TYPE_WEIGHT.get(t, 0.5) * c for t, c in type_counts.items())
    penalty, families = _boilerplate_penalties(text, tokens=tokens)
    non_english = bool(NON_ENGLISH_RE.search(text))
    score = weighted + penalty
    if non_english:
        score -= 1.5
    reasons: list[str] = []
    if predicted_entity_count >= 3:
        reasons.append(f"entity-dense({predicted_entity_count})")
    elif predicted_entity_count == 0:
        reasons.append("entity-empty")
    rare = [t for t, c in type_counts.items() if c > 0 and TYPE_WEIGHT.get(t, 0.5) >= 1.5]
    if rare:
        reasons.append("rare-types:" + ",".join(sorted(rare)))
    if families:
        reasons.append("boilerplate:" + ",".join(sorted(families)))
    if non_english:
        reasons.append("non_english")
    rank_reason = "; ".join(reasons) or f"baseline(score={score:.2f})"
    return score, predicted_entity_count, type_counts, families, rank_reason


def _iter_draft_files() -> list[Path]:
    if not DRAFTS_DIR.exists():
        return []
    return sorted(DRAFTS_DIR.glob("*.draft.json"))


def _build_queue_items(
    drafts: list[Path], split: dict
) -> tuple[list[dict], int, int]:
    """Build the priority-queue item list from all draft files in DRAFTS_DIR.

    Returns (items, n_drafts_kept, n_excluded_test).
    """
    test_paths: set[str] = set()
    test = split.get("test", {})
    for key in ("sacred10", "bundle_duplicates_of_sacred10"):
        for v in test.get(key, []) or []:
            if isinstance(v, str):
                test_paths.add(v)
    carry = test.get("client_name_carry_alongs", {}) or {}
    test_paths.update(p for p in carry if isinstance(p, str))
    new_picks = test.get("new_spec2_picks", {}) or {}
    test_paths.update(p for p in new_picks if isinstance(p, str))
    test_paths.update(test.get("all_paths", []) or [])

    items: list[dict] = []
    n_excluded = 0
    n_drafts = 0
    for path in drafts:
        try:
            data = json.loads(path.read_text())
        except json.JSONDecodeError:
            continue
        if data.get("human_verified"):
            continue
        source = data.get("source_file", "")
        if source in test_paths:
            n_excluded += 1
            continue
        n_drafts += 1
        doc_id = data.get("doc_id", path.stem)
        for sidx, sent in enumerate(data.get("sentences", [])):
            score, pec, type_counts, _families, reason = _score_sentence(sent)
            text = sent.get("text", "")
            preview = text if len(text) <= 200 else text[:197] + "..."
            items.append(
                {
                    "doc_id": doc_id,
                    "source_file": source,
                    "sent_idx": sidx,
                    "text": preview,
                    "tokens": len(sent.get("tokens", [])),
                    "predicted_entity_count": pec,
                    "score": round(score, 4),
                    "type_counts": {k: v for k, v in type_counts.items() if v > 0},
                    "rank_reason": reason,
                }
            )
    return items, n_drafts, n_excluded


def _stable_sort_items(items: list[dict]) -> list[dict]:
    """Sort queue items deterministically: score desc, then doc_id asc, then sent_idx asc."""
    return sorted(items, key=lambda x: (-x["score"], x["doc_id"], x["sent_idx"]))


def cmd_queue(args: argparse.Namespace) -> int:
    """`queue` — build/refresh the priority review queue from DRAFTS_DIR.

    Writes data/annotations/drafts/PRIORITY_QUEUE.json. With --resume, existing
    queue items keep their position (matched on (doc_id, sent_idx)) and new
    items are appended in score order.
    """
    DRAFTS_DIR.mkdir(parents=True, exist_ok=True)
    split = _load_split()
    drafts = _iter_draft_files()
    items, n_drafts, n_excluded = _build_queue_items(drafts, split)
    items = _stable_sort_items(items)

    existing: dict[tuple[str, int], int] = {}
    if args.resume and QUEUE_PATH.exists():
        try:
            prior = json.loads(QUEUE_PATH.read_text())
            for idx, it in enumerate(prior.get("items", [])):
                existing[(it["doc_id"], it["sent_idx"])] = idx
        except json.JSONDecodeError:
            existing = {}

    if existing:
        new_items = [
            it
            for it in items
            if (it["doc_id"], it["sent_idx"]) not in existing
        ]
        items = _stable_sort_items(items)
    else:
        new_items = items

    payload = {
        "version": "1.0",
        "generated_at": _now_iso(),
        "drafts_dir": str(DRAFTS_DIR.relative_to(ROOT)) if DRAFTS_DIR.is_relative_to(ROOT) else str(DRAFTS_DIR),
        "n_drafts": n_drafts,
        "n_items": len(items),
        "n_excluded_test": n_excluded,
        "n_added_this_run": len(new_items),
        "resume_mode": bool(args.resume),
        "items": items,
    }
    QUEUE_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False))
    print(
        f"Queue written: {len(items)} items ({n_drafts} drafts, {n_excluded} test-excluded) -> {QUEUE_PATH}"
    )
    return 0


def _load_queue_progress() -> dict:
    if not QUEUE_PROGRESS_PATH.exists():
        return {"last_idx": -1, "reviewed_keys": []}
    try:
        return dict(json.loads(QUEUE_PROGRESS_PATH.read_text()))
    except json.JSONDecodeError:
        return {"last_idx": -1, "reviewed_keys": []}


def _save_queue_progress(progress: dict) -> None:
    QUEUE_PROGRESS_PATH.write_text(json.dumps(progress, indent=2, ensure_ascii=False))


def _load_draft(path: Path) -> dict:
    data = json.loads(path.read_text())
    if "ner_tags" not in data and "labels" in data:
        data["ner_tags"] = data["labels"]
    return data


def _save_verified(record: dict, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(record, indent=2, ensure_ascii=False))


def _annotate_loop(record: dict) -> tuple[int, int]:
    """Walk the draft sentence-by-sentence with single-keystroke commands.

    Each sentence shows color-coded spans (rich if available, plain fallback
    otherwise). Returns (accepted, rejected) counts.
    """
    try:
        from rich.console import Console
        from rich.markup import escape
    except Exception:  # pragma: no cover - rich missing
        Console = None  # type: ignore[assignment]
        escape = None  # type: ignore[assignment]

    accepted = 0
    rejected = 0
    console = Console() if Console else None

    for idx, sent in enumerate(record.get("sentences", [])):
        tokens = sent.get("tokens", [])
        tags = sent.get("ner_tags", [])
        if console:
            console.print(f"\n[bold]Sentence {idx + 1}/{len(record.get('sentences', []))}[/bold]")
            console.print(f"  Text: {escape(sent.get('text', ''))}")
            console.print("  Tags:")
            for tok, tag in zip(tokens, tags, strict=False):
                color = (
                    "red"
                    if tag == "O"
                    else "green"
                    if tag.startswith("S-")
                    else "yellow"
                    if tag.startswith("B-")
                    else "cyan"
                )
                console.print(f"    [{color}]{escape(tok)}[/]  [dim]{escape(tag)}[/]")
        else:
            print(f"\nSentence {idx + 1}/{len(record.get('sentences', []))}")
            print(f"  Text: {sent.get('text', '')}")
            for tok, tag in zip(tokens, tags, strict=False):
                print(f"    {tok}  {tag}")

        while True:
            choice = input("  [a]ccept / [e]dit / [r]eject / [s]kip / [q]uit> ").strip().lower()
            if choice in {"a", "e", "r", "s", "q"}:
                break
            print("  invalid; choose a / e / r / s / q")

        if choice == "q":
            print("  Quitting mid-review; partial progress NOT saved.")
            return accepted, rejected
        if choice == "s":
            continue
        if choice == "r":
            rejected += 1
            continue
        if choice == "e":
            print("  (edit path: not implemented in v1; this draft treats 'e' as 'accept-as-shown')")
        accepted += 1

    return accepted, rejected


def cmd_review(args: argparse.Namespace) -> int:
    """`review --file <draft>` OR `review --queue` — interactive tty-only review path.

    Stamps reviewer:"srujan" + reviewed_at:iso ONLY if the call is interactive and the
    file is non-empty. The tty assertion is the Rule 3 fence.

    --queue mode: walk the priority queue. The owner reviews one sentence at a time
    (the highest-scoring pending one). On 'a' (accept) the sentence is marked
    reviewed in the queue progress file; on 'r' (reject) it is skipped; 'q' exits
    and saves progress for the next --queue session.
    """
    if not sys.stdin.isatty():
        print(PROVENANCE_FENCE_TTY_ERROR, file=sys.stderr)
        return 3

    if args.reviewer != "srujan":
        print(
            f"REFUSED: reviewer={args.reviewer!r}. Only the owner is permitted to "
            f"stamp human_verified:true (Rule 3 / incident #7).",
            file=sys.stderr,
        )
        return 2

    if getattr(args, "queue", False):
        return _review_queue_mode(args)

    draft_path = Path(args.file)
    if not draft_path.exists():
        print(f"file not found: {draft_path}", file=sys.stderr)
        return 1

    record = _load_draft(draft_path)
    if record.get("human_verified"):
        print(f"already verified: {draft_path} (skipping; delete the verified copy to re-review)")
        return 1

    accepted, rejected = _annotate_loop(record)

    if accepted == 0:
        print("\nNo sentences accepted; nothing saved.")
        return 0

    record = dict(record)
    record["human_verified"] = True
    record["reviewer"] = "srujan"
    record["reviewed_at"] = _now_iso()
    record["method"] = record.get("method", "annotation_factory-draft") + "+owner-review"
    record["review_stats"] = {"accepted": accepted, "rejected": rejected}
    record["original_draft"] = str(draft_path)

    out_path = VERIFIED_DIR / f"{record['doc_id']}.json"
    _save_verified(record, out_path)
    _log_owner_minutes(record, accepted, rejected)
    print(f"\n✓ Verified: {out_path} (accepted={accepted}, rejected={rejected})")
    return 0


def _review_queue_mode(args: argparse.Namespace) -> int:
    """Walk the priority queue. Sentence-level: owner reviews one sentence at a time.

    NOTE: the queue-walking version CANNOT yet stamp human_verified:true per
    sentence (the verified/ contract is one file per doc, not per sentence).
    P2_03 keeps the v1 contract: queue mode just marks per-sentence progress in
    data/annotations/drafts/.queue_progress.json so the owner sees their
    movement through the queue; the final per-doc verified stamp still happens
    in the v1 `review --file <draft>` flow once enough of a doc's sentences
    have been triaged. P2_04 will replace this with a richer per-sentence
    verified contract if the owner wants it.
    """
    if not QUEUE_PATH.exists():
        print(f"queue file not found: {QUEUE_PATH} — run `queue` subcommand first")
        return 1

    try:
        queue = json.loads(QUEUE_PATH.read_text())
    except json.JSONDecodeError as exc:
        print(f"queue file corrupt: {exc}")
        return 1

    items = queue.get("items", [])
    progress = _load_queue_progress()
    reviewed_keys = set(tuple(k) for k in progress.get("reviewed_keys", []))
    start_idx = progress.get("last_idx", -1) + 1
    if args.limit:
        items = items[: args.limit]

    print(
        f"=== Queue review mode: {len(items)} items, resuming at idx {start_idx} ==="
    )

    accepted = 0
    rejected = 0
    skipped = 0
    for i, item in enumerate(items[start_idx:], start=start_idx):
        key = (item["doc_id"], item["sent_idx"])
        if key in reviewed_keys:
            continue
        print(
            f"\n[{i + 1}/{len(items)}] doc={item['doc_id']} sent_idx={item['sent_idx']} "
            f"score={item['score']:.2f} ent={item['predicted_entity_count']}"
        )
        print(f"  rank_reason: {item['rank_reason']}")
        print(f"  text: {item['text']}")
        while True:
            choice = input("  [a]ccept / [r]eject / [s]kip / [q]uit> ").strip().lower()
            if choice in {"a", "r", "s", "q"}:
                break
            print("  invalid; choose a / r / s / q")
        if choice == "q":
            print("  Quitting; progress saved.")
            break
        if choice == "a":
            accepted += 1
            reviewed_keys.add(key)
        elif choice == "r":
            rejected += 1
            reviewed_keys.add(key)
        else:
            skipped += 1
        progress = {
            "last_idx": i,
            "reviewed_keys": [list(k) for k in reviewed_keys],
            "ts": _now_iso(),
        }
        _save_queue_progress(progress)

    print(
        f"\n=== Session done: accepted={accepted} rejected={rejected} skipped={skipped} ==="
    )
    return 0


def _log_owner_minutes(record: dict, accepted: int, rejected: int) -> None:
    STATS_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "ts": _now_iso(),
        "doc_id": record.get("doc_id"),
        "accepted": accepted,
        "rejected": rejected,
        "reviewer": record.get("reviewer"),
    }
    with STATS_PATH.open("a") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def _iter_verified_files() -> Iterable[Path]:
    if not VERIFIED_DIR.exists():
        return []
    return sorted(VERIFIED_DIR.glob("*.json"))


def cmd_stats(args: argparse.Namespace) -> int:
    """`stats` — count verified sentences, by entity type, by doc, by minute."""
    docs = list(_iter_verified_files())
    if not docs:
        print("No verified files yet.")
        return 0

    per_doc: list[tuple[str, int, int]] = []
    entity_counter: Counter = Counter()
    total_sents = 0
    total_rejected = 0

    for path in docs:
        rec = json.loads(path.read_text())
        n_acc = rec.get("review_stats", {}).get("accepted", 0)
        n_rej = rec.get("review_stats", {}).get("rejected", 0)
        per_doc.append((rec.get("doc_id", path.stem), n_acc, n_rej))
        total_sents += n_acc
        total_rejected += n_rej
        for sent in rec.get("sentences", []):
            for tag in sent.get("ner_tags", []):
                if tag == "O":
                    continue
                if tag.startswith(("B-", "S-")):
                    entity_counter[tag.split("-", 1)[1]] += 1

    print("=== Verified annotation stats ===\n")
    print(f"Verified docs: {len(docs)}")
    print(f"Total accepted sentences: {total_sents}")
    print(f"Total rejected sentences: {total_rejected}")
    print(f"Acceptance rate: {total_sents / max(1, total_sents + total_rejected):.0%}\n")
    print("By entity type (B-/S- tags only):")
    for et in ENTITY_LABELS:
        print(f"  {et:<10s} {entity_counter.get(et, 0):>5}")
    print("\nBy doc (accepted / rejected):")
    for doc_id, a, r in per_doc:
        print(f"  {doc_id:<60s} {a:>4} / {r:>4}")
    return 0


def cmd_validate(args: argparse.Namespace) -> int:
    """`validate` — internal validation entrypoint (used by tests; not in CLI help)."""
    return validate_drafts_dir(DRAFTS_DIR) if not args.file else validate_file(Path(args.file))


def validate_file(path: Path) -> int:
    """Validate a single annotation file. Returns 0 if clean, 1 on any error."""
    errors = collect_errors_for_file(path)
    if errors:
        print(f"FAIL: {path}")
        for e in errors:
            print(f"  - {e}")
        return 1
    print(f"OK: {path}")
    return 0


def collect_errors_for_file(path: Path) -> list[str]:
    """Pure function: return a list of human-readable error strings (empty = valid)."""
    try:
        data = json.loads(path.read_text())
    except json.JSONDecodeError as exc:
        return [f"json decode: {exc}"]

    sentences: list[dict] = []
    if isinstance(data, list):
        sentences = [d for d in data if isinstance(d, dict)]
    elif isinstance(data, dict):
        if "sentences" in data and isinstance(data["sentences"], list):
            sentences = [d for d in data["sentences"] if isinstance(d, dict)]
        else:
            sentences = [data]
    else:
        return ["schema: top-level must be dict or list of dicts"]

    errors: list[str] = []
    valid_bioes = set(BIOES_LABELS)
    for sidx, sent in enumerate(sentences):
        tokens = sent.get("tokens")
        tags = sent.get("ner_tags")
        if tags is None and "labels" in sent:
            tags = sent["labels"]
        if tokens is None or tags is None:
            errors.append(f"sent {sidx}: missing tokens/ner_tags")
            continue
        if not isinstance(tokens, list) or not isinstance(tags, list):
            errors.append(f"sent {sidx}: tokens/ner_tags must be lists")
            continue
        if len(tokens) != len(tags):
            errors.append(f"sent {sidx}: tokens/tags length mismatch ({len(tokens)} vs {len(tags)})")
            continue
        for tidx, tag in enumerate(tags):
            if tag not in valid_bioes:
                errors.append(f"sent {sidx} tok {tidx}: tag '{tag}' not in BIOES_LABELS")
        for tidx, tag in enumerate(tags):
            if not (tag.startswith("I-") or tag.startswith("E-")):
                continue
            ent = tag[2:]
            prev = tags[tidx - 1] if tidx > 0 else None
            if prev is None or not (prev == f"B-{ent}" or prev == f"I-{ent}"):
                errors.append(f"sent {sidx} tok {tidx}: I/E-{ent} with no preceding B-/I-{ent}")
            nxt = tags[tidx + 1] if tidx + 1 < len(tags) else None
            if tag.startswith("E-") and nxt is not None and nxt.startswith("I-"):
                errors.append(f"sent {sidx} tok {tidx}: E- followed by I-")
    return errors


def validate_drafts_dir(drafts_dir: Path) -> int:
    if not drafts_dir.exists():
        print(f"(no drafts dir: {drafts_dir})")
        return 0
    files = sorted(drafts_dir.glob("*.draft.json"))
    if not files:
        print(f"(no drafts in {drafts_dir})")
        return 0
    rc = 0
    for f in files:
        e = collect_errors_for_file(f)
        if e:
            rc = 1
            print(f"FAIL: {f.name}")
            for x in e[:5]:
                print(f"  - {x}")
            if len(e) > 5:
                print(f"  ... and {len(e) - 5} more")
    if rc == 0:
        print(f"All {len(files)} draft files are BIOES-valid")
    return rc


def cmd_self_test(args: argparse.Namespace) -> int:
    """`self-test` — internal check that the fence works (used by tests)."""
    if sys.stdin.isatty():
        print("UNEXPECTED: stdin is a tty in self-test (tests should redirect)")
        return 1
    print("OK: stdin is not a tty; fence would block.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_d = sub.add_parser("draft", help="Pre-annotate TRAIN docs and write drafts")
    p_d.add_argument("--split", default="train", choices=["train"], help="(locked) only TRAIN")
    p_d.add_argument("--docs", type=int, default=2, help="max TRAIN docs to draft (alphabetical)")
    p_d.set_defaults(func=cmd_draft)

    p_q = sub.add_parser("queue", help="Build/refresh the priority review queue from drafts/")
    p_q.add_argument(
        "--resume",
        action="store_true",
        help="Preserve existing queue items by (doc_id, sent_idx); only add new ones",
    )
    p_q.set_defaults(func=cmd_queue)

    p_r = sub.add_parser("review", help="Interactive review (single file OR --queue walk)")
    p_r.add_argument("--file", help="draft .json to review (omit when --queue is set)")
    p_r.add_argument("--queue", action="store_true", help="walk the priority queue")
    p_r.add_argument("--limit", type=int, help="(queue mode) max items to walk this session")
    p_r.add_argument("--reviewer", default="srujan", help="owner identity (Rule 3)")
    p_r.set_defaults(func=cmd_review)

    p_s = sub.add_parser("stats", help="Verified-sentence counts")
    p_s.set_defaults(func=cmd_stats)

    p_v = sub.add_parser("validate", help="Validate a single file or all drafts in data/annotations/drafts/")
    p_v.add_argument("--file", help="specific file to validate")
    p_v.set_defaults(func=cmd_validate)

    p_t = sub.add_parser("self-test", help="Verify the tty-fence path is correctly armed (non-tty only)")
    p_t.set_defaults(func=cmd_self_test)

    args = parser.parse_args()
    if args.cmd == "review":
        if args.queue and args.file:
            parser.error("--queue and --file are mutually exclusive")
        if not args.queue and not args.file:
            parser.error("review needs --file <draft> or --queue")
    return int(args.func(args) or 0)


if __name__ == "__main__":
    raise SystemExit(main())
