#!/usr/bin/env python3
"""Per-document fidelity audit tool — proves R1 to SWA.

SWA Requirement R1: per-document fidelity audit showing count of source BOQ
rows vs output rows, with diff of misses.

Usage:
    python3 scripts/fidelity_audit.py data/real_rfqs/swa_enquiries/04_adani/
    python3 scripts/fidelity_audit.py --enquiry 04_adani
    python3 scripts/fidelity_audit.py --all
    python3 scripts/fidelity_audit.py --enquiry 04_adani --save
    python3 scripts/fidelity_audit.py --all --save

For XLSX files: source row count = rows with non-empty description AND unit in source.
For PDF files: source row count = rows in the independent gold file.
"""

from __future__ import annotations

import argparse
import json
import sys
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.pipeline import Pipeline

ROW_GOLD_DIR = Path("data/real_rfqs/gold/rows")
ENQUIRY_DIR = Path("data/real_rfqs/swa_enquiries")
RESULTS_DIR = Path("results")

LOW_CONFIDENCE_THRESHOLD = 0.5

# Same enquiry registry as eval_honest_rows.py
ENQUIRIES: dict[str, dict] = {
    "01_gsecl": {
        "source": "01_gsecl_wanakbori_tmd8/RFQ-75810 TMD-8.pdf",
        "rowgold": "01_gsecl_wanakbori_tmd8.rowgold.json",
        "type": "pdf",
    },
    "02_isro": {
        "source": "02_isro_vssc/VSSC_BOQ_with_qty.xlsx",
        "rowgold": "02_isro_vssc.rowgold.json",
        "type": "xlsx",
    },
    "03_zydus_matoda": {
        "source": "03_zydus_matoda_osd/Zydus_Matoda_Insulation_Enquiry.xlsx",
        "rowgold": "03_zydus_matoda_osd.rowgold.json",
        "type": "xlsx",
    },
    "04_adani": {
        "source": [
            "04_adani/BOQ PAGEadani proj.pdf",
            "04_adani/BOQ PAGE2adani proj.pdf",
        ],
        "rowgold": "04_adani.rowgold.json",
        "type": "pdf",
    },
    "05_zydus_animal": {
        "source": "05_zydus_animal_pharmez/Copy of Insulation Enquiry-Zydus Animal Health Expansion Project - Pharmez-Ahmedabad_.xlsx",
        "rowgold": "05_zydus_animal_pharmez.rowgold.json",
        "type": "xlsx",
    },
    "06_avante": {
        "source": "06_avante_kirloskar_pune/Insulation Boq_132.pdf",
        "rowgold": "06_avante_kirloskar_pune.rowgold.json",
        "type": "pdf",
    },
    "07_grew": {
        "source": "07_grew_solar_narmadapuram/108, BOQ compliance, Grew Energy.pdf",
        "rowgold": "07_grew_solar_narmadapuram.rowgold.json",
        "type": "pdf",
    },
    "08_sael": {
        "source": "08_sael/Copy of Insulation Enquiry - SAEL.xlsx",
        "rowgold": "08_sael.rowgold.json",
        "type": "xlsx",
    },
    "09_gem": {
        "source": "09_gem_bid_7439924/GeM-Bidding-9218026.pdf",
        "rowgold": "09_gem_bid_7439924.rowgold.json",
        "type": "pdf",
    },
    "10_gem": {
        "source": "10_gem_bid_7552777/GeM-Bidding-9343469.pdf",
        "rowgold": "10_gem_bid_7552777.rowgold.json",
        "type": "pdf",
    },
}


def _parse_qty(value: object) -> float:
    """Parse quantity to float, returning 0.0 on failure."""
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _select_boq_sheet(wb) -> openpyxl.worksheet.worksheet.Worksheet:
    """Select the primary BOQ sheet (first with >=3 non-empty rows).
    Matches logic in XLSXRowPipeline so fidelity source counts reflect
    the actual BOQ data sheet (avoids double-counting COMPLIANCE etc).
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        non_empty = [r for r in rows if any(c for c in r)]
        if len(non_empty) >= 3:
            return ws
    return wb.active


def count_xlsx_source_rows(path: Path) -> int:
    """Count rows in the BOQ sheet of an XLSX with non-empty description AND non-empty unit.
    Uses the same sheet selection as the extractor so source_row_count matches
    the rows the pipeline actually sees (fixes 33/66 on 03 Zydus Matoda).
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = _select_boq_sheet(wb)
    # Detect header row: find columns for description and unit
    desc_col: int | None = None
    unit_col: int | None = None
    header_row: int = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip().lower()
            if desc_col is None and any(
                kw in cell_str for kw in ("description", "item", "material", "particulars", "work")
            ):
                desc_col = col_idx
            if unit_col is None and cell_str in ("unit", "uom", "units"):
                unit_col = col_idx
        if desc_col is not None and unit_col is not None:
            header_row = row_idx
            break

    # Fallback: columns 1 and 2 (0-indexed) if header not found
    if desc_col is None:
        desc_col = 1
    if unit_col is None:
        unit_col = 2

    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        desc_val = row[desc_col] if desc_col < len(row) else None
        unit_val = row[unit_col] if unit_col < len(row) else None
        if desc_val and str(desc_val).strip() and unit_val and str(unit_val).strip():
            count += 1
    wb.close()
    return count


def count_pdf_source_rows(path: Path) -> int:
    """Count plausible BOQ rows in a PDF independently of the pipeline.

    Uses pdfplumber table extraction and identifies BOQ tables by their header
    cells (Description/Qty/Unit/Amount/etc.). Only data rows from those BOQ
    tables are counted. Header, total, and spec rows are skipped.
    """
    try:
        import pdfplumber
    except Exception:
        return 0

    unit_keywords = {
        "sqm",
        "sq.m",
        "sq.meter",
        "m2",
        "nos",
        "no.",
        "no",
        "kg",
        "m",
        "mm",
        "rm",
        "rmt",
        "cum",
        "cu.m",
        "mtr",
        "meter",
        "each",
        "set",
        "unit",
    }

    def _is_boq_table_header(row: list[Any]) -> bool:
        cells = [str(c).replace("\n", " ").strip().lower() for c in row if c is not None and str(c).strip()]
        if not cells:
            return False
        has_desc = any(sig in cell for cell in cells for sig in {"description", "item", "work", "material"})
        has_qty = any(sig in cell for cell in cells for sig in {"qty", "quantity", "amount"})
        has_unit = any(sig in cell for cell in cells for sig in {"unit", "uom"})
        return has_desc and (has_qty or has_unit)

    def _looks_like_data_row(row: list[Any]) -> bool:
        if not row:
            return False
        cells = [str(c).replace("\n", " ").strip() for c in row if c is not None and str(c).strip()]
        if len(cells) < 2:
            return False
        first = cells[0].lower()
        if any(
            kw in first for kw in ("total", "sr.no", "sr. no", "s.no", "description of work", "unit rate", "amount")
        ):
            return False
        has_material = any(
            len(c) > 10 and not c.replace(".", "").replace(",", "").replace("-", "").replace(" ", "").isdigit()
            for c in cells
        )
        has_unit = any(any(u in c.lower() for u in unit_keywords) for c in cells)
        has_qty = False
        for c in cells:
            clean = c.replace(",", "").replace(" ", "")
            try:
                if clean.replace(".", "", 1).isdigit() and float(c.replace(",", "")) > 0:
                    has_qty = True
                    break
            except ValueError:
                continue
        return has_material and (has_unit or has_qty)

    count = 0
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            try:
                tables = page.extract_tables() or []
            except Exception:
                continue
            for table in tables:
                if not table or len(table) < 2:
                    continue
                if not _is_boq_table_header(table[0]):
                    continue
                for row in table[1:]:
                    if _looks_like_data_row(row):
                        count += 1
    return count


def is_independent_gold(gold_data: dict) -> bool:
    """Return True if the rowgold was verified by a human.

    A human-verified rowgold is trustworthy for source-row counts even if the
    transcription tool was pdfplumber, because a person checked the entries.
    """
    return bool(gold_data.get("human_verified", False))


def load_row_gold(rowgold_path: Path) -> tuple[list[dict], dict]:
    """Load gold entries and metadata from a .rowgold.json file."""
    if not rowgold_path.exists():
        return [], {}
    with open(rowgold_path) as f:
        data = json.load(f)
    return data.get("entries", []), data


def run_pipeline_on_sources(source: str | list[str], pipeline: Pipeline) -> list:
    """Run pipeline on one or more source files, concatenating boq_items."""
    sources = [source] if isinstance(source, str) else source
    boq_items: list = []
    for src in sources:
        sp = ENQUIRY_DIR / src
        result = pipeline.run(str(sp))
        boq_items.extend(result.boq_items)
    return boq_items


def audit_enquiry(eid: str, info: dict, pipeline: Pipeline) -> dict:
    """Run fidelity audit for a single enquiry.

    Returns a dict with all audit fields; 'error' key set if something failed.
    """
    raw_source = info["source"]
    file_type = info["type"]

    # Validate source files exist
    sources = [raw_source] if isinstance(raw_source, str) else raw_source
    for src in sources:
        sp = ENQUIRY_DIR / src
        if not sp.exists():
            return {"eid": eid, "error": f"Source not found: {sp}"}

    rowgold_path = ROW_GOLD_DIR / info["rowgold"]
    gold_entries, gold_meta = load_row_gold(rowgold_path)

    # Determine source_row_count independently of the pipeline output.
    # HARD GATE (Rule 2, restored P0_03): only use len(gold_entries) as the
    # source count when the gold is independent (human-verified). Pipeline-derived
    # gold (pdfplumber transcription that has NOT been human-verified) must NEVER
    # be used as the source count because that makes the audit a self-comparison
    # — the pipeline would be graded against its own output (incident #11).
    # Non-independent gold that is present is a HARD FAILURE for this enquiry:
    # we refuse to score it rather than silently using self-compare gold or
    # silently falling back to a weaker counter. No env/flag override exists.
    if len(gold_entries) > 0:
        if not is_independent_gold(gold_meta):
            return {
                "eid": eid,
                "error": (
                    f"non-independent gold for {eid}: method="
                    f"{gold_meta.get('method', 'unknown')!r}, "
                    f"human_verified={gold_meta.get('human_verified')!r} "
                    "— cannot score self-comparison gold per Rule 2; aborting audit"
                ),
            }
        source_row_count = len(gold_entries)
    elif file_type == "xlsx":
        source_row_count = count_xlsx_source_rows(ENQUIRY_DIR / sources[0])
    elif file_type == "pdf":
        source_row_count = sum(count_pdf_source_rows(ENQUIRY_DIR / src) for src in sources)
    else:
        source_row_count = 0

    # Run pipeline
    try:
        boq_items = run_pipeline_on_sources(raw_source, pipeline)
    except Exception as exc:
        return {"eid": eid, "error": f"Pipeline failed: {exc}"}

    extracted_count = len(boq_items)

    # Classify items
    low_confidence: list[dict] = []
    normal_items: list[dict] = []

    for idx, item in enumerate(boq_items, start=1):
        conf = float(getattr(item, "confidence", 1.0))
        qty = _parse_qty(getattr(item, "quantity", None))
        unit = str(getattr(item, "unit", "") or "").strip()
        material = str(getattr(item, "material", "") or "").strip()
        warnings = list(getattr(item, "warnings", []) or [])

        flags: list[str] = []
        if conf < LOW_CONFIDENCE_THRESHOLD:
            flags.append("low confidence")
        if qty == 0.0:
            flags.append("missing quantity")
        if not unit:
            flags.append("missing unit")
        if not material:
            flags.append("missing material")

        entry = {
            "item_no": idx,
            "material": material,
            "quantity": qty,
            "unit": unit,
            "confidence": round(conf, 2),
            "flags": flags,
            "warnings": warnings,
        }

        if flags:
            low_confidence.append(entry)
        else:
            normal_items.append(entry)

    missing = max(0, source_row_count - extracted_count)
    fidelity = min(1.0, extracted_count / source_row_count) if source_row_count > 0 else 0.0

    return {
        "eid": eid,
        "file_type": file_type,
        "source_row_count": source_row_count,
        "extracted_count": extracted_count,
        "missing": missing,
        "low_confidence_count": len(low_confidence),
        "fidelity": fidelity,
        "items": normal_items,
        "flagged": low_confidence,
    }


def format_audit_report(result: dict) -> str:
    """Format a single audit result as a human-readable report string."""
    lines: list[str] = []
    eid = result["eid"]

    if "error" in result:
        lines.append(f"=== FIDELITY AUDIT: {eid} ===")
        lines.append(f"ERROR: {result['error']}")
        return "\n".join(lines)

    source_count = result["source_row_count"]
    extracted = result["extracted_count"]
    missing = result["missing"]
    low_conf = result["low_confidence_count"]
    fidelity = result["fidelity"]

    lines.append(f"=== FIDELITY AUDIT: {eid} ===")
    lines.append(f"Source BOQ rows:    {source_count}")
    lines.append(f"Extracted rows:     {extracted}")
    lines.append(f"Missing:            {missing}")
    lines.append(f"Low-confidence:     {low_conf}  (flagged for review)")
    lines.append(f"Fidelity:          {fidelity:.0%}")
    lines.append("")

    all_items = result["items"] + result["flagged"]
    all_items.sort(key=lambda x: x["item_no"])

    if all_items:
        lines.append("Extracted items:")
        for item in all_items:
            qty_str = str(item["quantity"]) if item["quantity"] != 0.0 else "?"
            mat = item["material"][:60] if item["material"] else "(no material)"
            lines.append(
                f"  {item['item_no']:>3}. {mat} | qty={qty_str} {item['unit']} | conf={item['confidence']:.2f}"
            )
        lines.append("")

    if result["flagged"]:
        lines.append("Flagged for review:")
        for item in result["flagged"]:
            qty_str = str(item["quantity"]) if item["quantity"] != 0.0 else "?"
            mat = item["material"][:60] if item["material"] else "(no material)"
            flag_str = ", ".join(item["flags"])
            lines.append(
                f"  {item['item_no']:>3}. {mat} | qty={qty_str} {item['unit']} | conf={item['confidence']:.2f}"
                f"  ← {flag_str}"
            )
        lines.append("")

    return "\n".join(lines)


def format_summary_table(results: list[dict]) -> str:
    """Format a summary table across all enquiries."""
    lines: list[str] = []
    lines.append("=" * 72)
    lines.append("FIDELITY AUDIT SUMMARY — All SWA Enquiries")
    lines.append("=" * 72)
    lines.append(f"{'Enquiry':<22} {'Src':>5} {'Ext':>5} {'Miss':>5} {'LowConf':>8} {'Fidelity':>9}")
    lines.append("-" * 72)

    valid = []
    for r in results:
        if "error" in r:
            lines.append(f"  {r['eid']:<20} ERROR: {r['error']}")
        else:
            lines.append(
                f"  {r['eid']:<20} {r['source_row_count']:>5} {r['extracted_count']:>5} "
                f"{r['missing']:>5} {r['low_confidence_count']:>8} {r['fidelity']:>8.0%}"
            )
            valid.append(r)

    if valid:
        total_src = sum(r["source_row_count"] for r in valid)
        total_ext = sum(r["extracted_count"] for r in valid)
        total_miss = sum(r["missing"] for r in valid)
        total_lc = sum(r["low_confidence_count"] for r in valid)
        overall_fidelity = min(1.0, total_ext / total_src) if total_src > 0 else 0.0
        lines.append("-" * 72)
        lines.append(
            f"  {'TOTAL':<20} {total_src:>5} {total_ext:>5} {total_miss:>5} {total_lc:>8} {overall_fidelity:>8.0%}"
        )

    lines.append("=" * 72)
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Per-document fidelity audit tool (SWA R1)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--enquiry", help="Single enquiry ID (e.g., 04_adani)")
    parser.add_argument("--all", action="store_true", help="Run against all 10 SWA enquiries")
    parser.add_argument("--save", action="store_true", help="Save output to results/fidelity_audit_{id}.txt")
    args = parser.parse_args()

    if not args.enquiry and not args.all:
        parser.print_help()
        return 1

    pipeline = Pipeline()

    if args.all:
        enquiries_to_run = list(ENQUIRIES.items())
    else:
        if args.enquiry not in ENQUIRIES:
            print(f"Unknown enquiry: {args.enquiry}")
            print(f"Available: {', '.join(ENQUIRIES.keys())}")
            return 1
        enquiries_to_run = [(args.enquiry, ENQUIRIES[args.enquiry])]

    results = []
    for eid, info in enquiries_to_run:
        print(f"Auditing {eid}...", end=" ", flush=True)
        result = audit_enquiry(eid, info, pipeline)
        results.append(result)
        if "error" in result:
            print(f"ERROR: {result['error']}")
        else:
            print(
                f"fidelity={result['fidelity']:.0%}  "
                f"({result['extracted_count']}/{result['source_row_count']} rows, "
                f"{result['low_confidence_count']} flagged)"
            )

    # Print full reports
    print()
    for result in results:
        report = format_audit_report(result)
        print(report)

        if args.save and "error" not in result:
            RESULTS_DIR.mkdir(exist_ok=True)
            out_path = RESULTS_DIR / f"fidelity_audit_{result['eid']}.txt"
            out_path.write_text(report)
            print(f"Saved: {out_path}")

    # Print summary table if running all
    if args.all:
        summary = format_summary_table(results)
        print(summary)

        if args.save:
            RESULTS_DIR.mkdir(exist_ok=True)
            summary_path = RESULTS_DIR / "fidelity_audit_summary.txt"
            summary_path.write_text(summary)
            print(f"Saved: {summary_path}")

    # HARD GATE (P0_03): any enquiry error (incl. non-independent gold refusal)
    # fails the whole run so the CLI exit code is non-zero.
    if any("error" in r for r in results):
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
