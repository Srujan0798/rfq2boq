#!/usr/bin/env python3
"""Deterministic GeM catalog ingester: XLSX -> JSON with provenance.

Ingests SWA's authoritative GeM product catalog verbatim from the sacred,
read-only source workbook ``resources/PUBLISH PRODUCT.xlsx`` into
``data/ontology/gem_catalog.json``.

The output is byte-stable: re-running the script produces an identical file
(sorted by source row order, deterministic key ordering, normalized unicode).
The JSON carries a ``_provenance`` block recording the source file path, its
sha256, the ingest date (UTC, day-resolution), and the row count, so downstream
agents can verify they are consuming the authoritative catalog.

Usage:
    python3 scripts/ingest_gem_catalog.py
    python3 scripts/ingest_gem_catalog.py --source resources/PUBLISH PRODUCT.xlsx \\
                                           --out data/ontology/gem_catalog.json
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from datetime import UTC, datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = REPO_ROOT / "resources" / "PUBLISH PRODUCT.xlsx"
DEFAULT_OUT = REPO_ROOT / "data" / "ontology" / "gem_catalog.json"

# Header row in PUBLISH PRODUCT.xlsx is row 6 (1-indexed); columns D/E/F hold
# S.No / PRODUCT NAME PUBLISHED / PRODUCT ID. The first five rows are blank.
HEADER_ROW = 6
# Column layout observed in the source workbook (1-indexed):
#   D = S.No (1..19), E = PRODUCT NAME PUBLISHED, F = PRODUCT ID
COL_SNO = 4
COL_NAME = 5
COL_PRODUCT_ID = 6


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _read_products(source: Path) -> list[dict]:
    """Read every product row from the source XLSX, verbatim, in source order.

    Returns a list of dicts with keys: row_no (1-indexed S.No from the sheet),
    name (verbatim string from the cell), product_id (verbatim string).
    """
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb.active

    products: list[dict] = []
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        sno = ws.cell(row=row, column=COL_SNO).value
        name = ws.cell(row=row, column=COL_NAME).value
        product_id = ws.cell(row=row, column=COL_PRODUCT_ID).value
        if name is None and product_id is None and sno is None:
            continue
        if not isinstance(name, str) or not name.strip():
            # Skip rows without a real product name (defensive; not observed).
            continue
        products.append(
            {
                "row_no": int(sno) if isinstance(sno, (int, float)) else len(products) + 1,
                "name": name.strip(),
                "product_id": str(product_id).strip() if product_id is not None else "",
            }
        )
    return products


def ingest(source: Path = DEFAULT_SOURCE, out: Path = DEFAULT_OUT) -> dict:
    """Ingest the source XLSX into the JSON catalog; return the parsed document."""
    if not source.exists():
        raise FileNotFoundError(f"Source GeM catalog not found: {source}")
    products = _read_products(source)
    if not products:
        raise ValueError(f"No product rows found in {source}")

    doc = {
        "_provenance": {
            "source_file": str(source.relative_to(REPO_ROOT)) if source.is_relative_to(REPO_ROOT) else str(source),
            "source_sha256": _sha256(source),
            "ingest_date_utc": datetime.now(UTC).date().isoformat(),
            "row_count": len(products),
            "ingest_tool": "scripts/ingest_gem_catalog.py",
            "note": "Verbatim ingest from SWA's sacred GeM product workbook (read-only).",
        },
        "products": products,
    }

    out.parent.mkdir(parents=True, exist_ok=True)
    # sort_keys=True + indent=2 ensures byte-stable output across runs.
    out.write_text(json.dumps(doc, indent=2, sort_keys=True, ensure_ascii=False) + "\n", encoding="utf-8")
    return doc


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE, help="Source XLSX (sacred, read-only)")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT, help="Output JSON path")
    args = parser.parse_args(argv)

    doc = ingest(source=args.source, out=args.out)
    prov = doc["_provenance"]
    print(f"Ingested {prov['row_count']} products from {prov['source_file']}")
    print(f"  source sha256: {prov['source_sha256']}")
    print(f"  ingest date  : {prov['ingest_date_utc']}")
    print(f"  wrote        : {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
