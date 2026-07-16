#!/usr/bin/env python3
"""draft_source_truth.py — machine DRAFT of true source-row counts (T1).

WHY THIS EXISTS
---------------
The fidelity harnesses (scripts/measure_fidelity.py, scripts/fidelity_audit.py)
need a stable denominator: "how many real BOQ line items exist in the SOURCE
document". That number changed run-to-run in the past (01_gsecl: 3 -> 397;
09_gem: 22 -> 157 -> 207) because different code paths derived it differently,
and at least one path (the old naive `count_pdf_source_rows` in
fidelity_audit.py) scanned tables across the ENTIRE PDF -- including
boilerplate pages (bank-guarantee lists, agreements, indemnity bonds) that
have nothing to do with the BOQ -- which is how 3 real Schedule-B rows in
01_gsecl became "397".

THIS SCRIPT NEVER RUNS THE PIPELINE. It reads the raw source file (PDF via
pdfplumber, XLSX via openpyxl) directly. Running src/pipeline.py or
src/pipeline_xlsx.py and counting THEIR output would be self-comparison --
that is incident #3 in this project's cheating history and is forbidden
(see tasks/sonnet/RULES.md).

This produces a DRAFT only. Every entry in data/real_rfqs/source_truth.json
is written with `owner_confirmed: false`. Only Srujan can flip that flag,
after reviewing results/source_truth_review.md against the actual documents.

COUNTING RULE (v1 draft -- explicitly NOT silently decided, see review.md)
---------------------------------------------------------------------------
A "BOQ line item" is a table row that:
  1. Lives on a page/sheet whose table has a detectable BOQ-style header
     (a column that says "description/item/work/material/category" AND
     either a "unit/uom" column or a "qty/quantity" column). Tables/pages
     that never show such a header (bank-guarantee lists, indemnity bonds,
     signature blocks, "Specification Document / View File" portal links)
     are NEVER counted -- this is the fix for the 01_gsecl 397 bug.
  2. Has a non-trivial description (>=3 chars) that is not a boilerplate
     row (total/subtotal/grand total, "Note:", "In words:", "Option N"
     section dividers, signature blocks, GeM "View File" links).
  3. Has a non-empty value in whichever column the table's header
     identified as the item-defining column: prefer the UNIT column if the
     table has one; else fall back to the QUANTITY column (some GeM portal
     tables list Item/Category + Quantity with no separate Unit column at
     all -- see 09_gem, 10_gem).

DELIBERATE INCLUSIVE CHOICE (flagged for owner ratification):
  Rows with a unit but a "Rate Only" / "R.O." / "R0" / zero quantity are
  COUNTED, not excluded. Rationale: they are real, distinct rows that exist
  in the source; R1 requires "every detail captured or FLAGGED, zero loss".
  If the pipeline can't produce a quantity for such a row it should FLAG it,
  not have it silently vanish from the denominator. Prior human rowgold
  transcriptions were INCONSISTENT on this point (03_zydus_matoda included
  its "R0" rows; 05_zydus_animal, 06_avante, 07_grew excluded theirs). This
  script picks ONE rule and applies it uniformly; results/source_truth_review.md
  shows the alternate ("exclude rate-only") count for every affected doc so
  the owner can pick.

Section/category header rows and "parent" rows whose only role is to
introduce child sub-items (e.g. "E CHILLED WATER PIPING", or a Sr.No "11"
row carrying nothing but a wall of spec text) do NOT count on their own --
they never carry a value in the item-defining column. This falls out of
rule 3 automatically; no separate hierarchy-detection logic is needed.

TIMEOUTS
--------
09_gem and 10_gem previously hung extractors. Every file is processed under
a hard wall-clock timeout (default 90s, SIGALRM-based, Unix only). On
timeout the doc is marked and, if an existing independent human rowgold
transcription is available, that count is cited as a fallback (with the
timeout noted in evidence) rather than silently producing 0.
"""

from __future__ import annotations

import json
import re
import signal
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import pdfplumber

REPO_ROOT = Path(__file__).resolve().parent.parent
SWA_DIR = REPO_ROOT / "data/real_rfqs/swa_enquiries"
ROWGOLD_DIR = REPO_ROOT / "data/real_rfqs/gold/rows"
OUT_JSON = REPO_ROOT / "data/real_rfqs/source_truth.json"
OUT_REVIEW = REPO_ROOT / "results/source_truth_review.md"

TIMEOUT_SECONDS = 90

# ---------------------------------------------------------------------------
# Canonical doc registry -- the 10 sacred SWA enquiries.
# doc_id matches the directory name under data/real_rfqs/swa_enquiries/ AND
# the doc_ids already used by scripts/measure_fidelity.py.
# ---------------------------------------------------------------------------
DOCS: dict[str, dict[str, Any]] = {
    "01_gsecl_wanakbori_tmd8": {
        "type": "pdf",
        "files": ["01_gsecl_wanakbori_tmd8/RFQ-75810 TMD-8.pdf"],
        "rowgold": "01_gsecl_wanakbori_tmd8.rowgold.json",
    },
    "02_isro_vssc": {
        "type": "xlsx",
        "files": ["02_isro_vssc/VSSC_BOQ_with_qty.xlsx"],
        "rowgold": "02_isro_vssc.rowgold.json",
    },
    "03_zydus_matoda_osd": {
        "type": "xlsx",
        "files": ["03_zydus_matoda_osd/Zydus_Matoda_Insulation_Enquiry.xlsx"],
        "rowgold": "03_zydus_matoda_osd.rowgold.json",
    },
    "04_adani": {
        "type": "pdf",
        "files": ["04_adani/BOQ PAGEadani proj.pdf", "04_adani/BOQ PAGE2adani proj.pdf"],
        "rowgold": "04_adani.rowgold.json",
    },
    "05_zydus_animal_pharmez": {
        "type": "xlsx",
        "files": [
            "05_zydus_animal_pharmez/Copy of Insulation Enquiry-Zydus Animal Health Expansion Project - Pharmez-Ahmedabad_.xlsx"
        ],
        "rowgold": "05_zydus_animal_pharmez.rowgold.json",
    },
    "06_avante_kirloskar_pune": {
        "type": "pdf",
        "files": ["06_avante_kirloskar_pune/Insulation Boq_132.pdf"],
        "rowgold": "06_avante_kirloskar_pune.rowgold.json",
    },
    "07_grew_solar_narmadapuram": {
        "type": "pdf",
        "files": ["07_grew_solar_narmadapuram/108, BOQ compliance, Grew Energy.pdf"],
        "rowgold": "07_grew_solar_narmadapuram.rowgold.json",
    },
    "08_sael": {
        "type": "xlsx",
        "files": ["08_sael/Copy of Insulation Enquiry - SAEL.xlsx"],
        "rowgold": "08_sael.rowgold.json",
    },
    "09_gem_bid_7439924": {
        "type": "pdf",
        "files": ["09_gem_bid_7439924/GeM-Bidding-9218026.pdf"],
        "rowgold": "09_gem_bid_7439924.rowgold.json",
    },
    "10_gem_bid_7552777": {
        "type": "pdf",
        "files": ["10_gem_bid_7552777/GeM-Bidding-9343469.pdf"],
        "rowgold": "10_gem_bid_7552777.rowgold.json",
    },
}

# Column 0 = Sr./Item No. (or blank continuation), Column 1 = Description.
# Verified by direct inspection of all 10 sacred documents (see task report).
DESC_COL = 1

UNIT_KEYWORDS = re.compile(r"\bunit\b|\buom\b|\bunits\b", re.IGNORECASE)
QTY_KEYWORDS = re.compile(r"\bqty\b|\bquantity\b", re.IGNORECASE)
DESC_KEYWORDS = re.compile(
    r"\bdescription\b|\bitem\b|\bwork\b|\bparticulars\b|\bmaterial\b|\bcategory\b",
    re.IGNORECASE,
)

BOILERPLATE_DESC = re.compile(
    r"^(total|sub-?total|grand-?total|note\b|in words|specification document|"
    r"view file|signature|option\s*\d+|declaration|^\s*$)",
    re.IGNORECASE,
)
LABEL_VALUES = {"unit", "uom", "units", "qty", "quantity", "total qty", "rate", "amount"}
RATE_ONLY_MARKERS = {"r.o.", "ro", "r0", "r/o", "rate only"}


def _clean_text(v: Any) -> str:
    """Collapse whitespace/newlines in a description cell to single spaces."""
    if v is None:
        return ""
    return " ".join(str(v).split())


def _clean_value(v: Any) -> str:
    """Strip a unit/qty cell. pdfplumber sometimes renders a single numeric
    token vertically as one-digit-per-line (e.g. GeM portal tables render
    "7150" as "7\\n1\\n5\\n0") -- join with no separator to recover it."""
    if v is None:
        return ""
    s = str(v)
    if "\n" in s:
        # if every non-empty fragment is a single char/digit, it's a vertical
        # render -> join tight; otherwise it's real multi-line text -> space-join.
        parts = [p for p in s.split("\n") if p.strip()]
        if parts and all(len(p.strip()) <= 2 for p in parts):
            return "".join(p.strip() for p in parts)
        return " ".join(p.strip() for p in parts)
    return s.strip()


def is_rate_only(value_s: str) -> bool:
    return value_s.strip().lower() in RATE_ONLY_MARKERS


@dataclass
class TableCountResult:
    count: int = 0
    excl_rate_only_count: int = 0
    evidence_lines: list[str] = field(default_factory=list)


# A real column header is a short label ("UNIT", "Qty.", "Item/Category",
# "Description"), never a full sentence. This cap is what keeps boilerplate
# prose (e.g. "...any other unit of measurement may not be adopted...") from
# being mistaken for a "Unit" column header -- that exact false positive is
# what inflated 01_gsecl from 3 rows to 397 in the old naive counter.
HEADER_CELL_MAX_LEN = 24


def classify_table_header(rows: list[tuple], max_scan: int = 8) -> tuple[int | None, int | None, int | None]:
    """Find the header row (within the first `max_scan` rows) that names a
    unit column and/or a quantity column, as a SHORT label (not a keyword
    incidentally appearing inside a long prose cell). Returns
    (header_idx, unit_col, qty_col). If nothing is found, (None, None, None)
    -- this table is not a BOQ table and must be skipped entirely."""
    for idx, row in enumerate(rows[:max_scan]):
        unit_col = None
        qty_col = None
        for i, cell in enumerate(row):
            s = _clean_text(cell)
            if not s or len(s) > HEADER_CELL_MAX_LEN:
                continue
            if unit_col is None and UNIT_KEYWORDS.search(s):
                unit_col = i
            if qty_col is None and QTY_KEYWORDS.search(s):
                qty_col = i
        # Require EITHER a short unit/qty header cell that co-occurs with a
        # short description-like header cell in the same row (the normal
        # case: "Description | Unit | Qty"), OR -- fallback for tables whose
        # item title doubles as the header row and never says the word
        # "description" (e.g. 02_isro_vssc) -- a short unit/qty cell is
        # still accepted on its own, since HEADER_CELL_MAX_LEN already rules
        # out the prose false-positives that motivated this check.
        if unit_col is not None or qty_col is not None:
            return idx, unit_col, qty_col
    return None, None, None


def count_items_in_table(
    rows: list[tuple],
    carry_state: tuple[int | None, int | None] | None,
    page_label: str,
) -> tuple[TableCountResult, tuple[int | None, int | None] | None]:
    """Count BOQ line items in one table. `carry_state` lets a BOQ table that
    spans consecutive pages without repeating its header (e.g. 09_gem's
    Talcher-project / Lara-project item tables) reuse the previously
    identified column roles -- but only for the immediately following table,
    never indefinitely (see caller)."""
    header_idx, unit_col, qty_col = classify_table_header(rows)
    if header_idx is None:
        if carry_state is None:
            return TableCountResult(), None
        header_idx = -1
        unit_col, qty_col = carry_state
    else:
        carry_state = (unit_col, qty_col)

    value_col = unit_col if unit_col is not None else qty_col
    if value_col is None:
        return TableCountResult(), carry_state

    result = TableCountResult()
    for row in rows[header_idx + 1 :]:
        if len(row) <= DESC_COL or value_col >= len(row):
            continue
        desc_s = _clean_text(row[DESC_COL])
        val_s = _clean_value(row[value_col])
        if len(desc_s) < 3 or BOILERPLATE_DESC.search(desc_s):
            continue
        if not val_s or val_s.lower() in LABEL_VALUES:
            continue
        result.count += 1
        if not is_rate_only(val_s):
            result.excl_rate_only_count += 1
        result.evidence_lines.append(f"{page_label}: {desc_s[:70]} | value={val_s}")
    return result, carry_state


class TimeoutSignalError(Exception):
    pass


def _alarm_handler(signum, frame):  # noqa: ARG001
    raise TimeoutSignalError("timed out")


def with_timeout(seconds: int):
    """Context manager: raise TimeoutSignalError if the block runs longer than
    `seconds`. SIGALRM-based (Unix only, fine on macOS/Linux CI)."""

    class _Ctx:
        def __enter__(self):
            if hasattr(signal, "SIGALRM"):
                signal.signal(signal.SIGALRM, _alarm_handler)
                signal.alarm(seconds)
            return self

        def __exit__(self, exc_type, exc, tb):
            if hasattr(signal, "SIGALRM"):
                signal.alarm(0)
            return False

    return _Ctx()


def count_pdf_source_rows(paths: list[Path]) -> tuple[TableCountResult, bool]:
    """Count BOQ line items across one or more PDF files, restricted to
    tables whose header names a description+unit or description+qty column.
    Never touches src/pipeline.py. Returns (result, header_found_anywhere) --
    see count_xlsx_source_rows docstring for why that second value matters."""
    total = TableCountResult()
    header_found_anywhere = False
    for path in paths:
        carry_state: tuple[int | None, int | None] | None = None
        pages_since_header = 0
        with pdfplumber.open(str(path)) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                try:
                    tables = page.extract_tables() or []
                except Exception:
                    continue
                page_had_header = False
                for table in tables:
                    if not table:
                        continue
                    r, new_carry = count_items_in_table(
                        table,
                        carry_state if pages_since_header <= 1 else None,
                        page_label=f"{path.name} p{page_idx + 1}",
                    )
                    if new_carry is not None:
                        carry_state = new_carry
                        page_had_header = True
                        header_found_anywhere = True
                    total.count += r.count
                    total.excl_rate_only_count += r.excl_rate_only_count
                    total.evidence_lines.extend(r.evidence_lines)
                pages_since_header = 0 if page_had_header else pages_since_header + 1
    return total, header_found_anywhere


def _select_boq_sheet(wb) -> str:
    """Pick ONE sheet to count. Some workbooks carry a duplicate/near-duplicate
    'COMPLIANCE' sheet alongside the real 'BOQ' sheet (see 03_zydus_matoda) --
    counting both would double the total (the exact 33-vs-66 bug this project
    has already hit once). Prefer a sheet literally named like 'boq'; else the
    first sheet (workbook order) whose first 8 rows show a detectable header."""
    for name in wb.sheetnames:
        if "boq" in name.lower():
            return name
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True, max_row=8))
        header_idx, unit_col, qty_col = classify_table_header(rows)
        if header_idx is not None:
            return name
    return wb.sheetnames[0]


def count_xlsx_source_rows(path: Path) -> tuple[TableCountResult, str, bool]:
    """Count BOQ line items in the (single, selected) BOQ sheet of an XLSX.

    Returns (result, sheet_name, header_found). header_found=False means
    classify_table_header() never located a BOQ-style header row at all --
    this is NOT the same as "confirmed zero items" (see draft_one_doc):
    some real BOQs (e.g. numbered-clause outlines with no literal column
    labels) have no detectable header and would otherwise silently report a
    misleading 0.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    sheet_name = _select_boq_sheet(wb)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header_idx, _, _ = classify_table_header(rows)
    result, _ = count_items_in_table(rows, carry_state=None, page_label=f"sheet '{sheet_name}'")
    return result, sheet_name, header_idx is not None


def load_existing_rowgold(doc_id: str) -> dict[str, Any] | None:
    info = DOCS[doc_id]
    path = ROWGOLD_DIR / info["rowgold"]
    if not path.exists():
        return None
    return json.loads(path.read_text())


def draft_one_doc(doc_id: str) -> dict[str, Any]:
    info = DOCS[doc_id]
    file_type = info["type"]
    paths = [SWA_DIR / f for f in info["files"]]
    missing = [p for p in paths if not p.exists()]
    if missing:
        return {
            "doc_id": doc_id,
            "source_row_count": 0,
            "unit": "BOQ line items",
            "method": "human count",
            "evidence": f"ERROR: source file(s) not found: {missing}",
            "counted_by": "machine-draft",
            "owner_confirmed": False,
            "error": "missing_source_file",
        }

    timed_out = False
    header_found = False
    try:
        with with_timeout(TIMEOUT_SECONDS):
            if file_type == "pdf":
                result, header_found = count_pdf_source_rows(paths)
                sheet_note = ""
            else:
                result, sheet_note, header_found = count_xlsx_source_rows(paths[0])
    except TimeoutSignalError:
        timed_out = True
        result = TableCountResult()
        sheet_note = ""

    if timed_out:
        gold = load_existing_rowgold(doc_id)
        if gold and gold.get("entries"):
            n = len(gold["entries"])
            return {
                "doc_id": doc_id,
                "source_row_count": n,
                "unit": "BOQ line items",
                "method": "human count",
                "evidence": (
                    f"TIMEOUT: machine extraction exceeded {TIMEOUT_SECONDS}s on "
                    f"{[str(p.name) for p in paths]}. Fell back to existing "
                    f"human-verified rowgold transcription "
                    f"(data/real_rfqs/gold/rows/{info['rowgold']}, method="
                    f"{gold.get('method')!r}, dated {gold.get('date')}), n={n}."
                ),
                "counted_by": "machine-draft (timeout fallback to prior rowgold)",
                "owner_confirmed": False,
            }
        return {
            "doc_id": doc_id,
            "source_row_count": 0,
            "unit": "BOQ line items",
            "method": "human count",
            "evidence": f"TIMEOUT after {TIMEOUT_SECONDS}s, no rowgold fallback available. NEEDS MANUAL COUNT.",
            "counted_by": "machine-draft",
            "owner_confirmed": False,
            "error": "timeout_no_fallback",
        }

    evidence_first3 = result.evidence_lines[:3]
    evidence_last3 = result.evidence_lines[-3:] if len(result.evidence_lines) > 3 else []
    sheet_bit = f"; {sheet_note}" if sheet_note else ""
    files_bit = ", ".join(p.name for p in paths)

    # A count of 0 with NO header ever detected means the auto-counter found
    # nothing it recognized as a BOQ table -- it is NOT the same claim as
    # "confirmed: this document has zero BOQ line items". Surface that
    # distinction explicitly rather than let a misleading bare "0" reach
    # the fidelity harness or the owner review doc unflagged.
    needs_manual_count = result.count == 0 and not header_found

    if needs_manual_count:
        evidence = (
            f"files=[{files_bit}]{sheet_bit}; NO BOQ-STYLE HEADER DETECTED "
            f"(description+unit or description+qty column) in any table/page. "
            f"This is NOT a confirmed zero -- it means the auto-counter's "
            f"header heuristic found nothing it recognized (e.g. numbered-"
            f"clause outlines with no literal column labels look like this). "
            f"NEEDS MANUAL COUNT."
        )
    else:
        evidence = (
            f"files=[{files_bit}]{sheet_bit}; rows counted={result.count} "
            f"(excluding rate-only/'R.O.' rows: {result.excl_rate_only_count}); "
            f"first3=[{' || '.join(evidence_first3)}]; "
            f"last3=[{' || '.join(evidence_last3)}]"
        )

    return {
        "doc_id": doc_id,
        "source_row_count": result.count,
        "unit": "BOQ line items",
        "method": "human count",
        "evidence": evidence,
        "counted_by": "machine-draft",
        "owner_confirmed": False,
        "needs_manual_count": needs_manual_count,
        "_draft_excl_rate_only_count": result.excl_rate_only_count,
        "_draft_evidence_lines": result.evidence_lines,
    }


def build_review_markdown(entries: list[dict[str, Any]]) -> str:
    lines = [
        "# Source truth review — for Srujan to confirm or correct",
        "",
        "Generated by `scripts/draft_source_truth.py`. Every count below was read",
        "directly from the source PDF/XLSX (never from pipeline output). Counting",
        "rule (v1 draft, see script docstring for full rationale):",
        "",
        "- A row counts as a BOQ line item iff it sits in a table with a detected",
        "  BOQ-style header (description + unit, or description + quantity) AND it",
        "  carries a non-empty value in that column.",
        "- Section/category headers and descriptive parent rows that only introduce",
        "  child sub-items do NOT count on their own (they never carry a unit/qty).",
        "- Rate-only ('R.O.'/'R0') and zero-quantity rows ARE counted by default",
        "  (inclusive rule) — the alternate exclusive count is shown alongside so",
        "  you can pick. Prior human transcriptions were inconsistent on this point.",
        "",
        "**Please check each row below against the actual document, then flip",
        "`owner_confirmed` in `data/real_rfqs/source_truth.json` yourself (or tell",
        "the agent which rows to correct and to what value).**",
        "",
        "| doc_id | draft count (inclusive) | draft count (excl. rate-only) | notes |",
        "|---|---|---|---|",
    ]
    for e in entries:
        n = e["source_row_count"]
        n_excl = e.get("_draft_excl_rate_only_count", n)
        note = "MISMATCH excl-rate-only differs" if n_excl != n else ""
        if e.get("error"):
            note = f"ERROR: {e['error']}"
        lines.append(f"| {e['doc_id']} | {n} | {n_excl} | {note} |")

    lines.append("")
    lines.append("## Per-document detail")
    lines.append("")
    for e in entries:
        lines.append(f"### {e['doc_id']}")
        lines.append("")
        lines.append(f"- Draft count (inclusive of rate-only rows): **{e['source_row_count']}**")
        lines.append(
            f"- Draft count (excluding rate-only rows): **{e.get('_draft_excl_rate_only_count', e['source_row_count'])}**"
        )
        lines.append(f"- Evidence: {e['evidence']}")
        gold = load_existing_rowgold(e["doc_id"])
        if gold and gold.get("entries"):
            lines.append(
                f"- Prior human rowgold transcription: {len(gold['entries'])} entries "
                f"(method={gold.get('method')!r}, human_verified={gold.get('human_verified')}) "
                "— UNTRUSTED per T2 until you re-confirm, shown here only for cross-check."
            )
        ev_lines = e.get("_draft_evidence_lines", [])
        if ev_lines:
            lines.append("")
            lines.append("First 3 rows counted:")
            for ln in ev_lines[:3]:
                lines.append(f"  - {ln}")
            if len(ev_lines) > 3:
                lines.append("Last 3 rows counted:")
                for ln in ev_lines[-3:]:
                    lines.append(f"  - {ln}")
        lines.append("")
    return "\n".join(lines)


def main() -> int:
    entries = []
    for doc_id in DOCS:
        print(f"Counting {doc_id} ...", end=" ", file=sys.stderr, flush=True)
        e = draft_one_doc(doc_id)
        print(f"{e['source_row_count']} items", file=sys.stderr)
        entries.append(e)

    # Write source_truth.json WITHOUT the internal-only draft debug fields.
    json_entries = []
    for e in entries:
        clean = {
            "doc_id": e["doc_id"],
            "source_row_count": e["source_row_count"],
            "unit": e["unit"],
            "method": e["method"],
            "evidence": e["evidence"],
            "counted_by": e["counted_by"],
            "owner_confirmed": e["owner_confirmed"],
            "needs_manual_count": e.get("needs_manual_count", False),
        }
        if e.get("error"):
            clean["error"] = e["error"]
        json_entries.append(clean)

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps({"entries": json_entries}, indent=2) + "\n")
    print(f"\nWrote {OUT_JSON}", file=sys.stderr)

    OUT_REVIEW.parent.mkdir(parents=True, exist_ok=True)
    OUT_REVIEW.write_text(build_review_markdown(entries))
    print(f"Wrote {OUT_REVIEW}", file=sys.stderr)

    print("\n" + "=" * 60)
    print("DRAFT SOURCE TRUTH SUMMARY (all owner_confirmed=false)")
    print("=" * 60)
    for e in entries:
        print(f"  {e['doc_id']:<32} {e['source_row_count']:>4} items")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
