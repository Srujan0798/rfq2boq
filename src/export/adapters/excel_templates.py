"""Generic Excel BOQ template (priced CPWD/DSR formats removed - S1 unpriced BOQ scope).

Only GenericTemplate remains for basic unpriced output.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.domain.models import ExtractionResult


class GenericTemplate:
    @staticmethod
    def generate(result: ExtractionResult, output_path: str, template_name: str = "standard") -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = f"BOQ-{template_name.upper()}"

        header_fill = PatternFill("solid", fgColor="1565C0")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        headers = [
            "Item No",
            "Description",
            "Material",
            "Grade",
            "Dimensions",
            "Location",
            "Quantity",
            "Unit",
            # Rate/Amount/DSR Code removed (unpriced BOQ)
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, item in enumerate(result.boq_items, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=getattr(item, "description_raw", "") or getattr(item, "material", ""))
            ws.cell(row=row_idx, column=3, value=getattr(item, "material", ""))
            ws.cell(row=row_idx, column=4, value=getattr(item, "grade", ""))
            dims = getattr(item, "dimensions", [])
            ws.cell(row=row_idx, column=5, value=", ".join(dims) if dims else "")
            ws.cell(row=row_idx, column=6, value=getattr(item, "location", ""))
            ws.cell(row=row_idx, column=7, value=str(getattr(item, "quantity", "")))
            ws.cell(row=row_idx, column=8, value=getattr(item, "unit", ""))

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)


# Priced export helpers removed (S1)
# def export_cpwd... removed
# def export_dsr... removed


def generate(result: ExtractionResult, output_path: str, template_name: str = "standard") -> None:
    GenericTemplate.generate(result, output_path, template_name)
