"""Create and save unpriced BOQ Excel template (rates/DSR removed per S1)."""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOQ"

title_font = Font(bold=True, size=14, color="1F4E78")
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="double"),
)

ws.cell(row=1, column=1, value="BILL OF QUANTITIES (UNPRICED - rates/DSR removed)")
ws.cell(row=1, column=1).font = title_font
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
ws.row_dimensions[1].height = 22

labels = ["Project Name:", "Location:", "Date:", "RFQ Reference:", "Contractor:"]
for i, label in enumerate(labels, 2):
    ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
    ws.row_dimensions[i].height = 16

ws.row_dimensions[7].height = 8

col_headers = ["S.No", "Description", "Unit", "Quantity", "Notes"]  # unpriced (rates/pricing removed)
for col, header in enumerate(col_headers, 1):
    cell = ws.cell(row=9, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[9].height = 28

widths = [6, 11, 42, 8, 10, 13, 14, 20]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

wb.save("src/export/templates/cpwd_template.xlsx")
print("Template saved to src/export/templates/cpwd_template.xlsx")
