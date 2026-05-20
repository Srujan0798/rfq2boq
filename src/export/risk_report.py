"""Risk report export to Excel and PDF formats."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill


def generate_risk_excel(boq: list[dict], output_path: str) -> None:
    """Generate Excel risk report with heatmap highlighting."""
    from openpyxl.styles import PatternFill  # noqa: N817

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ Risk Analysis"

    headers = ["Item No", "Material", "Grade", "Quantity", "Unit", "Risk Score", "Risk Factors", "Status"]
    ws.append(headers)

    from src.domain.risk_engine import RiskEngine
    engine = RiskEngine()

    for item in boq:
        risk = engine.score_item(item)

        if risk.score >= 30:
            status = "HIGH RISK"
        elif risk.score >= 15:
            status = "MEDIUM"
        else:
            status = "OK"

        ws.append([
            item.get("item_no", ""),
            item.get("material", ""),
            item.get("grade", ""),
            item.get("quantity", ""),
            item.get("unit", ""),
            risk.score,
            ", ".join(risk.factors) if risk.factors else "",
            status,
        ])

    for row in ws.iter_rows(min_row=2):
        risk_val = row[5].value
        if risk_val and risk_val >= 30:
            for cell in row:
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif risk_val and 15 <= risk_val < 30:
            for cell in row:
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    wb.save(output_path)


def export_risk_heatmap(report: Any, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Overview"

    ws.append(["Item #", "Material", "Total Risk", "Price Outlier", "Scope Ambiguity", "Missing Standard", "Coverage", "Flags"])

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for score in report.item_scores:
        ws.append([
            score.item_no,
            score.material,
            f"{score.total_risk_score:.0%}",
            f"{score.price_outlier_score:.0%}",
            f"{score.scope_ambiguity_score:.0%}",
            f"{score.missing_standard_score:.0%}",
            f"{score.coverage_score:.0%}",
            "; ".join(f"{f.flag_type}" for f in score.flags) if score.flags else "—",
        ])

    color_scale = ColorScaleRule(
        start_type="num", start_value=0, start_color="00B050",
        mid_type="num", mid_value=0.5, mid_color="FFFF00",
        end_type="num", end_value=1, end_color="FF0000",
    )
    ws.conditional_formatting.add("C2:C1000", color_scale)

    risk_fill_low = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    risk_fill_med = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    risk_fill_high = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        risk_cell = row[2]
        if isinstance(risk_cell.value, str) and risk_cell.value.endswith("%"):
            val = float(risk_cell.value.rstrip("%")) / 100
            if val > 0.7:
                for cell in row:
                    cell.fill = risk_fill_high
            elif val > 0.4:
                for cell in row:
                    cell.fill = risk_fill_med
            else:
                for cell in row:
                    cell.fill = risk_fill_low

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    summary = wb.create_sheet("Summary")
    summary.append(["Aggregate Risk Score", f"{report.aggregate_risk_score:.0%}"])
    summary.append(["High Risk Items", report.high_risk_items])
    summary.append(["Flagged Items", report.flagged_items])
    summary.append(["Coverage", f"{report.coverage_percentage:.0%}"])
    for warning in report.warnings:
        summary.append(["Warning", warning])

    wb.save(output_path)


def export_risk_summary_text(report: Any) -> str:
    lines = [
        "BOQ Risk Analysis Summary",
        f"{'=' * 40}",
        f"Aggregate Risk: {report.aggregate_risk_score:.0%}",
        f"High Risk Items: {report.high_risk_items}",
        f"Flagged Items: {report.flagged_items}",
        f"Coverage: {report.coverage_percentage:.0%}",
        "",
        "Top Risks:",
    ]

    sorted_scores = sorted(report.item_scores, key=lambda s: s.total_risk_score, reverse=True)
    for score in sorted_scores[:10]:
        if score.flags:
            lines.append(f"  Item {score.item_no} ({score.material}): {score.total_risk_score:.0%}")
            for flag in score.flags:
                lines.append(f"    - {flag.flag_type}: {flag.message}")

    if report.warnings:
        lines.append("")
        lines.append("Warnings:")
        for w in report.warnings:
            lines.append(f"  - {w}")

    return "\n".join(lines)
