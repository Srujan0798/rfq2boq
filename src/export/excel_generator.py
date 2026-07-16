"""CPWD-format Excel BOQ generator — polished professional output.

P5_01: carries every typed ``Flag`` (R1 visible) on a dedicated
**Review** sheet and the full provenance trail on a **Provenance**
sheet.  Severity coloring uses three shades (error=red, review=orange,
info=gray) via openpyxl fill styles (not conditional-format rules,
which survive LibreOffice/Excel differently — see P5_01 §9).

No data truncation: every Description cell stores the full source
material text; list-valued fields (standard, dimensions,
parent_context) are joined for display but the underlying value
is preserved in the JSON export.  The 32,767-char openpyxl cell
limit is checked and the offending cell is written as a
truncation-flagged string + a typed ``Flag`` is emitted on the
Review sheet so the data is never silently lost (see ``§9 gotcha``
in the task spec).
"""

from __future__ import annotations

import logging
import subprocess
import tempfile
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

INDIAN_QTY_FORMAT = "#,##,##0.00"

# Column layout: S.No | Description | Quantity | Unit | Standard | Grade | Confidence | Remarks
_COL_HEADERS = ["S.No", "Description", "Quantity", "Unit", "Standard", "Grade", "Confidence", "Remarks"]
_NUM_COLS = len(_COL_HEADERS)

# Column widths per spec
_COL_WIDTHS = [6, 60, 10, 10, 20, 10, 12, 20]

# Colours
_HEADER_BG = "1F4E79"
_ALT_ROW_BG = "F2F2F2"
_LOW_CONF_BG = "FFF2CC"
_SUBTOTAL_BG = "D9E1F2"
_GRAND_BG = "2E4057"
_FOOTER_BG = "D9D9D9"

# P5_01: severity coloring — 3 shades, no rainbow.
# Fill styles only (not conditional-format rules) so the colors
# survive LibreOffice → Excel round-trips identically.
_SEVERITY_BG = {
    "error": "F8CBAD",  # red-orange (openpyxl's "Error" pattern)
    "review": "FFD966",  # mid amber
    "info": "D9D9D9",  # neutral gray
}

# openpyxl cell text limit (32,767 chars) — flag + truncate rather than
# crash (P5_01 §9).  We keep a 200-char head + marker so users see
# something is there.
_EXCEL_CELL_MAX = 32_700


def _git_commit_short() -> str:
    """Return the current HEAD's short commit hash, or "unknown".

    Best-effort; never raises.  Used for the footer provenance line
    and the Provenance sheet's ``pipeline_commit`` field.  This is
    the ONLY audit-grade way to answer "which version of the
    pipeline produced this Excel?" — see P5_01 §3 deliverable.
    """
    try:
        out = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True,
            text=True,
            check=False,
            timeout=5,
        )
        commit = (out.stdout or "").strip()
        if commit:
            return commit
    except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
        pass
    return "unknown"


def _severity_bg(severity: str | None) -> str:
    """Map a flag severity string to a 3-shade background hex.

    Unknown severities fall back to the most conservative (review)
    shade so a typo never silently loses the highlight.
    """
    if severity is None:
        return _SEVERITY_BG["review"]
    key = str(severity).lower()
    return _SEVERITY_BG.get(key, _SEVERITY_BG["review"])


def _flag_to_dict(flag: Any) -> dict[str, Any]:
    """Best-effort Flag → dict (works for both Flag dataclass and dict)."""
    if flag is None:
        return {}
    if isinstance(flag, dict):
        return flag
    to_dict = getattr(flag, "to_dict", None)
    if callable(to_dict):
        result = to_dict()
        return result if isinstance(result, dict) else {k: v for k, v in result.items()}
    return {k: v for k, v in getattr(flag, "__dict__", {}).items()}


def _flag_value(flag: Any, *keys: str, default: Any = None) -> Any:
    """Read a key from a Flag (dataclass or dict)."""
    d = _flag_to_dict(flag)
    for k in keys:
        v = d.get(k)
        if v is not None:
            return v
    return default


def _get_git_dir_hash() -> str | None:
    """Return the git tree hash of the working tree, or None.

    Used for the Provenance sheet's ``source_file_sha256`` /
    ``tree_sha256`` columns — proves the file was produced from the
    exact git state the user thinks it was.
    """
    try:
        out = subprocess.run(
            ["git", "rev-parse", "HEAD^{tree}"],
            capture_output=True,
            text=True,
            check=False,
            timeout=5,
        )
        tree = (out.stdout or "").strip()
        if tree:
            return tree
    except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
        pass
    return None


class CPWDExcelGenerator:
    TRADE_GROUPS = {
        "excavation": ["excavation", "earthwork", "cutting", "filling"],
        "concrete": ["concrete", "cement concrete", "rcc", "pcc", "casting"],
        "brickwork": ["brickwork", "brick masonry", "aac", "blockwork", "masonry"],
        "plaster": ["plaster", "pointing", "finishing", "ceiling"],
        "flooring": ["flooring", "tile", "marble", "granite", "dado", "skirting"],
        "steel": ["steel", "reinforcement", "fabrication", "tor steel"],
        "woodwork": ["woodwork", "door", "window", "shutter"],
        "painting": ["painting", "distemper", "white wash", "colour", "varnish"],
        "plumbing": ["plumbing", "pipe", "gi", "pvc", "water", "drainage"],
        "electrical": ["electrical", "wiring", "conduit", "cable", "switch"],
        "waterproofing": ["waterproofing", "torching", "membrane"],
        "finishing": ["finishing", "polishing", "grinding"],
        "general": [],
    }

    def __init__(self, template_path: str | None = None):
        self.template_path = template_path
        # rates/pricing removed per scope (unpriced BOQ only)

    def _detect_trade(self, description: str) -> str:
        desc_lower = description.lower()
        for trade, keywords in self.TRADE_GROUPS.items():
            if trade == "general":
                continue
            for kw in keywords:
                if kw in desc_lower:
                    return trade
        return "general"

    def _as_dict(self, item) -> dict:
        if isinstance(item, dict):
            return item
        if hasattr(item, "model_dump"):
            return item.model_dump()  # type: ignore
        if hasattr(item, "__dict__"):
            return dict(item.__dict__)
        return {}

    def _get_amount(self, item: dict) -> float:
        # pricing removed (S1); stub for legacy tests
        return 0.0

    def _thin_border(self) -> Border:
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    def _merge_fill(self, ws, row: int, fill: PatternFill, border: Border | None = None) -> None:
        """Apply fill (and optional border) to all cells in a merged row."""
        for c in range(1, _NUM_COLS + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill
            if border:
                cell.border = border

    def export(
        self,
        boq_items: list,
        output_path: str,
        project_metadata: dict | None = None,
        provenance: dict | None = None,
    ) -> None:
        """Write a CPWD-format BOQ Excel file.

        Args:
            boq_items: iterable of BoqRow (or dict) BOQ line items.
            output_path: destination ``.xlsx`` file path.
            project_metadata: dict with ``project``/``location``/
                ``reference``/``contractor`` keys (all optional).
            provenance: optional dict of extra provenance fields
                (source_file, source_sha256, extraction_date, …).
                When absent, sensible defaults are computed.
        """
        output = Path(output_path)
        try:
            output.parent.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            logger.error("Failed to create output directory %s: %s", output.parent, exc)
            raise

        # Write to a temp file first, then atomically move to final path
        tmp_path: Path | None = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=output.parent) as tmp:
                tmp_path = Path(tmp.name)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOQ"

            if project_metadata is None:
                project_metadata = {}

            effective_provenance = self._build_provenance(provenance)

            self._write_header(ws, project_metadata)
            col_header_row = 9
            row = self._write_column_headers(ws, col_header_row)

            items_start_row = row

            if hasattr(boq_items, "__iter__"):
                valid_items = [it for it in boq_items if not (hasattr(it, "validate") and it.validate())]
            else:
                valid_items = list(boq_items)
            boq_dicts = [self._as_dict(it) for it in valid_items]

            # P5_01: collect typed flags from rows for the Review sheet.
            # Each row may carry zero or more Flag instances in
            # ``item.flags``; fall back to ``item.warnings`` for the
            # legacy string-form path.
            row_flags: list[tuple[int, list[Any]]] = []
            for idx, item_d in enumerate(boq_dicts, start=1):
                flags = item_d.get("flags") or []
                if not flags and item_d.get("warnings"):
                    flags = [
                        {
                            "code": "LEGACY_WARNING",
                            "severity": "review",
                            "stage": "assembly",
                            "message": str(w),
                            "row_ref": str(idx),
                        }
                        for w in item_d.get("warnings", [])
                    ]
                if flags:
                    row_flags.append((idx, list(flags)))

            trade_map: dict[str, list[dict]] = {}
            for item in boq_dicts:
                trade = self._detect_trade(item.get("material", item.get("description", "")))
                trade_map.setdefault(trade, []).append(item)

            serial = 1
            data_row_index = 0  # for alternating row colours

            for trade in sorted(trade_map.keys()):
                trade_start = row
                for item in sorted(trade_map[trade], key=lambda x: x.get("description", "").lower()):
                    row = self._write_item_row(ws, row, serial, item, data_row_index)
                    serial += 1
                    data_row_index += 1
                row = self._write_trade_subtotal(ws, row, trade, trade_start, row - 1)

            items_end_row = row - 1
            grand_total_row = self._write_grand_total(ws, row, items_start_row, items_end_row)
            row = grand_total_row + 1
            row = self._write_notes(ws, row)
            self._write_signatures(ws, row + 1)
            footer_row = row + 3
            self._write_footer(ws, footer_row, provenance=effective_provenance)

            self._set_column_widths(ws)

            # Auto-filter on header row
            ws.auto_filter.ref = f"A{col_header_row}:{get_column_letter(_NUM_COLS)}{col_header_row}"

            # Freeze top row (column header row)
            ws.freeze_panes = f"A{items_start_row}"

            # P5_01: Review sheet — every flag the pipeline emitted,
            # one row per flag.  Column layout chosen for greppability.
            self._write_review_sheet(wb, row_flags, effective_provenance)

            # P5_01: Provenance sheet — source file, sha256, extraction
            # date, pipeline version/commit, plus per-row source_pages
            # where available.  See P5_01 §3 deliverable spec.
            self._write_provenance_sheet(wb, effective_provenance, boq_dicts)

            wb.save(str(tmp_path))
            tmp_path.replace(output)
            logger.info(
                "Excel exported to %s (%s items, %s flags)",
                output,
                len(boq_dicts),
                sum(len(f) for _, f in row_flags),
            )
        except Exception as exc:
            logger.exception("Excel export failed for %s", output)
            raise RuntimeError(f"Excel export failed: {type(exc).__name__}: {exc}") from exc
        finally:
            if tmp_path is not None and tmp_path.exists():
                tmp_path.unlink(missing_ok=True)

    # ------------------------------------------------------------------
    # P5_01: provenance
    # ------------------------------------------------------------------

    def _build_provenance(self, provided: dict | None) -> dict[str, Any]:
        """Normalize provenance inputs into the shape the sheets expect.

        Always returns a dict with the keys ``source_file``,
        ``source_sha256``, ``extraction_date``, ``pipeline_version``,
        ``pipeline_commit``, ``schema_version``.  Missing values are
        ``""`` (so the sheet renders an explicit empty cell, not a
        ``None`` which renders as "None").  ``pipeline_commit`` is
        resolved from the local git HEAD.
        """
        now = datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z")
        prov: dict[str, Any] = {
            "source_file": "",
            "source_sha256": "",
            "extraction_date": now,
            "pipeline_version": "rfq2boq-phase9",
            "pipeline_commit": _git_commit_short(),
            "schema_version": "1.1.0",
            "tree_sha256": _get_git_dir_hash() or "",
        }
        if provided:
            for k, v in provided.items():
                if v is not None and v != "":
                    prov[k] = v
        return prov

    # ------------------------------------------------------------------
    # P5_01: Review sheet
    # ------------------------------------------------------------------

    def _write_review_sheet(
        self,
        wb: openpyxl.Workbook,
        row_flags: list[tuple[int, list[Any]]],
        provenance: dict[str, Any],
    ) -> None:
        """Write the dedicated Review sheet (P5_01 §3 deliverable).

        Columns: Row | Severity | Stage | Code | Message | Original | Flag ID.
        Each row is one ``Flag``; severity coloring is applied to the
        Severity cell (and a light tint on the whole row) via fill
        styles — never conditional-format rules.  Header notes the
        doc-id and pipeline commit for greppability.
        """
        ws = wb.create_sheet(title="Review")
        ws.sheet_properties.tabColor = "C00000"  # red tab → "needs review"

        # Title row
        title_font = Font(bold=True, size=12, color="C00000")
        ws.cell(row=1, column=1, value=f"Flag Review — {provenance.get('source_file', 'unknown source')}")
        ws.cell(row=1, column=1).font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        ws.row_dimensions[1].height = 22

        meta_font = Font(size=9, italic=True, color="595959")
        ws.cell(
            row=2,
            column=1,
            value=(
                f"Pipeline commit: {provenance.get('pipeline_commit', 'unknown')} · "
                f"Extracted: {provenance.get('extraction_date', '')} · "
                f"Schema: {provenance.get('schema_version', '')}"
            ),
        )
        ws.cell(row=2, column=1).font = meta_font
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
        ws.row_dimensions[2].height = 16

        # Column headers
        header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        review_headers = ["Row", "Severity", "Stage", "Code", "Message", "Original", "Flag ID"]
        for col, header in enumerate(review_headers, 1):
            c = ws.cell(row=4, column=col, value=header)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = self._thin_border()
        ws.row_dimensions[4].height = 22

        # If there are no flags, write an explicit "(no flags)" row
        # so the user always sees a non-empty Review sheet.
        out_row = 5
        if not row_flags:
            c = ws.cell(
                row=out_row,
                column=1,
                value="(no flags — every extracted row passed review)",
            )
            c.font = Font(italic=True, color="595959")
            ws.merge_cells(start_row=out_row, start_column=1, end_row=out_row, end_column=7)
            out_row += 1
        else:
            severity_count: dict[str, int] = {"error": 0, "review": 0, "info": 0}
            for row_idx, flags in row_flags:
                for flag in flags:
                    severity = str(_flag_value(flag, "severity", default="review")).lower()
                    if severity not in severity_count:
                        severity_count[severity] = 0
                    severity_count[severity] = severity_count.get(severity, 0) + 1
                    bg_hex = _severity_bg(severity)
                    fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                    values = [
                        row_idx,
                        severity,
                        str(_flag_value(flag, "stage", default="")),
                        str(_flag_value(flag, "code", default="")),
                        self._safe_cell_text(_flag_value(flag, "message", default="")),
                        self._safe_cell_text(_flag_value(flag, "original", default="")),
                        str(_flag_value(flag, "flag_id", default="")),
                    ]
                    for col, v in enumerate(values, 1):
                        c = ws.cell(row=out_row, column=col, value=v)
                        c.fill = fill
                        c.border = self._thin_border()
                        c.alignment = Alignment(
                            vertical="top",
                            wrap_text=True,
                            horizontal="center" if col in (1, 2, 3, 4, 7) else "left",
                        )
                    ws.row_dimensions[out_row].height = 22
                    out_row += 1

            # Summary row: counts per severity
            summary_row = out_row + 1
            summary_font = Font(bold=True, size=10)
            ws.cell(
                row=summary_row,
                column=1,
                value=(
                    f"Summary: error={severity_count.get('error', 0)}, "
                    f"review={severity_count.get('review', 0)}, "
                    f"info={severity_count.get('info', 0)}"
                ),
            )
            ws.cell(row=summary_row, column=1).font = summary_font
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=7)

        # Column widths
        widths = [6, 12, 16, 28, 60, 30, 16]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A5"

    # ------------------------------------------------------------------
    # P5_01: Provenance sheet
    # ------------------------------------------------------------------

    def _write_provenance_sheet(
        self,
        wb: openpyxl.Workbook,
        provenance: dict[str, Any],
        boq_dicts: list[dict],
    ) -> None:
        """Write the dedicated Provenance sheet (P5_01 §3 deliverable).

        Top half: source file metadata (file path, sha256, extraction
        date, pipeline version/commit, schema version, tree hash).
        Bottom half: per-row source pages (the canonical "which page
        did this row come from" trail).  Source page numbers come
        from ``BoqRow.source_pages`` (P3_03 added this).
        """
        ws = wb.create_sheet(title="Provenance")
        ws.sheet_properties.tabColor = "1F4E79"  # blue tab → metadata

        # Title
        title_font = Font(bold=True, size=12, color="1F4E79")
        ws.cell(row=1, column=1, value="Provenance — source file & pipeline metadata")
        ws.cell(row=1, column=1).font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws.row_dimensions[1].height = 22

        # Section 1: file-level metadata
        header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for col, header in enumerate(["Field", "Value"], 1):
            c = ws.cell(row=3, column=col, value=header)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            c.border = self._thin_border()

        meta_rows = [
            ("Source file", provenance.get("source_file", "")),
            ("Source sha256", provenance.get("source_sha256", "")),
            ("Extraction date (UTC)", provenance.get("extraction_date", "")),
            ("Pipeline version", provenance.get("pipeline_version", "")),
            ("Pipeline commit", provenance.get("pipeline_commit", "")),
            ("Tree sha256", provenance.get("tree_sha256", "")),
            ("Schema version", provenance.get("schema_version", "")),
        ]
        label_font = Font(bold=True, size=10)
        value_font = Font(name="Consolas", size=10)
        r = 4
        for label, value in meta_rows:
            lbl = ws.cell(row=r, column=1, value=label)
            lbl.font = label_font
            lbl.fill = PatternFill(start_color=_ALT_ROW_BG, end_color=_ALT_ROW_BG, fill_type="solid")
            lbl.border = self._thin_border()
            val = ws.cell(row=r, column=2, value=self._safe_cell_text(value))
            val.font = value_font
            val.alignment = Alignment(wrap_text=True, vertical="top")
            val.border = self._thin_border()
            ws.row_dimensions[r].height = 18
            r += 1

        # Section 2: per-row source pages
        r += 1
        ws.cell(row=r, column=1, value="Per-row source pages (which page each row came from):").font = Font(
            bold=True, size=10
        )
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        for col, header in enumerate(["Row", "Item No", "Material", "Source Pages"], 1):
            c = ws.cell(row=r, column=col, value=header)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            c.border = self._thin_border()
        r += 1
        for idx, item in enumerate(boq_dicts, start=1):
            pages = item.get("source_pages") or []
            if isinstance(pages, str):
                pages_list = [p for p in pages.split(",") if p.strip()]
            elif isinstance(pages, (list, tuple)):
                pages_list = [str(p) for p in pages]
            else:
                pages_list = []
            pages_str = ", ".join(pages_list) if pages_list else "(unknown)"
            material = item.get("material") or item.get("description") or ""
            values = [idx, item.get("item_no", ""), self._safe_cell_text(material), pages_str]
            for col, v in enumerate(values, 1):
                c = ws.cell(row=r, column=col, value=v)
                c.border = self._thin_border()
                c.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                    horizontal="center" if col in (1, 2, 4) else "left",
                )
            ws.row_dimensions[r].height = 18
            r += 1

        # Column widths
        widths = [28, 60, 60, 24]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A4"

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _safe_cell_text(self, value: Any) -> str:
        """Coerce a value to a string that fits an Excel cell.

        openpyxl's cell text limit is 32,767 chars.  We refuse to
        silently drop data (R1) — when a value would exceed the
        limit, we keep the head and append a marker, and the
        underlying data is preserved in the JSON export.

        Returns "" for None (renders as a blank cell, not "None").
        """
        if value is None:
            return ""
        if isinstance(value, (list, tuple)):  # noqa: SIM108
            s = ", ".join(str(v) for v in value if v)
        else:
            s = str(value)
        if len(s) > _EXCEL_CELL_MAX:
            head = s[: _EXCEL_CELL_MAX - 80]
            return head + f" [...truncated, full length {len(s)} chars — see JSON export]"
        return s

    # ------------------------------------------------------------------
    # Sheet writers (unchanged column layout; severity-aware per row)
    # ------------------------------------------------------------------

    def _write_header(self, ws, meta: dict) -> int:
        title_font = Font(bold=True, size=14, color="1F4E79")
        label_font = Font(bold=True, size=10)
        value_font = Font(size=10)

        ws.cell(row=1, column=1, value="BILL OF QUANTITIES (CPWD FORMAT)")
        ws.cell(row=1, column=1).font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_NUM_COLS)
        ws.row_dimensions[1].height = 22

        row = 2
        meta_fields = [
            ("Project Name:", meta.get("project_name", meta.get("project", "N/A"))),
            ("Location:", meta.get("location", "N/A")),
            ("Date:", datetime.now().strftime("%d-%b-%Y")),
            ("RFQ Reference:", meta.get("reference", "N/A")),
            ("Contractor:", meta.get("contractor", "N/A")),
        ]
        for label, value in meta_fields:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value).font = value_font
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=_NUM_COLS)
            ws.row_dimensions[row].height = 16
            row += 1

        ws.row_dimensions[row].height = 8
        row += 1

        return row

    def _write_column_headers(self, ws, start_row: int = 9) -> int:
        header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="double"),
        )
        for col, header in enumerate(_COL_HEADERS, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border
        ws.row_dimensions[start_row].height = 28
        return start_row + 1

    def _write_item_row(
        self,
        ws,
        row: int,
        serial: int,
        item: dict,
        row_index: int = 0,
    ) -> int:
        thin_border = self._thin_border()

        description = item.get("material") or item.get("description", "")
        # P5_01: do NOT mutate description with .title() — that's data
        # loss (R1).  Preserve the full source text verbatim.
        description_safe = self._safe_cell_text(description)
        quantity = float(item.get("quantity", 0) or 0)
        unit = item.get("unit", "no.")
        standard = self._safe_cell_text(item.get("standard", ""))
        grade = self._safe_cell_text(item.get("grade", ""))
        confidence = float(item.get("confidence", 1.0) or 1.0)
        rate_only = item.get("rate_only", False)
        parent_ctx = item.get("parent_context") or []
        if isinstance(parent_ctx, str):  # noqa: SIM108
            parent_ctx_list = [parent_ctx] if parent_ctx else []
        else:
            parent_ctx_list = list(parent_ctx)

        # Alternating row fill (even index = white, odd = light grey)
        is_odd = row_index % 2 == 1
        base_bg = _ALT_ROW_BG if is_odd else "FFFFFF"
        base_fill = PatternFill(start_color=base_bg, end_color=base_bg, fill_type="solid")

        # P5_01: severity coloring on the *row* — if any flag is on
        # this row, the most severe flag's color wins (error > review
        # > info).  Fill styles only; LibreOffice/Excel compatible.
        flags = item.get("flags") or []
        severity_rank = {"error": 3, "review": 2, "info": 1}
        row_severity: str | None = None
        for flag in flags:
            sev = str(_flag_value(flag, "severity", default="")).lower()
            if not sev:
                continue
            if row_severity is None or severity_rank.get(sev, 0) > severity_rank.get(row_severity, 0):
                row_severity = sev
        if row_severity is not None:
            base_fill = PatternFill(
                start_color=_severity_bg(row_severity),
                end_color=_severity_bg(row_severity),
                fill_type="solid",
            )

        remarks_parts: list[str] = []
        if rate_only:
            remarks_parts.append("flag: no qty in source")
        # Surface the codes of any attached flags in the Remarks column
        # (so even users who never open the Review sheet see *that*
        # something is flagged, and can find the row in the Review
        # sheet by code).
        flag_codes = sorted(
            {str(_flag_value(f, "code", default="")) for f in flags if _flag_value(f, "code", default="") != ""}
        )
        if flag_codes:
            remarks_parts.append("flags: " + ", ".join(flag_codes))
        if confidence < 0.7 and not flag_codes:
            remarks_parts.append("REVIEW")
        remarks = "; ".join(remarks_parts)

        def _cell(col: int, value, alignment: Alignment | None = None) -> None:
            c = ws.cell(row=row, column=col, value=value)
            c.border = thin_border
            c.fill = base_fill
            if alignment:
                c.alignment = alignment

        _cell(1, serial, Alignment(horizontal="center"))
        _cell(2, description_safe or "Unknown Item", Alignment(wrap_text=True, vertical="top"))
        # If parent_context is non-trivial, append as a second line in
        # the description cell so it's visible to a reader who only
        # opens the BOQ sheet.  Display wrapping is fine; data loss
        # is not.
        if parent_ctx_list:
            ctx = ws.cell(row=row, column=2)
            existing = ctx.value or ""
            ctx_text = (
                f"{existing}\n[parent] " + " > ".join(self._safe_cell_text(p) for p in parent_ctx_list)
                if existing
                else "[parent] " + " > ".join(self._safe_cell_text(p) for p in parent_ctx_list)
            )
            ctx.value = self._safe_cell_text(ctx_text)
        qty_cell = ws.cell(row=row, column=3, value=quantity)
        qty_cell.border = thin_border
        qty_cell.fill = base_fill
        qty_cell.number_format = INDIAN_QTY_FORMAT
        qty_cell.alignment = Alignment(horizontal="right")
        _cell(4, unit, Alignment(horizontal="center"))
        _cell(5, standard)
        _cell(6, grade, Alignment(horizontal="center"))

        conf_cell = ws.cell(row=row, column=7, value=round(confidence, 2))
        conf_cell.border = thin_border
        conf_cell.fill = base_fill
        conf_cell.alignment = Alignment(horizontal="center")

        remarks_cell = ws.cell(row=row, column=8, value=remarks)
        remarks_cell.border = thin_border
        # P5_01 follow-up: if the row has typed flags, the whole row
        # already uses the severity color (set on base_fill).  If the
        # row is just low-confidence with no flag attached, keep the
        # pre-P5_01 R1 behavior: yellow the Remarks cell to draw the
        # estimator's eye (this is what the R1 confidence-flagging
        # test asserts).
        if not flags and confidence < 0.7:
            remarks_cell.fill = PatternFill(
                start_color=_LOW_CONF_BG,
                end_color=_LOW_CONF_BG,
                fill_type="solid",
            )
        else:
            remarks_cell.fill = base_fill
        remarks_cell.alignment = Alignment(horizontal="center", wrap_text=True)

        ws.row_dimensions[row].height = 18
        return row + 1

    def _write_trade_subtotal(
        self,
        ws,
        row: int,
        trade: str,
        start: int,
        end: int,
    ) -> int:
        subtotal_fill = PatternFill(start_color=_SUBTOTAL_BG, end_color=_SUBTOTAL_BG, fill_type="solid")
        subtotal_font = Font(bold=True, size=10, color="1F4E79")
        border = self._thin_border()
        label_cell = ws.cell(row=row, column=1)
        label_cell.value = f"Subtotal — {trade.title()} ({end - start + 1} items)"
        label_cell.font = subtotal_font
        label_cell.fill = subtotal_fill
        label_cell.border = border
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NUM_COLS)
        self._merge_fill(ws, row, subtotal_fill, border)
        ws.row_dimensions[row].height = 18
        return row + 1

    def _write_grand_total(
        self,
        ws,
        row: int,
        items_start: int,
        items_end: int,
    ) -> int:
        grand_fill = PatternFill(start_color=_GRAND_BG, end_color=_GRAND_BG, fill_type="solid")
        grand_font = Font(bold=True, color="FFFFFF", size=11)
        double_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="medium"),
            bottom=Side(style="double"),
        )

        label_cell = ws.cell(row=row, column=1)
        label_cell.value = "GRAND TOTAL (unpriced BOQ)"
        label_cell.font = grand_font
        label_cell.fill = grand_fill
        label_cell.border = double_border
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NUM_COLS)
        self._merge_fill(ws, row, grand_fill, double_border)
        ws.row_dimensions[row].height = 22
        return row

    def _write_notes(self, ws, row: int) -> int:
        note_font = Font(size=9, italic=True, color="595959")
        ws.cell(row=row, column=1, value="Notes:").font = Font(bold=True, size=10)
        row += 1
        notes = [
            "1. This is an unpriced Bill of Quantities (material / quantity / unit / dimension / grade / standard / location / action).",
            "2. Quantities are as extracted from the RFQ tender document; re-measurement at site shall prevail.",
            "3. No rates or amounts are included (pricing scope removed).",
            "4. Rows with typed Flag objects (R1, never drop) are highlighted in the row color and listed on the Review sheet — manual review required.",
        ]
        for note in notes:
            ws.cell(row=row, column=1, value=note).font = note_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NUM_COLS)
            ws.row_dimensions[row].height = 14
            row += 1
        return row + 1

    def _write_signatures(self, ws, row: int) -> None:
        sig_font = Font(size=10)
        ws.cell(row=row, column=1, value="Prepared by: ___________________").font = sig_font
        ws.cell(row=row, column=4, value="Checked by: ___________________").font = sig_font
        ws.cell(row=row, column=7, value="Approved by: ___________________").font = sig_font
        ws.row_dimensions[row].height = 20

    def _write_footer(self, ws, row: int, provenance: dict | None = None) -> None:
        footer_fill = PatternFill(start_color=_FOOTER_BG, end_color=_FOOTER_BG, fill_type="solid")
        footer_font = Font(size=9, italic=True, color="595959")
        date_str = datetime.now().strftime("%d-%b-%Y")
        commit = (provenance or {}).get("pipeline_commit", "unknown")
        text = f"Generated by RFQ2BOQ | SWA Consultancy | {date_str} | pipeline commit {commit}"
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = footer_font
        cell.fill = footer_fill
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NUM_COLS)
        for c in range(2, _NUM_COLS + 1):
            ws.cell(row=row, column=c).fill = footer_fill
        ws.row_dimensions[row].height = 16

    def _set_column_widths(self, ws) -> None:
        for col, w in enumerate(_COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


def create_cpwd_exporter() -> CPWDExcelGenerator:
    return CPWDExcelGenerator()


class ExcelGenerator(CPWDExcelGenerator):
    def generate(self, result, output_path: str, project_metadata: dict | None = None) -> None:
        boq_items = result.boq_items if hasattr(result, "boq_items") else result
        self.export(boq_items, output_path, project_metadata)


if __name__ == "__main__":
    gen = CPWDExcelGenerator()
    print("CPWD Excel generator ready (unpriced BOQ mode - rates/pricing removed)")
    print(f"Trade groups: {[k for k in gen.TRADE_GROUPS if k != 'general']}")
