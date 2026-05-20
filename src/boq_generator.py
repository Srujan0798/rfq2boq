"""High-level BOQ output generator."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.domain.models import BoqRow, ExtractionMetadata, ExtractionResult


class BOQGenerator:
    """Format BOQ rows for JSON/Excel consumers and legacy scripts."""

    def generate(self, items_or_result: ExtractionResult | list[BoqRow] | list[dict[str, Any]]) -> dict[str, Any]:
        if isinstance(items_or_result, ExtractionResult):
            rows = list(items_or_result.boq_items)
            metadata = items_or_result.metadata
            doc_id = items_or_result.doc_id
            project_name = items_or_result.project_name
        else:
            rows = [self._coerce_row(item, index) for index, item in enumerate(items_or_result, start=1)]
            metadata = ExtractionMetadata(
                total_items=len(rows),
                avg_confidence=self._average_confidence(rows),
            )
            doc_id = "generated-boq"
            project_name = "Untitled"

        formatted_items = [self._format_item(row) for row in rows]
        return {
            "metadata": {
                "doc_id": doc_id,
                "project_name": project_name,
                "generated_at": datetime.now(UTC).isoformat(),
                "total_items": len(formatted_items),
                "avg_confidence": metadata.avg_confidence,
            },
            "boq": {
                "summary": {
                    "total_items": len(formatted_items),
                    "avg_confidence": metadata.avg_confidence,
                },
                "items": formatted_items,
            },
        }

    def generate_boq_output(self, extraction_result: ExtractionResult) -> dict[str, Any]:
        return self.generate(extraction_result)

    def generate_json(
        self,
        items_or_result: ExtractionResult | list[BoqRow] | list[dict[str, Any]],
        output_path: str,
    ) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as handle:
            json.dump(self.generate(items_or_result), handle, indent=2, default=self._json_default)
        return path

    def generate_excel(
        self,
        items_or_result: ExtractionResult | list[BoqRow] | list[dict[str, Any]],
        output_path: str,
    ) -> Path:
        data = self.generate(items_or_result)
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Bill of Quantities"
        headers = ["Item Code", "Description", "Material", "Quantity", "Unit", "Confidence"]
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        for item in data["boq"]["items"]:
            sheet.append(
                [
                    item["item_code"],
                    item["description"],
                    item["material"],
                    float(item["quantity"] or 0),
                    item["unit"],
                    item["confidence"],
                ]
            )
        workbook.save(path)
        return path

    def _format_item(self, item: BoqRow) -> dict[str, Any]:
        description = item.description_raw or item.material
        return {
            "item_code": f"BOQ-{item.item_no:03d}",
            "item_no": item.item_no,
            "description": description,
            "material": item.material,
            "quantity": item.quantity,
            "unit": item.unit,
            "action": item.action,
            "grade": item.grade,
            "standard": item.standard,
            "location": item.location,
            "confidence": item.confidence,
            "source_pages": item.source_pages,
        }

    def _coerce_row(self, item: BoqRow | dict[str, Any], index: int) -> BoqRow:
        if isinstance(item, BoqRow):
            if item.item_no <= 0:
                item.item_no = index
            return item

        return BoqRow(
            item_no=int(item.get("item_no") or index),
            material=str(item.get("material") or ""),
            quantity=Decimal(str(item.get("quantity") or 0)),
            unit=str(item.get("unit") or "no."),
            action=str(item.get("action") or "supply"),
            grade=str(item.get("grade") or ""),
            standard=list(item.get("standard") or []),
            location=str(item.get("location") or ""),
            confidence=float(item.get("confidence") or 0.0),
            description_raw=str(item.get("description") or item.get("description_raw") or item.get("material") or ""),
            source_pages=list(item.get("source_pages") or []),
        )

    def _average_confidence(self, rows: list[BoqRow]) -> float:
        if not rows:
            return 0.0
        return round(sum(row.confidence for row in rows) / len(rows), 3)

    def _json_default(self, value: Any) -> Any:
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, datetime):
            return value.isoformat()
        raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")
