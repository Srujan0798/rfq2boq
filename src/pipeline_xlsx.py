"""XLSX row-preservation pipeline — each XLSX row maps 1:1 to one BoqRow.

XLSX row-preservation pipeline.

The goal is a general, structure-driven extractor that works on any real
construction tender XLSX (insulation, civil, electrical, etc.) without
hard-coded knowledge of specific files.

Design:
- Prefer real headers when present.
- Discover all plausible quantity columns (numeric, not rate/amount/total/serial).
- For "wide matrix" tenders (multiple system qty columns, common in insulation
  enquiries), emit one BoqRow per (material row × relevant system qty col).
- Strong filters for section headers, totals, spec paragraphs, empty rows.
- No filename-specific logic or sheet-name overrides for the sacred 10.
  All behavior must be driven by header text + data patterns so it generalizes
  to new tenders the owner may provide tomorrow.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from config.settings import settings

from src.domain.flags import Flag, table_type_flag
from src.domain.models import BoqRow
from src.domain.xlsx_column_mapper import XLSXColumnMapper
from src.ingest.table_classifier import (
    TableType,
    classification_to_flag,
    classify_table,
)
from src.ingest.xlsx_hierarchy import apply_parent_context

TOTAL_KEYWORD_PATTERN = re.compile(
    r"\b(total|sub-?total|grand-?total)\b",
    re.IGNORECASE,
)


@dataclass
class _ColMapping:
    material_col: int | None = None
    quantity_cols: list[int] | None = None
    unit_col: int | None = None
    location_col: int | None = None
    grade_col: int | None = None
    dimension_col: int | None = None
    standard_col: int | None = None
    action_col: int | None = None
    total_col: int | None = None


class XLSXRowPipeline:
    DESCR_COLNAMES = ("description", "material", "item desc", "scope", "work desc")

    def __init__(self) -> None:
        self._fidelity: dict[str, int] = {
            "source_rows": 0,
            "header_rows": 0,
            "empty_rows": 0,
            "total_rows": 0,
            "spec_rows": 0,
            "section_header_rows": 0,
            "extracted_rows": 0,
            "flagged_low_confidence": 0,
            "flagged_rate_only": 0,
            "non_boq_document_flag": 0,
            "non_boq_rows": 0,
        }
        # P3_03: document-level classification of the input table.
        self._table_type: TableType = TableType.UNKNOWN
        # P3_03: typed document-level flag for non-BOQ tables (e.g.
        # "COMPLIANCE_CHECKLIST", "MAKE_LIST", "VENDOR_LIST", "GENERIC_SPEC").
        # None when the table is BOQ or the pipeline hasn't been run yet.
        self._document_flag: str | None = None
        # P3_03: per-table classification records (one per sheet's
        # classified table) -- list of (table_type, flag).
        self._table_classifications: list[tuple[TableType, str | None]] = []

    @property
    def fidelity_report(self) -> dict[str, int]:
        """Return the fidelity report from the last ``run()`` call.

        Tracks every source row disposition so callers can assert zero
        silent data loss (E2 requirement).
        """
        return dict(self._fidelity)

    @property
    def table_type(self) -> TableType:
        """The classified TableType of the primary table in the last run.

        BOQ when the table was a real bill-of-quantities; COMPLIANCE_CHECKLIST
        / MAKE_LIST / VENDOR_LIST / GENERIC_SPEC / UNKNOWN for other shapes.
        """
        return self._table_type

    @property
    def document_flag(self) -> str | None:
        """Typed document-level flag for the last run.

        None for BOQ; a stable UPPER_SNAKE_CASE string for every other type
        (e.g. "COMPLIANCE_CHECKLIST", "MAKE_LIST", "VENDOR_LIST", "GENERIC_SPEC",
        "UNKNOWN"). Multiple sheets that each classify as non-BOQ contribute
        a comma-joined flag string (e.g. "COMPLIANCE_CHECKLIST,MAKE_LIST").
        """
        flags = [f for _, f in self._table_classifications if f]
        if not flags:
            return None
        if len(flags) == 1:
            return flags[0]
        # Deduplicate while preserving order.
        seen: set[str] = set()
        deduped: list[str] = []
        for f in flags:
            if f not in seen:
                seen.add(f)
                deduped.append(f)
        return ",".join(deduped)

    @property
    def typed_flags(self) -> list[Flag]:
        """Typed document-level :class:`Flag` objects for the last run (P3_04).

        Returns one :class:`Flag` per non-BOQ table classification
        (e.g. COMPLIANCE_CHECKLIST → table_type_flag("COMPLIANCE_CHECKLIST")).
        Empty list when the primary table was a real BOQ.
        """
        flags: list[Flag] = []
        seen_codes: set[str] = set()
        for table_type, _ in self._table_classifications:
            if table_type == TableType.BOQ:
                continue
            code = table_type.value
            if code in seen_codes:
                continue
            seen_codes.add(code)
            flags.append(table_type_flag(code))
        return flags

    def run(self, xlsx_path: str | Path) -> list[BoqRow]:
        self._fidelity = {
            "source_rows": 0,
            "header_rows": 0,
            "empty_rows": 0,
            "total_rows": 0,
            "spec_rows": 0,
            "section_header_rows": 0,
            "extracted_rows": 0,
            "flagged_low_confidence": 0,
            "flagged_rate_only": 0,
            "non_boq_document_flag": 0,
            "non_boq_rows": 0,
        }
        self._table_type = TableType.UNKNOWN
        self._document_flag = None
        self._table_classifications = []

        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
        filename = Path(xlsx_path).name
        sheets = self._select_boq_sheets(wb, filename)
        if not sheets:
            return []

        # Multi-sheet workbooks (e.g. Insulation ARFF / Medical): extract every
        # BOQ-like sheet and concatenate. Single-sheet files are unchanged.
        all_rows: list[BoqRow] = []
        for sheet in sheets:
            sheet_rows = self._extract_from_sheet(sheet, filename)
            all_rows.extend(sheet_rows)

        for i, boq_row in enumerate(all_rows, start=1):
            boq_row.item_no = i
        return all_rows

    def _extract_from_sheet(
        self, sheet: openpyxl.worksheet.worksheet.Worksheet, filename: str
    ) -> list[BoqRow]:
        """Extract BOQ rows from a single worksheet (shared by multi-sheet run)."""
        raw_rows = list(sheet.iter_rows(values_only=True))
        if len(raw_rows) < 2:
            return []

        has_real_header, header_idx, data_rows = self._find_header_and_data(raw_rows, filename)
        self._fidelity["source_rows"] = self._fidelity.get("source_rows", 0) + len(data_rows)

        if has_real_header:
            headers: list[str] = [str(h).strip() if h is not None else "" for h in raw_rows[header_idx]]
            # P3_03: classify the table by header shape + sample rows.
            # If the table is a known non-BOQ type (compliance checklist,
            # make list, vendor list, generic spec), emit 0 rows + a typed
            # document-level flag. UNKNOWN is treated as BOQ-with-flag (we
            # still extract, but the flag surfaces the ambiguity for audit).
            sample_for_classify = [list(row) for row in data_rows[:10]]
            table_type = classify_table(headers, sample_for_classify)
            self._table_type = table_type
            doc_flag = classification_to_flag(table_type)
            self._table_classifications.append((table_type, doc_flag))
            if table_type not in (TableType.BOQ, TableType.UNKNOWN):
                # Non-BOQ table: stop the row-extraction path. R1-honest
                # zero rows + a typed flag, never invented BOQ rows.
                # Count every source row as a non-BOQ row so fidelity accounting
                # still balances (multi-sheet workbooks often include a
                # compliance / make / vendor sheet alongside the real BOQ).
                self._fidelity["non_boq_document_flag"] = 1
                self._fidelity["non_boq_rows"] += len(data_rows)
                return []

            mapper = XLSXColumnMapper()
            sample_rows = [list(row) for row in data_rows[:10]]
            header_col_map = mapper.map_columns(headers, sample_rows)
            quantity_cols = self._discover_quantity_columns(
                headers, data_rows[:15], mapper, header_col_map.material_col
            )
            if header_col_map.quantity_col is not None and header_col_map.quantity_col not in quantity_cols:
                quantity_cols = [header_col_map.quantity_col] + quantity_cols
            total_col = self._find_total_col(headers, data_rows[:15])
            col_map = _ColMapping(
                material_col=header_col_map.material_col,
                quantity_cols=quantity_cols,
                unit_col=header_col_map.unit_col,
                location_col=header_col_map.location_col,
                grade_col=header_col_map.grade_col,
                dimension_col=header_col_map.dimension_col,
                standard_col=header_col_map.standard_col,
                action_col=header_col_map.action_col,
                total_col=total_col,
            )
        else:
            col_map = self._map_columns_from_data(data_rows[:10])
            quantity_cols = col_map.quantity_cols or []
            total_col = None

        mapping = _ColMapping(
            material_col=col_map.material_col,
            quantity_cols=quantity_cols,
            unit_col=col_map.unit_col,
            location_col=col_map.location_col,
            grade_col=col_map.grade_col,
            dimension_col=col_map.dimension_col,
            standard_col=col_map.standard_col,
            action_col=col_map.action_col,
            total_col=total_col,
        )

        wide_matrix = len(quantity_cols) >= 3 and mapping.total_col is None
        # Multi-qty-column + TOTAL layout: a sheet that has several system /
        # area quantity columns AND a sheet-level TOTAL column. Per owner
        # D5 ruling (2026-07-06) Rule A: emit ONE row per material line,
        # qty = the sheet's TOTAL column. Rows with TOTAL=0 / R.O. are
        # rate-only requests, not billable BOQ items, and are skipped
        # (this is the 05_zydus_animal case: 8 qty cols + TOTAL, gold=20).
        # A future sheet whose qty columns carry DIFFERENT units gets
        # `mixed_unit_breakdown=True` on the parent row instead of being
        # summed (per D5 pack §4 variant).
        multi_qty_with_total = mapping.total_col is not None and len(quantity_cols) >= 2
        boq_rows: list[BoqRow] = []
        # P3_03: parallel list of source item numbers (e.g. "11.1.1") for each
        # BoqRow. Used to build parent_context after extraction. A blank string
        # means the row had no source item number (fallback / wide-matrix emit).
        source_item_nos: list[str] = []
        item_no = 1

        # Pre-filter: drop fully-empty rows and total/sub-total rows BEFORE
        # the wrap assembly. Otherwise a "Sub-Total" row (empty item number,
        # "Sub-Total" material, no unit/qty) would be mis-classified as a
        # description continuation and joined to the previous BOQ row,
        # corrupting the material text. The downstream extraction loop
        # still does its own empty/total checks for any rows that survive
        # here (defense in depth). We keep fidelity reporting accurate by
        # incrementing the appropriate counter for every row we filter.
        pre_filtered: list[tuple[Any, ...]] = []
        for row in data_rows:
            if self._is_empty_row(row):
                self._fidelity["empty_rows"] += 1
                continue
            if self._is_total_row(row):
                self._fidelity["total_rows"] += 1
                continue
            pre_filtered.append(row)
        data_rows = pre_filtered

        # Merge two-row patterns where row 1 has a short material name + item number
        # and row 2 has the long description + unit/qty (common in ISRO-style tenders).
        data_rows = self._merge_short_name_rows(data_rows, mapping)
        # P3_03: assemble multi-line / wrapped descriptions under their
        # item-numbered parent row. Only joins a row into the previous one
        # if the previous row was a real BOQ row (had an item number OR a
        # qty/unit signal). Section-header / spec-paragraph rows are
        # emitted as-is and filtered downstream by the existing
        # _is_section_header / _is_spec_paragraph checks. Notes (rows
        # whose material starts with "Note:") are NEVER treated as
        # continuations -- they are kept as standalone rows.
        pre_wrap_count = len(data_rows)
        data_rows = self._assemble_wrapped_rows(data_rows, mapping)
        # The wrap may reduce row count (continuations joined into parents).
        # Track the difference in `wrapped_continuation_rows` so the
        # "every source row accounted for" invariant still holds.
        self._fidelity["wrapped_continuation_rows"] = pre_wrap_count - len(data_rows)

        for row in data_rows:
            if self._is_empty_row(row):
                self._fidelity["empty_rows"] += 1
                continue
            if self._is_total_row(row):
                self._fidelity["total_rows"] += 1
                continue

            material = self._cell(row, mapping.material_col)
            if not material or len(material) <= 2:
                self._fidelity["header_rows"] += 1
                continue

            unit = self._cell(row, mapping.unit_col)

            qty_cols = mapping.quantity_cols or []
            any_has_qty = self._row_has_any_quantity(row, qty_cols)
            # TOTAL column is also a quantity signal (e.g. "Total qty").
            if not any_has_qty and mapping.total_col is not None:
                any_has_qty = self._row_has_any_quantity(row, [mapping.total_col])

            # Check if this row has a valid item number in first column
            has_item_number = False
            if len(row) > 0 and row[0] is not None:
                first = str(row[0]).strip()
                if self._ITEM_NUMBER_PATTERN.match(first):
                    has_item_number = True

            # Pure dimensions (e.g., "15MM", "20MM" pipe/duct sizes) are
            # usually column-header size codes in wide-matrix insulation
            # enquiries -- but NOT always: 03_zydus_matoda's real gold (hand-
            # verified 2026-06-18) has 17 legitimate per-pipe-diameter BOQ
            # rows that are bare dimension strings with real, distinct
            # non-zero quantities (e.g. "15MM"->6 rmt, "15mm OD"->30 rmt are
            # BOTH real, different rows, not duplicates/header noise). A row
            # with a genuine item number + quantity + unit is a real BOQ
            # item regardless of whether its description is just a
            # dimension -- only filter when that real-item signal is absent
            # (incident #8, reintroduced 4x on 2026-07-05; this is the
            # verified-correct, non-regressing form).
            if self._is_pure_dimension(material) and not (has_item_number and any_has_qty and unit and unit.strip()):
                self._fidelity["header_rows"] += 1
                continue

            if self._is_section_header(row, mapping, material, unit, any_has_qty, has_item_number):
                self._fidelity["section_header_rows"] += 1
                continue

            # Don't filter spec-like text if the row has a real unit + quantity
            # (long descriptions are common in real BOQs -- e.g. 02_isro's real
            # 1300-sqm insulation item has a >1000-char technical description
            # citing ASTM/IS standards, and was being silently dropped by a
            # length-based override here; fixed 2026-07-05).  Parent rows with
            # item numbers in grouped tenders (e.g. Sael) are also legitimate
            # items.  Per R1 (never silently drop): a row with a genuine
            # quantity + unit is a real BOQ item regardless of description
            # length -- length alone must never override that signal.
            if self._is_spec_paragraph(material) and not (unit and any_has_qty) and not has_item_number:
                self._fidelity["spec_rows"] += 1
                continue

            if self._has_excel_error(material):
                self._fidelity["header_rows"] += 1
                continue

            if wide_matrix:
                system_rows = self._emit_wide_matrix_rows(row, mapping, material, item_no)
                # Capture the source item number (e.g. "11.1.1") once for the
                # whole wide-matrix emit (each emitted row inherits it).
                src_no = ""
                if len(row) > 0 and row[0] is not None:
                    first = str(row[0]).strip()
                    if self._ITEM_NUMBER_PATTERN.match(first):
                        src_no = first
                for br in system_rows:
                    boq_rows.append(br)
                    source_item_nos.append(src_no)
                    self._fidelity["extracted_rows"] += 1
                    if br.confidence < 0.7:
                        self._fidelity["flagged_low_confidence"] += 1
                    if br.rate_only:
                        self._fidelity["flagged_rate_only"] += 1
                    item_no += 1
            else:
                qty = self._find_best_quantity(row, mapping)
                # When TOTAL column exists, it is the authoritative quantity source.
                # Per R1 (flag-never-drop): do not silently drop rows with TOTAL <= 0.
                # Emit them (they will be marked rate_only=True, conf=0.70, flagged in audit).
                # Previously this continue caused 20/48 fidelity on 05 Zydus Animal.
                #
                # Rule A exception (owner D5, 2026-07-06, P1_03): in the
                # multi-qty-column + TOTAL layout, a material line whose
                # TOTAL is 0 / R.O. is a rate-only request, not a billable
                # BOQ item. The owner-verified gold for 05_zydus_animal has
                # 20 rows, one per material line WITH a positive TOTAL; the
                # other 28 source rows are R.O. / 0-qty and not billable
                # items. We skip them here (counted as `total_rows` for
                # fidelity) so the output matches the gold structure.
                if mapping.total_col is not None:
                    total_val = row[mapping.total_col] if mapping.total_col < len(row) else None
                    total_qty = self._parse_qty(total_val)
                    if multi_qty_with_total and total_qty <= 0:
                        self._fidelity["total_rows"] += 1
                        continue
                    if total_qty <= 0:
                        self._fidelity["total_rows"] += 1
                        # fallthrough to emit below (no continue)

                # D5 pack §4 variant: in the multi-qty-column + TOTAL
                # layout, mark the parent row `mixed_unit_breakdown` if the
                # qty-column headers carry different units (e.g. one column
                # is m, another is m²) so the per-area sum is invalid. For
                # 05_zydus_animal all 8 qty columns share the row's unit
                # (RMT / sq.m), so this stays False. The check inspects
                # the qty-column HEADER text for a unit token at a word
                # boundary (so 'm' doesn't match inside 'manifold').
                mixed_units = False
                if multi_qty_with_total and headers:
                    unit_tokens = (
                        "rmt",
                        "rm",
                        "m",
                        "meter",
                        "metre",
                        "mtr",
                        "sqm",
                        "sq.m",
                        "sq.m.",
                        "m2",
                        "sq ft",
                        "sqft",
                        "kg",
                        "kgs",
                        "tonne",
                        "tonnes",
                        "mt",
                        "t",
                        "nos",
                        "no.",
                        "no",
                        "set",
                        "sets",
                        "litre",
                        "litres",
                        "l",
                        "lts",
                        "cum",
                        "cu.m",
                        "cu.m.",
                        "m3",
                    )
                    row_unit = (unit or "").strip().lower().rstrip(".")
                    seen_qty_header_units: set[str] = set()
                    for qc in mapping.quantity_cols or []:
                        if qc is None or qc >= len(headers):
                            continue
                        h = (headers[qc] or "").lower()
                        # word-boundary match: token is a whole word, OR
                        # token is a known short unit appearing surrounded
                        # by non-letters (parens, slashes, punctuation).
                        for tok in unit_tokens:
                            if re.search(r"(?:^|[^a-z])" + re.escape(tok) + r"(?:$|[^a-z])", h):
                                seen_qty_header_units.add(tok)
                                break
                    if row_unit:
                        seen_qty_header_units.discard(row_unit)
                    mixed_units = bool(seen_qty_header_units)

                unit_val = unit or "no."
                location = self._cell(row, mapping.location_col) or ""
                grade = self._cell(row, mapping.grade_col) or ""
                action = self._detect_action(material)
                rate_only = qty <= 0 and any(
                    self._is_rate_only_marker(c) for c in row if c is not None
                )

                # Drop non-billable parent/section intros that lack unit+qty.
                # Keep long SAEL-style rate-only group parents (item-numbered
                # multi-paragraph specs with blank qty cells — gold counts them).
                unit_is_placeholder = unit_val.strip().lower() in ("", "no.", "no", "nos.", "n/a")
                if float(qty) <= 0 and not rate_only and unit_is_placeholder:
                    mat_l = material.lower()
                    # "SECTION TITLE Supply/Providing..." — mostly-uppercase prefix
                    leading_caps = False
                    caps_prefix = re.match(
                        r"^(.*?)(?:supply|providing|sitc)",
                        material,
                        flags=re.IGNORECASE | re.DOTALL,
                    )
                    if caps_prefix:
                        prefix = caps_prefix.group(1).strip()
                        letters = [c for c in prefix if c.isalpha()]
                        if len(prefix) >= 10 and letters:
                            leading_caps = (
                                sum(1 for c in letters if c.isupper()) / len(letters)
                            ) >= 0.7
                    # Short "Providing/Supplying..." parent intros whose size
                    # children carry the real unit+qty (Insulation.xlsx).
                    short_parent_intro = len(material) < 600 and (
                        mat_l.startswith("providing") or mat_l.startswith("supplying")
                    )
                    short_title = (not has_item_number) and len(material) < 80
                    if leading_caps or short_parent_intro or short_title:
                        self._fidelity["section_header_rows"] += 1
                        continue

                new_row = BoqRow(
                    item_no=item_no,
                    material=material,
                    quantity=qty,
                    unit=unit_val,
                    description_raw=material,
                    grade=grade,
                    location=location,
                    action=action,
                    confidence=(
                        settings.XLSX_DIRECT_QUANTITY_CONFIDENCE
                        if qty > 0
                        else settings.XLSX_RECOVERED_QUANTITY_CONFIDENCE
                    ),
                    rate_only=rate_only,
                )
                if mixed_units:
                    new_row.warnings = list(new_row.warnings) + ["mixed_unit_breakdown"]
                boq_rows.append(new_row)
                # P3_03: capture source item number (e.g. "11.1.1") for
                # parent_context resolution.
                src_no = ""
                if len(row) > 0 and row[0] is not None:
                    first = str(row[0]).strip()
                    if self._ITEM_NUMBER_PATTERN.match(first):
                        src_no = first
                source_item_nos.append(src_no)
                self._fidelity["extracted_rows"] += 1
                if new_row.confidence < 0.7:
                    self._fidelity["flagged_low_confidence"] += 1
                if new_row.rate_only:
                    self._fidelity["flagged_rate_only"] += 1
                item_no += 1

        # Fallback: if main extraction completely failed, try a raw scan.
        if len(boq_rows) == 0 and len(data_rows) > 10:
            boq_rows, source_item_nos = self._fallback_raw_extraction(data_rows, mapping)

        for i, boq_row in enumerate(boq_rows, start=1):
            boq_row.item_no = i

        # P3_03: apply hierarchy parent_context (no row merging, no drops).
        # Each child row's parent_context is its ancestor description chain.
        self._apply_hierarchy(boq_rows, source_item_nos)

        return boq_rows

    def _fallback_raw_extraction(
        self, data_rows: list[tuple[Any, ...]], mapping: _ColMapping
    ) -> tuple[list[BoqRow], list[str]]:
        """Last-resort extraction: keep any row with plausible material text and a numeric column.

        Returns (boq_rows, source_item_nos). source_item_nos is the parallel
        list of source item numbers captured for parent_context resolution
        (P3_03 hierarchy feature).
        """
        boq_rows: list[BoqRow] = []
        source_item_nos: list[str] = []
        item_no = 1
        seen_materials: set[str] = set()

        for row in data_rows:
            if self._is_empty_row(row):
                continue
            if self._is_total_row(row):
                continue

            # Try each column as potential material
            best_material = ""
            best_material_col = -1
            for ci, cell in enumerate(row):
                if cell is None:
                    continue
                text = str(cell).strip()
                if len(text) > len(best_material) and len(text) > 5:
                    # Skip pure numeric cells and item numbers
                    if self._parse_qty(text) > 0 and len(text) < 20:
                        continue
                    if re.match(r"^\d+(\.\d+)*$", text):
                        continue
                    best_material = text
                    best_material_col = ci

            if not best_material or len(best_material) <= 5:
                continue
            if self._is_spec_paragraph(best_material):
                continue
            if self._is_pure_dimension(best_material):
                continue
            if best_material in seen_materials:
                continue
            seen_materials.add(best_material)

            # Find best quantity from any numeric column except the material column
            best_qty = Decimal("0")
            for ci, cell in enumerate(row):
                if ci == best_material_col:
                    continue
                qty = self._parse_qty(cell)
                if qty > best_qty:
                    best_qty = qty

            action = self._detect_action(best_material)
            boq_rows.append(
                BoqRow(
                    item_no=item_no,
                    material=best_material,
                    quantity=best_qty,
                    unit="no.",
                    description_raw=best_material,
                    action=action,
                    confidence=settings.XLSX_RECOVERED_QUANTITY_CONFIDENCE,
                )
            )
            # P3_03: capture source item number (fallback rarely sees one,
            # but record it consistently).
            src_no = ""
            if len(row) > 0 and row[0] is not None:
                first = str(row[0]).strip()
                if self._ITEM_NUMBER_PATTERN.match(first):
                    src_no = first
            source_item_nos.append(src_no)
            item_no += 1

        return boq_rows, source_item_nos

    def _find_best_quantity(self, row: tuple[Any, ...], mapping: _ColMapping) -> Decimal:
        # Prefer TOTAL column when available — it's the canonical quantity.
        # Return even if <=0 (R1: emit rate_only rows with the authoritative total).
        if mapping.total_col is not None:
            total_qty = self._parse_qty(self._cell_or_none(row, mapping.total_col))
            return total_qty
        qty_cols = mapping.quantity_cols or []
        if not qty_cols:
            return Decimal("0")
        best_qty = Decimal("0")
        for qc in qty_cols:
            qty = self._parse_qty(self._cell_or_none(row, qc))
            if qty > best_qty:
                best_qty = qty
        return best_qty if best_qty > 0 else self._parse_qty(self._cell_or_none(row, qty_cols[0]))

    def _row_has_any_quantity(self, row: tuple[Any, ...], qty_cols: list[int]) -> bool:
        for qc in qty_cols:
            if qc is not None and qc < len(row):
                val = row[qc]
                if val is not None and str(val).strip() not in ("", "None", "nan", "-"):
                    return True
        return False

    _RATE_ONLY_MARKERS = frozenset({"r0", "r.o.", "r.o", "ro", "rate only", "rate_only"})

    def _is_rate_only_marker(self, val: Any) -> bool:
        if val is None:
            return False
        s = str(val).strip().lower()
        if not s or s in ("none", "nan", "-"):
            return False
        return s in self._RATE_ONLY_MARKERS

    _SERIAL_HEADER_KEYWORDS = (
        "serial",
        "sno",
        "s.no",
        "s. no",
        "s no",
        "sno.",
        "sr.no",
        "sr no",
        "sr. no",
        "sr.no.",
        "item no",
        "item no.",
        "sl.no",
        "sl. no",
        "sl no",
        "row no",
        "row",
        "index",
        "no.",
        "no ",
    )

    def _looks_like_serial_header(self, header: str) -> bool:
        h = header.lower().strip().rstrip(".")
        if not h:
            return False
        # Normalize spaces around dots: "s. no" / "s.no" / "s . no"
        h_norm = re.sub(r"\s*\.\s*", ".", h)
        h_norm = re.sub(r"\s+", " ", h_norm).strip()
        candidates = {h, h_norm, h_norm.replace(".", " "), h_norm.replace(".", "")}
        if any(c in ("no", "sr", "sno", "sno.", "item", "row", "index", "slno", "sl no") for c in candidates):
            return True
        return any(
            kw == h or h.startswith(kw) or kw in h_norm or h_norm.startswith(kw.replace(" ", ""))
            for kw in self._SERIAL_HEADER_KEYWORDS
        )

    def _discover_quantity_columns(
        self,
        headers: list[Any],
        data_rows: list[tuple[Any, ...]],
        mapper: XLSXColumnMapper,
        material_col: int | None,
    ) -> list[int]:
        numeric_cols: list[int] = []
        for i, h in enumerate(headers):
            if i == material_col:
                continue
            h_str = str(h).lower().strip() if h is not None else ""
            # Keep "total qty" / "total quantity" as quantity columns; skip
            # financial totals (total amount/rate) and remarks/serial.
            if h_str and any(kw in h_str for kw in (" remarks", "rate", "amount", "serial", "sno")):
                continue
            if h_str and "total" in h_str and not re.search(r"\btotal\s*(qty|quantity|qnty)\b", h_str):
                continue
            if self._looks_like_serial_header(h_str):
                continue
            samples = [row[i] if i < len(row) else None for row in data_rows]
            non_null = [s for s in samples if s is not None]
            if not non_null:
                continue
            # Skip serial/item-number columns by HEADER only — do not use the
            # item-number regex on cell values (decimal quantities like 1.5 /
            # 11.1 look identical to hierarchical item numbers and would wipe
            # legitimate multi-area qty columns, e.g. 05_zydus).
            if i == 0 or self._looks_like_serial_header(h_str):
                continue
            numeric_count = sum(1 for s in non_null if mapper._is_numeric(s))
            rate_only_count = sum(1 for s in non_null if self._is_rate_only_marker(s))
            other_count = len(non_null) - numeric_count - rate_only_count
            if other_count > 0:
                continue
            if numeric_count >= 1 and (numeric_count + rate_only_count) >= max(3, len(non_null) * 0.5):
                # Reject columns that look like item/serial numbers (small sequential ints)
                numeric_vals = [s for s in non_null if mapper._is_numeric(s)]
                if len(numeric_vals) >= 2:
                    try:
                        ints = [int(float(str(s).replace(",", ""))) for s in numeric_vals]
                        if all(1 <= v <= 50 for v in ints):
                            # Check if sequential or nearly sequential
                            sorted_vals = sorted(set(ints))
                            if len(sorted_vals) >= 2 and max(sorted_vals) - min(sorted_vals) <= len(sorted_vals) + 2:
                                continue
                    except (ValueError, TypeError):
                        pass
                numeric_cols.append(i)
        return numeric_cols

    def _find_total_col(self, headers: list[Any], data_rows: list[tuple[Any, ...]]) -> int | None:
        for i, h in enumerate(headers):
            h_str = str(h).lower().strip() if h is not None else ""
            if TOTAL_KEYWORD_PATTERN.search(h_str):
                return i
        return None

    def _emit_wide_matrix_rows(
        self,
        row: tuple[Any, ...],
        mapping: _ColMapping,
        material: str,
        start_item_no: int,
    ) -> list[BoqRow]:
        results: list[BoqRow] = []
        unit = self._cell(row, mapping.unit_col) or "no."
        location = self._cell(row, mapping.location_col) or ""
        grade = self._cell(row, mapping.grade_col) or ""
        action = self._detect_action(material)

        # Always prefer TOTAL column for wide matrix — it prevents duplicate items.
        if mapping.total_col is not None:
            total_qty = self._parse_qty(self._cell_or_none(row, mapping.total_col))
            if total_qty >= 0:
                return [
                    BoqRow(
                        item_no=start_item_no,
                        material=material,
                        quantity=total_qty,
                        unit=unit,
                        description_raw=material,
                        grade=grade,
                        location=location,
                        action=action,
                        confidence=settings.XLSX_DIRECT_QUANTITY_CONFIDENCE,
                    )
                ]

        for qc in mapping.quantity_cols or []:
            qty = self._parse_qty(self._cell_or_none(row, qc))
            if qty > 0:
                results.append(
                    BoqRow(
                        item_no=start_item_no,
                        material=material,
                        quantity=qty,
                        unit=unit,
                        description_raw=material,
                        grade=grade,
                        location=location,
                        action=action,
                        confidence=settings.XLSX_DIRECT_QUANTITY_CONFIDENCE,
                    )
                )
                start_item_no += 1

        if results:
            return results

        qty = self._parse_qty(self._cell_or_none(row, mapping.quantity_cols[0]) if mapping.quantity_cols else None)
        results.append(
            BoqRow(
                item_no=start_item_no,
                material=material,
                quantity=qty,
                unit=unit,
                description_raw=material,
                grade=grade,
                location=location,
                action=action,
                confidence=settings.XLSX_RECOVERED_QUANTITY_CONFIDENCE,
            )
        )
        return results

    def _find_header_and_data(
        self, raw_rows: list[tuple[Any, ...]], filename: str
    ) -> tuple[bool, int, list[tuple[Any, ...]]]:
        # Fully general: no filename-specific scanners.
        # Try standard header detection first.
        header_idx = self._find_header_row(raw_rows)
        data_rows = raw_rows[header_idx + 1 :]
        return True, header_idx, data_rows

    _ITEM_NUMBER_PATTERN = re.compile(r"^\d+(\.\d+)*$")

    def _find_header_row(self, raw_rows: list[tuple[Any, ...]]) -> int:
        for i, row in enumerate(raw_rows):
            non_empty = [c for c in row if c is not None and str(c).strip() not in ("", "None", "nan")]
            if len(non_empty) < 2:
                continue
            text_vals = [str(c).lower().strip() for c in non_empty]
            header_indicators = (
                "description",
                "material",
                "qty",
                "quantity",
                "unit",
                "remarks",
                "sr. no.",
                "s r no",
            )
            has_header_text = any(any(ind in t for ind in header_indicators) for t in text_vals)
            if not has_header_text:
                continue
            # Reject rows that look like data rows (item numbers + quantities)
            first_cell = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            looks_like_item_number = bool(self._ITEM_NUMBER_PATTERN.match(first_cell))
            has_quantity_in_row = any(
                self._parse_qty(c) > 0 for c in row[1:] if c is not None and str(c).strip() not in ("", "None", "nan")
            )
            if looks_like_item_number and has_quantity_in_row:
                continue
            # Reject rows that are overwhelmingly long description text
            total_text_len = sum(len(str(c)) for c in non_empty)
            if total_text_len > 800 and not any(t in text_vals for t in ("sr. no.", "s.no", "item no")):
                continue
            return i
        return 0

    @staticmethod
    def _score_data_rows(rows: list[tuple]) -> int:
        """Count rows that contain at least one numeric cell — a proxy for BOQ data content."""
        count = 0
        for row in rows:
            if any(isinstance(c, (int, float)) for c in row if c is not None):
                count += 1
        return count

    def _select_boq_sheet(self, wb: openpyxl.Workbook, filename: str) -> openpyxl.worksheet.worksheet.Worksheet:
        """Back-compat: return the single highest-scoring BOQ sheet."""
        sheets = self._select_boq_sheets(wb, filename)
        return sheets[0] if sheets else wb.active

    def _select_boq_sheets(
        self, wb: openpyxl.Workbook, filename: str
    ) -> list[openpyxl.worksheet.worksheet.Worksheet]:
        """Return all BOQ-like sheets, best-first.

        Multi-sheet insulation packs (ARFF, Medical) put billable lines on
        Sheet1 *and* Sheet2. Prefer every sheet with a meaningful numeric
        data score; fall back to the single best sheet when only one qualifies.
        """
        candidates: list[tuple[int, str, openpyxl.worksheet.worksheet.Worksheet]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            non_empty = [r for r in rows if any(c for c in r)]
            if len(non_empty) >= 3:
                data_score = self._score_data_rows(non_empty)
                if data_score > 0:
                    candidates.append((data_score, sheet_name, ws))
        if not candidates:
            return [wb.active]
        candidates.sort(key=lambda x: (-x[0], x[1]))
        # Keep all sheets that look like real BOQ tables (score >= 2), else top-1.
        multi = [ws for score, _name, ws in candidates if score >= 2]
        if len(multi) >= 2:
            # Preserve workbook order for stable multi-sheet concatenation.
            order = {name: i for i, name in enumerate(wb.sheetnames)}
            multi.sort(key=lambda ws: order.get(ws.title, 999))
            return multi
        return [candidates[0][2]]

    def _is_empty_row(self, row: tuple[Any, ...]) -> bool:
        return all(c is None or str(c).strip() in ("", "None", "nan") for c in row)

    def _is_total_row(self, row: tuple[Any, ...]) -> bool:
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            if TOTAL_KEYWORD_PATTERN.search(text):
                return True
        return False

    _NOTE_RE = re.compile(r"^Note[\s:]", re.IGNORECASE)

    # Descriptive header text that should be filtered even if it contains dimensions.
    _HEADER_STARTS = (
        "drain piping",
        "manifold will be",
        "colour shall be",
        "supply, installation, testing and commissioning",
        "supply, installation, testing & commissioning",
        "supply and installation of thermal insulation",
        "optional",
        "structure & civil",
        "thermal insulation",
    )

    # Material nouns — if a cell has no material noun, it's likely a dimension header, not a material.
    _MATERIAL_NOUNS = frozenset(
        {
            "pipe",
            "pipes",
            "duct",
            "ducts",
            "insulation",
            "lining",
            "laying",
            "cement",
            "concrete",
            "steel",
            "rebar",
            "rod",
            "rods",
            "bar",
            "bars",
            "aggregate",
            "sand",
            "mortar",
            "grout",
            "screed",
            "plaster",
            "paint",
            "painting",
            "primer",
            "coating",
            "waterproofing",
            "tile",
            "tiles",
            "flooring",
            "granite",
            "marble",
            "glass",
            "plywood",
            "board",
            "panel",
            "panels",
            "cladding",
            "sheet",
            "sheets",
            "wool",
            "mattress",
            "mattresses",
            "cable",
            "cables",
            "wire",
            "wires",
            "conductor",
            "conduit",
            "junction",
            "box",
            "aluminum",
            "brick",
            "bricks",
            "block",
            "blocks",
            "paver",
            "pavers",
            "adhesive",
            "sealant",
            "tape",
            "fastener",
            "bolt",
            "nut",
            "washer",
            "valve",
            "fitting",
            "elbow",
            "tee",
            "flange",
            "gasket",
        }
    )

    # Pure-dimension pattern: matches cells like "15MM", "20 mm" (number + unit ONLY, no qualifier).
    # Does NOT match "15mm OD" or "15mm thick" — those have qualifiers that give material context.
    _PURE_DIMENSION_RE = re.compile(
        r"^\s*\d+\s*(mm|cm|m|inch|in|ft)\s*$",
        re.IGNORECASE,
    )

    # Dimension with qualifier: "15mm OD", "15mm ID", "15mm thick", "20mm dia" — VALID (has material context).
    _DIMENSION_WITH_QUALIFIER_RE = re.compile(
        r"^\s*\d+\s*(mm|cm|m|inch|in|ft)\s*(od|id|thick|thick\.|dia|diameter)\s*$",
        re.IGNORECASE,
    )

    def _is_pure_dimension(self, material: str) -> bool:
        """Check if material is purely a dimension value with no material context.

        Examples of pure dimensions (REJECT):
        - "15MM"
        - "20 mm"
        - "32mm"
        - "50MM"

        Examples of valid materials (KEEP):
        - "15mm OD" (has qualifier "OD" — outer diameter, gives material context)
        - "15mm ID" (has qualifier "ID" — inner diameter)
        - "15mm thick" (has qualifier "thick")
        - "15mm dia" (has qualifier "dia")
        - "15mm thick insulation" (has material noun "insulation")
        - "20mm dia pipes" (has material noun "pipes")
        - "13 mm thick insulation for supply air ducts" (has material noun "insulation")
        """
        if not material:
            return False
        mat_stripped = material.strip()
        if not mat_stripped:
            return False
        # If it has a dimension qualifier (OD, ID, thick, dia), it's valid (has material context).
        if self._DIMENSION_WITH_QUALIFIER_RE.match(mat_stripped):
            return False
        # If it has a material noun, it's valid.
        mat_lower = mat_stripped.lower()
        if any(noun in mat_lower.split() for noun in self._MATERIAL_NOUNS):
            return False
        # Must match pure-dimension pattern (number + unit, nothing else).
        return bool(self._PURE_DIMENSION_RE.match(mat_stripped))

    def _is_section_header(
        self,
        row: tuple[Any, ...],
        mapping: _ColMapping,
        material: str,
        unit: str,
        any_has_qty: bool,
        has_item_number: bool = False,
    ) -> bool:
        if not material or material in ("", "None", "nan"):
            return True

        # Notes are never BOQ items
        if self._NOTE_RE.search(material):
            return True

        # If row has a valid unit, it's likely a real BOQ item.
        if unit and unit.strip() not in ("", "None", "nan"):
            return False

        # If row has ANY quantity, keep it (unit may be embedded in description).
        if any_has_qty:
            return False

        # Known header starts are always section headers, even with dimension text.
        mat_lower = material.lower()
        if any(mat_lower.startswith(w) for w in self._HEADER_STARTS):
            return True

        # Very long text with no unit and no quantity = section header / spec
        # paragraph, unless it is an item-numbered parent row in a grouped
        # tender (e.g. Sael rows 11.1 and 8).
        if not unit and not any_has_qty and len(material) > 100:
            return not has_item_number

        # Short item-numbered titles without unit/qty are section headings
        # (e.g. "UNDERDECK INSULATION", "COPPER REFRIGERANT PIPING…"), not
        # billable lines. Long item-numbered parents with billable verbs
        # remain legitimate rate-only rows (ISRO / Sael).
        if has_item_number and not unit and not any_has_qty:
            billable_verbs = (
                "supply",
                "install",
                "providing",
                "fixing",
                "laying",
                "fixish",
                "erect",
                "apply",
                "commission",
            )
            return len(material) < 90 and not any(v in mat_lower for v in billable_verbs)

        # If material contains dimension-like text ("mm dia", "mm thick", etc.)
        # it's probably a real item even without explicit unit/qty.
        # Otherwise, no unit and no quantity = likely a section header or note.
        return not bool(re.search(r"\d+\s*mm\s+(dia|thick|width|height|depth|length)", material, re.IGNORECASE))

    _EXCEL_ERRORS = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "#GETTING_DATA"}

    # Stronger spec-paragraph detection to catch multi-paragraph tender specs.
    _SPEC_PHRASES = (
        "application to be",
        "above mentioned specifications",
        "third party test certificates",
        "nabl accredited labs",
        "contracting agency",
        "manufacturer representative",
        "pmc/consultant",
        "thermal conductivity",
        "fire performance",
        "water absorption",
        "test to be conducted",
        "to be submitted",
        "to be borne by",
        "to be approved",
        "specifications to be",
        "certified by",
        "installation training",
        "site visit",
        "good workmanship",
        "well ventilated",
        "as per astm",
        "as per en",
        "as per is",
        "compliance to",
        "deviations to be filled",
    )

    def _is_spec_paragraph(self, material: str) -> bool:
        lower = material.lower()
        if len(material) > 1000:
            return True
        if 200 < len(material) <= 1000:
            phrase_count = sum(1 for p in self._SPEC_PHRASES if p in lower)
            if phrase_count >= 2:
                return True
        # Catch medium-length spec blocks (150-200) with very high spec density
        if 150 < len(material) <= 200:
            phrase_count = sum(1 for p in self._SPEC_PHRASES if p in lower)
            if phrase_count >= 3:
                return True
        return False

    def _has_excel_error(self, material: str) -> bool:
        return any(err in material for err in self._EXCEL_ERRORS)

    def _cell(self, row: tuple[Any, ...], col: int | None) -> str:
        if col is None or col >= len(row):
            return ""
        val = row[col]
        if val is None:
            return ""
        return str(val).rstrip("\n").strip()

    def _cell_or_none(self, row: tuple[Any, ...], col: int | None) -> Any:
        if col is None or col >= len(row):
            return None
        return row[col]

    def _parse_qty(self, raw: Any) -> Decimal:
        if raw is None:
            return Decimal("0")
        s = str(raw).strip().replace(",", "")
        if s in ("", "None", "nan", "-"):
            return Decimal("0")
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return Decimal("0")

    def _detect_action(self, material: str) -> str:
        m = material.lower()
        for kw in (
            "supply",
            "install",
            "provide",
            "fix",
            "furnish",
            "erect",
            "apply",
            "commission",
            "lay",
        ):
            if m.startswith(kw):
                return kw
        return "supply"

    def _map_columns_from_data(self, data_rows: list[tuple[Any, ...]]) -> _ColMapping:
        if not data_rows or len(data_rows[0]) < 4:
            return _ColMapping()
        return _ColMapping(
            material_col=1,
            quantity_cols=[3],
            unit_col=2,
        )

    # ------------------------------------------------------------------
    # Two-row short-name / detail-row merging
    # ------------------------------------------------------------------

    _ITEM_NUMBER_RE = re.compile(r"^\d+(\.\d+)*$")

    def _merge_short_name_rows(self, data_rows: list[tuple[Any, ...]], mapping: _ColMapping) -> list[tuple[Any, ...]]:
        """Merge short-name rows with their following detail rows.

        Some tenders (e.g. ISRO-style) place the concise material name in one
        row with an item number but no unit/qty, and the full specification +
        quantities in the very next row.  Detect this pattern and synthesise a
        single row that keeps the short name but carries the unit/qty.
        """
        if mapping.material_col is None:
            return data_rows

        merged: list[tuple[Any, ...]] = []
        i = 0
        while i < len(data_rows):
            row = data_rows[i]
            if i + 1 >= len(data_rows):
                merged.append(row)
                i += 1
                continue

            next_row = data_rows[i + 1]

            if self._is_short_name_row(row, mapping) and self._is_detail_row(next_row, mapping, row):
                # Synthesise a row that keeps the short material but copies
                # unit / quantity / total from the detail row.
                max_len = max(
                    len(row),
                    len(next_row),
                    (mapping.material_col or 0) + 1,
                    (mapping.unit_col or 0) + 1,
                )
                new_row = list(row) + [None] * (max_len - len(row))

                if mapping.unit_col is not None and mapping.unit_col < len(next_row):
                    new_row[mapping.unit_col] = next_row[mapping.unit_col]
                if mapping.total_col is not None and mapping.total_col < len(next_row):
                    new_row[mapping.total_col] = next_row[mapping.total_col]
                for qc in mapping.quantity_cols or []:
                    if qc < len(next_row):
                        new_row[qc] = next_row[qc]
                # Prefer the child's long description as the material text; the
                # parent's short name is just a heading (e.g. ISRO tenders).
                if mapping.material_col is not None and mapping.material_col < len(next_row):
                    child_mat = next_row[mapping.material_col]
                    if child_mat is not None and str(child_mat).strip():
                        new_row[mapping.material_col] = child_mat

                merged.append(tuple(new_row))
                i += 2  # skip both consumed rows
            else:
                merged.append(row)
                i += 1

        return merged

    def _is_short_name_row(self, row: tuple[Any, ...], mapping: _ColMapping) -> bool:
        """Row looks like a concise material name with an item number but no qty/unit."""
        # Must have an item-number-ish first cell
        if len(row) == 0 or row[0] is None:
            return False
        first = str(row[0]).strip()
        if not self._ITEM_NUMBER_RE.match(first):
            return False

        material = self._cell(row, mapping.material_col)
        if not material or len(material) < 5 or len(material) > 120:
            return False

        # Must NOT look like a note / total / spec block
        if self._NOTE_RE.search(material):
            return False
        if TOTAL_KEYWORD_PATTERN.search(material):
            return False
        mat_lower = material.lower()
        if any(mat_lower.startswith(w) for w in self._HEADER_STARTS):
            return False
        if self._is_spec_paragraph(material):
            return False

        # Must NOT have a real unit
        unit = self._cell(row, mapping.unit_col)
        if unit and unit.strip() and not self._looks_like_header_unit(unit):
            return False

        # Must NOT have any quantity
        return not self._row_has_any_quantity(row, mapping.quantity_cols or [])

    def _is_detail_row(
        self,
        next_row: tuple[Any, ...],
        mapping: _ColMapping,
        short_row: tuple[Any, ...],
    ) -> bool:
        """Next row looks like the long-description + qty/unit partner."""
        # First cell should be empty (or at least not an item number)
        if len(next_row) > 0 and next_row[0] is not None:
            first = str(next_row[0]).strip()
            if first and self._ITEM_NUMBER_RE.match(first):
                return False

        material = self._cell(next_row, mapping.material_col)
        if not material or len(material) < 50:
            return False

        short_material = self._cell(short_row, mapping.material_col)
        # Detail must be substantially longer than the short name
        if len(material) <= len(short_material) * 2 and len(material) < 200:
            return False

        # Must carry a real unit or a real quantity (incl. TOTAL col)
        unit = self._cell(next_row, mapping.unit_col)
        has_unit = bool(unit and unit.strip() and not self._looks_like_header_unit(unit))
        qty_cols = list(mapping.quantity_cols or [])
        if mapping.total_col is not None and mapping.total_col not in qty_cols:
            qty_cols.append(mapping.total_col)
        has_qty = self._row_has_any_quantity(next_row, qty_cols)
        return bool(has_unit or has_qty)

    def _looks_like_header_unit(self, unit: str) -> bool:
        u = unit.lower().strip()
        return u in ("unit", "uom", "units", "measure", "qty", "quantity")

    # ------------------------------------------------------------------
    # P3_03: hierarchy + parent_context application
    # ------------------------------------------------------------------

    def _apply_hierarchy(self, boq_rows: list[BoqRow], source_item_nos: list[str]) -> None:
        """Attach parent_context to each BoqRow in place.

        parent_context is the chain of ancestor material descriptions
        (outermost parent first, item's own material last). It enables
        BOQ assembly + export display to show children in their parent's
        frame (e.g. 11.1.1 -> ["11 THERMAL INSULATION",
        "11.1 Duct insulation", "11.1.1 50mm thick ..."]).

        Malformed / unparseable item numbers get parent_context = [material]
        (orphan node, no parents).

        No rows are merged or dropped. Pure data transform.
        """
        if not boq_rows:
            return
        if len(boq_rows) != len(source_item_nos):
            # Defensive: if the parallel lists got out of sync, skip
            # parent_context resolution (still preserve rows; R1).
            return
        items: list[dict] = []
        for row, src_no in zip(boq_rows, source_item_nos, strict=True):
            items.append(
                {
                    "item_no_raw": src_no,
                    "material": row.material,
                    "_row": row,
                }
            )
        apply_parent_context(items, item_no_key="item_no_raw", material_key="material", context_key="parent_context")
        for it in items:
            ctx = it.get("parent_context") or []
            row_ref = it["_row"]
            row_ref.parent_context = list(ctx)

    # ------------------------------------------------------------------
    # P3_03: multi-line wrap assembly
    # ------------------------------------------------------------------

    def _assemble_wrapped_rows(
        self,
        data_rows: list[tuple[Any, ...]],
        mapping: _ColMapping,
    ) -> list[tuple[Any, ...]]:
        """Assemble multi-line / wrapped description rows.

        A "wrapped" row is one where ALL of these are true:
          - item-number column is empty/None
          - material column has text
          - NO quantity cell (no value in any detected quantity column)
          - NO unit cell
        Consecutive such rows below a REAL BOQ row are joined into the
        parent's material cell.

        The qty+unit+item# guard pattern (incident #8 / `f3affab`) is
        respected: rows with a real qty or unit are NEVER treated as
        continuations -- they are independent BOQ items and must be
        extracted as their own row (R1: no silent merging).

        We additionally only join a row into the previous one if the
        previous emitted row was a real BOQ row (had item number OR
        qty/unit signal). Otherwise the "previous" is a section-header
        / spec paragraph and the current row is emitted as-is so the
        downstream _is_section_header / _is_spec_paragraph checks can
        handle it.

        No rows are added; continuation rows are absorbed into the parent
        only when they carry zero BOQ signal of their own AND the parent
        is also a real BOQ row.
        """
        if not data_rows:
            return data_rows
        if mapping.material_col is None:
            return data_rows
        item_col = 0
        mat_col = mapping.material_col
        qty_cols = mapping.quantity_cols or []
        unit_col = mapping.unit_col

        def _has_qty_signal(cells: list[Any]) -> bool:
            for qc in qty_cols:
                if qc is not None and qc < len(cells):
                    v = cells[qc]
                    if v is not None and str(v).strip() not in ("", "None", "nan", "-"):
                        return True
            return False

        def _has_unit_signal(cells: list[Any]) -> bool:
            if unit_col is None:
                return False
            if unit_col >= len(cells):
                return False
            v = cells[unit_col]
            if v is None:
                return False
            s = str(v).strip()
            if not s or s in ("None", "nan", "-"):
                return False
            return not self._looks_like_header_unit(s)

        def _has_item_signal(cells: list[Any]) -> bool:
            if item_col >= len(cells):
                return False
            v = cells[item_col]
            if v is None:
                return False
            s = str(v).strip()
            return bool(s) and bool(self._ITEM_NUMBER_PATTERN.match(s))

        def _is_note_or_total(cells: list[Any]) -> bool:
            mat = cells[mat_col] if mat_col < len(cells) else None
            if mat is None:
                return False
            s = str(mat).strip()
            if not s:
                return False
            if self._NOTE_RE.search(s):
                return True
            return bool(TOTAL_KEYWORD_PATTERN.search(s))

        def _is_real_boq(cells: list[Any]) -> bool:
            """A row is a real BOQ row if it has at least one real signal."""
            return _has_item_signal(cells) or _has_qty_signal(cells) or _has_unit_signal(cells)

        out: list[tuple[Any, ...]] = []
        for row in data_rows:
            cells = list(row)
            while len(cells) <= mat_col:
                cells.append(None)
            if _is_real_boq(cells) or _is_note_or_total(cells):
                # Independent BOQ row, or a note/total row: emit as-is
                out.append(tuple(cells))
            else:
                # Continuation row: only join if the previous emitted row
                # was a real BOQ row. Otherwise emit as-is (downstream
                # section-header / spec-paragraph filters handle it).
                if out and _is_real_boq(list(out[-1])):
                    prev_cells = list(out[-1])
                    while len(prev_cells) <= mat_col:
                        prev_cells.append(None)
                    cont = cells[mat_col]
                    if cont is not None and str(cont).strip():
                        prev_mat = prev_cells[mat_col] or ""
                        sep = "" if (not prev_mat or str(prev_mat).endswith((" ", "\n"))) else " "
                        prev_cells[mat_col] = (str(prev_mat) + sep + str(cont).strip()).strip()
                    out[-1] = tuple(prev_cells)
                else:
                    out.append(tuple(cells))
        return out
