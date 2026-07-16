"""XLSX parsing using openpyxl."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass(slots=True)
class SheetData:
    name: str
    headers: list[str] = field(default_factory=list)
    rows: list[list[Any]] = field(default_factory=list)


@dataclass(slots=True)
class XLSXParseResult:
    sheets: list[SheetData] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)


class XLSXParser:
    """Parse XLSX files and extract sheet data."""

    def parse(self, path: str | Path) -> XLSXParseResult:
        """Parse an XLSX file and return all sheets with headers and rows."""
        from openpyxl import load_workbook

        path = Path(path)
        wb = load_workbook(path, data_only=True, read_only=True)
        sheets: list[SheetData] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows: list[list[Any]] = []

            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    all_rows.append(list(row))

            headers: list[str] = []
            rows: list[list[Any]] = []

            if all_rows:
                headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
                rows = [[str(c).strip() if c is not None else "" for c in r] for r in all_rows[1:]]

            sheets.append(
                SheetData(
                    name=sheet_name,
                    headers=headers,
                    rows=rows,
                )
            )

        max_row = max((len(s.rows) for s in sheets), default=0)
        total_rows = sum(len(s.rows) for s in sheets)

        return XLSXParseResult(
            sheets=sheets,
            metadata={
                "sheet_count": len(sheets),
                "total_rows": total_rows,
                "max_row": max_row,
            },
        )
