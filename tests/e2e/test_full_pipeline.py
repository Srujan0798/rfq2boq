import json

from src.boq_generator import BOQGenerator
from src.unit_normalization import calculate_quantity, normalize_dimension, normalize_unit, parse_dimension


class TestFullPipeline:
    def test_end_to_end_boq_generation(self, tmp_path):
        generator = BOQGenerator()
        boq_items = [
            {"description": "Cement", "material": "cement", "quantity": 100, "unit": "bags", "confidence": 0.9},
            {"description": "Steel", "material": "steel", "quantity": 200, "unit": "kg", "confidence": 0.85},
            {"description": "Concrete", "material": "concrete", "quantity": 50, "unit": "m³", "confidence": 0.88},
        ]

        output = generator.generate(boq_items)

        assert "metadata" in output
        assert "boq" in output
        assert len(output["boq"]["items"]) == 3
        assert output["boq"]["summary"]["total_items"] == 3

    def test_json_output_generation(self, tmp_path):
        generator = BOQGenerator()
        boq_items = [
            {"description": "Cement", "material": "cement", "quantity": 100, "unit": "bags", "confidence": 0.9},
        ]
        output_path = tmp_path / "boq.json"

        generator.generate_json(boq_items, str(output_path))

        assert output_path.exists()
        with open(output_path) as f:
            data = json.load(f)
        assert "boq" in data
        assert len(data["boq"]["items"]) == 1

    def test_excel_output_generation(self, tmp_path):
        from openpyxl import load_workbook

        generator = BOQGenerator()
        boq_items = [
            {"description": "Cement", "material": "cement", "quantity": 100, "unit": "bags", "confidence": 0.9},
        ]
        output_path = tmp_path / "boq.xlsx"

        generator.generate_excel(boq_items, str(output_path))

        assert output_path.exists()
        wb = load_workbook(output_path)
        assert "Bill of Quantities" in wb.sheetnames

    def test_item_code_sequential(self):
        generator = BOQGenerator()
        boq_items = [
            {"description": "Item 1", "material": "cement", "quantity": 100, "unit": "bags"},
            {"description": "Item 2", "material": "steel", "quantity": 200, "unit": "kg"},
            {"description": "Item 3", "material": "concrete", "quantity": 50, "unit": "m³"},
        ]

        output = generator.generate(boq_items)
        item_codes = [item["item_code"] for item in output["boq"]["items"]]
        assert item_codes == ["BOQ-001", "BOQ-002", "BOQ-003"]


class TestUnitNormalization:
    def test_normalize_unit_m2(self):
        assert normalize_unit("sq.m") == "m²"
        assert normalize_unit("sqm") == "m²"
        assert normalize_unit("m2") == "m²"

    def test_normalize_unit_m3(self):
        assert normalize_unit("cum") == "m³"
        assert normalize_unit("cubic meter") == "m³"
        assert normalize_unit("m3") == "m³"

    def test_normalize_unit_count(self):
        assert normalize_unit("nos") == "no."
        assert normalize_unit("nos.") == "no."
        assert normalize_unit("each") == "no."

    def test_normalize_dimension(self):
        assert normalize_dimension("20mm thick") == "20mm"
        assert normalize_dimension("2mm") == "2mm"

    def test_calculate_running_meter(self):
        qty, unit = calculate_quantity("100rm of 20cm wide marble", 100, "rm", "20cm")
        assert qty == 100
        assert unit == "rm"


class TestQuantityCalculations:
    def test_compound_dimension_calculation(self):
        result = parse_dimension("3m x 2m")
        assert result["length"] == 3.0
        assert result["width"] == 2.0

    def test_thickness_parsing(self):
        result = parse_dimension("20mm")
        assert result["thickness"] == 0.020
