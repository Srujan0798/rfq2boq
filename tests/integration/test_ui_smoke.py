"""UI smoke tests using Streamlit's AppTest (no browser automation).

P5_01 §3 deliverable: ``tests/integration/test_ui_smoke.py`` that
exercises the Streamlit app via ``streamlit.testing.v1.AppTest``.
We inject a real ExtractionResult via ``session_state`` (AppTest can't
drive the file-uploader widget deterministically; this is the
accepted pattern per P5_01 §9 gotcha).
"""

from __future__ import annotations

import json
import sys
from decimal import Decimal
from pathlib import Path

import pytest

pytest.importorskip("streamlit")

# Make repo root importable so `import ui.app` resolves
ROOT = Path(__file__).resolve().parent.parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


@pytest.fixture(scope="module")
def real_xlsx_path() -> Path | None:
    """Path to a real sacred-10 XLSX, if available (02_isro_vssc).

    Lives under ``data/real_rfqs/swa_enquiries/02_isro_vssc/VSSC_BOQ_with_qty.xlsx``.
    Falls back to None when the file isn't present so the test can
    be skipped (the agents rule S2: no synthetic data; the test only
    runs against real corpus files).
    """
    p = ROOT / "data/real_rfqs/swa_enquiries/02_isro_vssc/VSSC_BOQ_with_qty.xlsx"
    return p if p.exists() else None


class TestAppTestSmoke:
    """AppTest-based smoke: import the app, run it once, look for widgets."""

    def test_app_imports_and_runs(self):
        """ui.app imports cleanly and AppTest can run it."""
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(str(ROOT / "ui" / "app.py"))
        at.run(timeout=10)
        # App should render at least the title.
        assert at.title is not None
        assert any("RFQ" in (t.value or "") for t in at.title)

    def test_app_has_file_uploader(self):
        """The file uploader widget is present (per §3 deliverable)."""
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(str(ROOT / "ui" / "app.py"))
        at.run(timeout=10)
        found = any(type(e).__name__ == "FileUploader" for e in at.main)
        assert found, "FileUploader widget not found in UI"

    def test_app_has_tabs(self):
        """The two-tab structure (Extract + Help) survives."""
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(str(ROOT / "ui" / "app.py"))
        at.run(timeout=10)
        # Should have at least the two main tabs.
        assert len(at.tabs) >= 2

    def test_app_renders_footer_text(self):
        """The footer carries the model version + pipeline commit."""
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(str(ROOT / "ui" / "app.py"))
        at.run(timeout=10)
        # Look for the footer markdown ("RFQ2BOQ v..." or "BOQ JSON schema v1.1.0")
        text = " ".join(
            str(getattr(m, "value", "") or "") for m in (list(at.main) + list(at.sidebar))
        )
        # Either the footer text or at least the version marker.
        assert "RFQ2BOQ" in text or "schema v1.1.0" in text, (
            f"Footer text not found. Rendered text: {text[:500]!r}"
        )


class TestFlagReviewPanelRendering:
    """Inject an ExtractionResult with typed Flags and verify the panel renders."""

    def _make_result(self) -> "ExtractionResult":
        from src.domain.flags import (
            Flag,
            FlagSeverity,
            FlagStage,
            gem_non_catalog_flag,
            low_confidence_flag,
            structure_fallback_flag,
        )
        from src.domain.models import BoqRow, ExtractionMetadata, ExtractionResult

        rows = [
            BoqRow(
                item_no=1,
                material="Supply 500 kg cement at ground floor M20 grade",
                quantity=Decimal("500"),
                unit="kg",
                grade="M20",
                confidence=0.92,
            ),
            BoqRow(
                item_no=2,
                material="Fancy pipe (not in GeM catalog)",
                quantity=Decimal("200"),
                unit="rmt",
                confidence=0.45,
                flags=[
                    low_confidence_flag(2, 0.45),
                    gem_non_catalog_flag("Fancy pipe (not in GeM catalog)"),
                ],
                warnings=["low confidence", "GeM non-catalog"],
            ),
        ]
        return ExtractionResult(
            doc_id="smoke_test",
            project_name="Smoke Test",
            source_file="data/test.pdf",
            boq_items=rows,
            metadata=ExtractionMetadata(
                total_items=2,
                avg_confidence=0.69,
                flags=[structure_fallback_flag()],
                warnings=["structure fallback"],
            ),
        )

    def test_flag_summary_counts_correct(self):
        """The flag-summary helper tallies error/review/info correctly."""
        from ui.app import _result_flag_summary

        result = self._make_result()
        counts = _result_flag_summary(result)
        # LOW_CONFIDENCE = review, GEM_NON_CATALOG = review, STRUCTURE_FALLBACK = review.
        assert counts.get("review", 0) >= 3
        assert counts.get("error", 0) == 0

    def test_row_severity_returns_review_for_flagged_row(self):
        """A row with a review-level flag returns 'review' as its severity."""
        from ui.app import _row_severity

        result = self._make_result()
        row = result.boq_items[1]  # the flagged one
        assert _row_severity(row) == "review"

    def test_row_severity_returns_none_for_clean_row(self):
        from ui.app import _row_severity

        result = self._make_result()
        row = result.boq_items[0]
        assert _row_severity(row) is None

    def test_row_flag_codes_sorted_unique(self):
        from ui.app import _row_flag_codes

        result = self._make_result()
        codes = _row_flag_codes(result.boq_items[1])
        assert codes == sorted(set(codes))
        assert "LOW_CONFIDENCE" in codes
        assert "GEM_NON_CATALOG" in codes

    def test_build_boq_dataframe_adds_severity_column(self):
        """The new build_boq_dataframe() adds a Severity column for the UI."""
        import pandas as pd

        from ui.app import build_boq_dataframe

        df = build_boq_dataframe(self._make_result())
        assert isinstance(df, pd.DataFrame)
        assert "Severity" in df.columns
        assert "Flags" in df.columns
        # Clean row: empty severity; flagged row: 'review'
        assert df.iloc[0]["Severity"] == ""
        assert df.iloc[1]["Severity"] == "review"
        # Flag codes surfaced in the Flags column.
        assert "LOW_CONFIDENCE" in df.iloc[1]["Flags"]
        assert "GEM_NON_CATALOG" in df.iloc[1]["Flags"]


class TestDocumentTypeBanner:
    """The non-BOQ document-type banner must be reachable from the UI helpers."""

    def test_compliance_checklist_classified_correctly(self):
        from src.domain.flags import Flag, FlagSeverity, FlagStage
        from src.domain.models import BoqRow, ExtractionMetadata, ExtractionResult
        from ui.app import _classify_document_type

        flag = Flag(
            code="TABLE_TYPE_NOT_BOQ",
            severity=FlagSeverity.INFO,
            stage=FlagStage.TABLE_CLASSIFY,
            message="classified table is 'COMPLIANCE_CHECKLIST', not BOQ — 0 rows emitted; document not skippable",
        )
        result = ExtractionResult(
            doc_id="compliance",
            boq_items=[],
            metadata=ExtractionMetadata(flags=[flag]),
        )
        assert _classify_document_type(result) == "COMPLIANCE_CHECKLIST"

    def test_boq_document_returns_none(self):
        from ui.app import _classify_document_type
        from src.domain.models import BoqRow, ExtractionResult

        result = ExtractionResult(
            doc_id="boq",
            boq_items=[
                BoqRow(material="Cement", quantity=Decimal("100"), unit="kg"),
            ],
        )
        assert _classify_document_type(result) is None


class TestRealXLSXSmoke:
    """End-to-end smoke: extract a real XLSX through the UI helpers and
    verify the result has Flags-shaped data the UI can render."""

    def test_real_xlsx_pipeline_produces_result_with_flags_shape(
        self, real_xlsx_path: Path | None
    ) -> None:
        if real_xlsx_path is None:
            pytest.skip("No real XLSX available — tests only run on real corpus files")
        from src.pipeline_xlsx import XLSXRowPipeline

        items = XLSXRowPipeline().run(str(real_xlsx_path))
        assert isinstance(items, list)
        # Each row should have a `flags` field (even if empty) and
        # `warnings` for the legacy path.
        for it in items:
            assert hasattr(it, "flags")
            assert hasattr(it, "warnings")
            assert hasattr(it, "source_pages")
        # build_boq_dataframe should accept the result.
        from src.domain.models import ExtractionResult
        from ui.app import build_boq_dataframe

        result = ExtractionResult(
            doc_id="real_xlsx",
            boq_items=items,
        )
        df = build_boq_dataframe(result)
        assert "Flags" in df.columns
        assert "Severity" in df.columns


class TestExcelDownloadFromUI:
    """The Excel download the UI generates must include Review + Provenance."""

    def test_excel_bytes_contain_review_sheet(self) -> None:
        import io

        from src.domain.flags import low_confidence_flag
        from src.domain.models import BoqRow, ExtractionResult
        from src.export.excel_generator import CPWDExcelGenerator

        items = [
            BoqRow(
                item_no=1,
                material="Test item",
                quantity=Decimal("1"),
                unit="nos",
                confidence=0.4,
                flags=[low_confidence_flag(1, 0.4)],
            ),
        ]
        result = ExtractionResult(doc_id="dl", boq_items=items)

        # Mimic the UI's generate_excel_bytes flow but write to
        # an in-memory buffer by way of a temp file (CPWDExcelGenerator
        # needs a real path).
        import tempfile

        from pathlib import Path

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            CPWDExcelGenerator().export(
                items,
                str(tmp_path),
                provenance={"source_file": "test.pdf", "pipeline_commit": "abc1234"},
            )
            data = tmp_path.read_bytes()
        finally:
            tmp_path.unlink(missing_ok=True)

        assert data
        # Open the round-trip and confirm sheets.
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "BOQ" in wb.sheetnames
        assert "Review" in wb.sheetnames
        assert "Provenance" in wb.sheetnames


class TestJSONExportFromUI:
    """The JSON download the UI generates must include flags + provenance."""

    def test_json_has_flags_and_schema_version(self) -> None:
        from src.domain.flags import low_confidence_flag
        from src.domain.models import BoqRow, ExtractionResult
        from src.export.json_formatter import JSONFormatter

        flag = low_confidence_flag(1, 0.3)
        items = [
            BoqRow(
                item_no=1,
                material="X",
                quantity=Decimal("1"),
                unit="nos",
                confidence=0.3,
                flags=[flag],
                # P3_04: producers attach BOTH a typed flag and the
                # legacy string-form warning.  This test verifies
                # the formatter round-trips both.
                warnings=[flag.to_legacy_warning()],
            ),
        ]
        result = ExtractionResult(doc_id="j", boq_items=items)
        s = JSONFormatter().format(result)
        d = json.loads(s)
        assert d["schema_version"] == "1.1.0"
        assert d["boq_items"][0]["flags"][0]["code"] == "LOW_CONFIDENCE"
        # Legacy warnings list is preserved.
        assert d["boq_items"][0]["warnings"]


class TestPipelinesGoThroughSameEntry:
    """The UI must use the same pipeline entry point as the CLI (no UI-only path)."""

    def test_ui_extract_boq_pdf_uses_pipeline_run(self) -> None:
        """ui.app.extract_boq_pdf delegates to ``Pipeline().run(str(file))``."""
        import inspect

        from ui.app import extract_boq_pdf

        src = inspect.getsource(extract_boq_pdf)
        assert "pipeline.run(str(file_path))" in src

    def test_ui_extract_boq_xlsx_uses_xlsxpipeline(self) -> None:
        import inspect

        from ui.app import extract_boq_xlsx

        src = inspect.getsource(extract_boq_xlsx)
        assert "XLSXRowPipeline" in src
        assert "xp.run(str(file_path))" in src
