"""Tests for export validation (E4: BoqRow.validate before export)."""

from __future__ import annotations

import json
import tempfile
from decimal import Decimal
from pathlib import Path

from src.domain.models import BoqRow, ExtractionResult
from src.export.excel_generator import CPWDExcelGenerator
from src.export.json_formatter import JSONFormatter


class TestJSONExportValidation:
    def test_invalid_rows_skipped(self) -> None:
        result = ExtractionResult(
            doc_id="test",
            boq_items=[
                BoqRow(material="Cement", quantity=Decimal("100"), unit="kg"),
                BoqRow(material="", quantity=Decimal("0"), unit=""),  # invalid
                BoqRow(material="Sand", quantity=Decimal("50"), unit="cum"),
            ],
        )
        formatter = JSONFormatter()
        output = formatter.format(result)
        data = json.loads(output)
        assert len(data["boq_items"]) == 2
        assert data["boq_items"][0]["material"] == "Cement"
        assert data["boq_items"][1]["material"] == "Sand"

    def test_rate_only_rows_not_skipped(self) -> None:
        result = ExtractionResult(
            doc_id="test",
            boq_items=[
                BoqRow(material="Item", quantity=Decimal("0"), unit="Sqm.", rate_only=True),
            ],
        )
        formatter = JSONFormatter()
        output = formatter.format(result)
        data = json.loads(output)
        assert len(data["boq_items"]) == 1
        assert data["boq_items"][0]["rate_only"] is True

    def test_all_invalid_returns_empty(self) -> None:
        result = ExtractionResult(
            doc_id="test",
            boq_items=[
                BoqRow(material="", quantity=Decimal("0"), unit=""),
            ],
        )
        formatter = JSONFormatter()
        output = formatter.format(result)
        data = json.loads(output)
        assert len(data["boq_items"]) == 0


class TestExcelExportValidation:
    def test_invalid_rows_skipped_in_excel(self) -> None:
        items = [
            BoqRow(material="Cement", quantity=Decimal("100"), unit="kg"),
            BoqRow(material="", quantity=Decimal("0"), unit=""),  # invalid
            BoqRow(material="Sand", quantity=Decimal("50"), unit="cum"),
        ]
        gen = CPWDExcelGenerator()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        gen.export(items, path)

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        data_rows = [r for r in ws.iter_rows(min_row=10, values_only=True) if r[0] is not None and isinstance(r[0], int)]
        assert len(data_rows) == 2

    def test_rate_only_row_shows_flag_in_excel(self) -> None:
        items = [
            BoqRow(material="Rate Item", quantity=Decimal("0"), unit="Sqm.", rate_only=True, confidence=0.70),
        ]
        gen = CPWDExcelGenerator()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        gen.export(items, path)

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        data_rows = [r for r in ws.iter_rows(min_row=10, values_only=True) if r[0] is not None and isinstance(r[0], int)]
        assert len(data_rows) == 1
        remarks = data_rows[0][7]  # Remarks column
        assert "flag" in str(remarks).lower() or "qty" in str(remarks).lower()

    def test_empty_material_skipped_not_crashed(self) -> None:
        items = [
            BoqRow(material="", quantity=Decimal("0"), unit=""),
            BoqRow(material="Valid", quantity=Decimal("10"), unit="kg"),
        ]
        gen = CPWDExcelGenerator()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        gen.export(items, path)
        assert Path(path).exists()
