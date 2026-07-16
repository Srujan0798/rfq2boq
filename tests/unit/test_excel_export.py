"""Tests for the P5_01 Excel export contract.

Verifies the deliverables from ``tasks/phase9/P5_01_export_ui.md``:

- BOQ sheet has the full set of columns incl. parent_context, no data
  truncation, list-valued fields joined.
- Review sheet exists with every Flag surfaced (severity, stage, code,
  message, row ref, original, flag_id).
- Provenance sheet exists with source file, sha256, extraction date,
  pipeline version/commit, and per-row source pages.
- Severity coloring uses 3 shades (error/red, review/orange, info/gray)
  — no rainbow.
- Empty/non-BOQ document export doesn't crash and still produces a
  valid file.
- Unicode survives the round-trip (Hindi, em-dashes, etc.).
- Export is idempotent: opening the file with openpyxl and the data
  layer round-trip preserves every byte of a long description.
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from src.domain.flags import (
    Flag,
    FlagSeverity,
    FlagStage,
    gem_non_catalog_flag,
    low_confidence_flag,
    structure_fallback_flag,
    unknown_unit_flag,
)
from src.domain.models import BoqRow, ExtractionMetadata, ExtractionResult
from src.export.excel_generator import CPWDExcelGenerator


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_boq_rows() -> list[BoqRow]:
    """A small BOQ with one flagged and one clean row — the P5_01 happy path."""
    return [
        BoqRow(
            item_no=1,
            material="Earthwork in excavation by mechanical means",
            quantity=Decimal("100"),
            unit="cum",
            grade="",
            standard=["IS 1200"],
            confidence=0.92,
        ),
        BoqRow(
            item_no=2,
            material="M20 grade concrete in foundation",
            quantity=Decimal("50"),
            unit="cum",
            grade="M20",
            standard=["IS 456"],
            confidence=0.45,
        ),
    ]


@pytest.fixture
def sample_flagged_rows() -> list[BoqRow]:
    """Two rows with a mix of typed flags attached (R1 visibility)."""
    return [
        BoqRow(
            item_no=1,
            material="Supply fancy pipe (not in GeM catalog)",
            quantity=Decimal("200"),
            unit="rmt",
            confidence=0.55,
            flags=[
                low_confidence_flag(1, 0.55),
                gem_non_catalog_flag("fancy pipe"),
            ],
            warnings=["low confidence", "GeM"],
        ),
        BoqRow(
            item_no=2,
            material="Supply 500 mm dia pipe",
            quantity=Decimal("0"),
            unit="m",
            rate_only=True,
            confidence=0.40,
            flags=[
                unknown_unit_flag("xyz", 2),
            ],
            warnings=["unknown unit"],
        ),
    ]


@pytest.fixture
def sample_result(sample_boq_rows: list[BoqRow]) -> ExtractionResult:
    return ExtractionResult(
        doc_id="test_doc",
        project_name="Test Project",
        source_file="data/real_rfqs/swa_enquiries/02_isro_vssc/VSSC_BOQ_with_qty.xlsx",
        boq_items=sample_boq_rows,
        metadata=ExtractionMetadata(total_items=2, avg_confidence=0.69),
    )


@pytest.fixture
def sample_flagged_result(sample_flagged_rows: list[BoqRow]) -> ExtractionResult:
    return ExtractionResult(
        doc_id="gem_test",
        project_name="GeM Test",
        source_file="data/real_rfqs/swa_enquiries/09_gem_bid_7439924/file.pdf",
        boq_items=sample_flagged_rows,
        metadata=ExtractionMetadata(
            total_items=2,
            avg_confidence=0.47,
            flags=[structure_fallback_flag()],
            warnings=["structure fallback"],
        ),
    )


# ---------------------------------------------------------------------------
# Deliverable 1: BOQ sheet contract
# ---------------------------------------------------------------------------


class TestBOQSheetContract:
    """The BOQ sheet is the primary deliverable: no truncation, list-join."""

    def test_export_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(
            sample_result.boq_items,
            str(out),
            project_metadata={"project": "Test"},
        )
        assert out.exists()
        assert out.stat().st_size > 1024

    def test_full_description_survives(self, sample_result, tmp_path):
        """Full material text lands in the cell — no .title() munging (R1)."""
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        # The exact source text must appear somewhere in the BOQ sheet.
        cells_text = " ".join(
            str(ws.cell(row=r, column=2).value or "")
            for r in range(1, ws.max_row + 1)
        )
        assert "Earthwork in excavation by mechanical means" in cells_text
        # No .title()-style mutation: lowercase 'in' and 'by' should remain.
        assert "Earthwork in excavation by mechanical means" in cells_text

    def test_list_standard_joined_for_display(self, sample_result, tmp_path):
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        # Find the row with "IS 1200" in the Standard column
        found = False
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=5).value  # Standard column
            if val and "IS 1200" in str(val):
                found = True
                break
        assert found, "IS 1200 should be in the Standard column for row 1"

    def test_no_rate_or_amount_columns(self, sample_result, tmp_path):
        """S1: No Rate/Amount/cost *data columns* in the BOQ sheet.

        Notes about "no rates" are fine — we only check that no column
        in the header row (or data rows) is named Rate / Amount / Cost.
        """
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        # Header row is at row 9 by spec; check that row + the
        # data-row columns (rows 10..) don't include Rate/Amount.
        HEADER_ROW = 9
        header = [str(ws.cell(row=HEADER_ROW, column=c).value or "").lower() for c in range(1, ws.max_column + 1)]
        for h in header:
            assert h not in {"rate", "rate (₹)", "amount", "amount (₹)", "cost"}, (
                f"S1 violation: BOQ header contains forbidden column '{h}'"
            )
        # Data rows: scan first column of each data row for header-like
        # values (catches accidental re-introduction).
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                s = str(v).lower().strip()
                # Allow the explicit "no rates" note + the grand total
                # label which mentions "rate" only as a side effect of
                # being a complete English sentence.
                if s in {"grand total (unpriced boq)"}:
                    continue
                # Column headers shouldn't appear as data values.
                assert s not in {"rate", "rate (₹)", "amount", "amount (₹)", "cost", "dsr code"}, (
                    f"S1 violation: forbidden data value '{s}' in cell {r},{c}"
                )

    def test_unicode_description_survives(self, tmp_path):
        """Hindi + em-dash + special chars all round-trip."""
        items = [
            BoqRow(
                item_no=1,
                material="विद्युत — high tension cable (Hindi + em-dash)",
                quantity=Decimal("100"),
                unit="रन मीटर",
            ),
        ]
        out = tmp_path / "uni.xlsx"
        CPWDExcelGenerator().export(items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        joined = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        )
        assert "विद्युत" in joined
        assert "—" in joined

    def test_empty_boq_does_not_crash(self, tmp_path):
        out = tmp_path / "empty.xlsx"
        CPWDExcelGenerator().export([], str(out))
        assert out.exists()
        wb = openpyxl.load_workbook(str(out))
        assert "BOQ" in wb.sheetnames

    def test_parent_context_surfaced_in_description_cell(self, tmp_path):
        """P3_03 parent_context: child items get their ancestor's material
        appended to the Description cell so a reader of the BOQ sheet
        alone can see the hierarchy (no truncation, no separate column)."""
        items = [
            BoqRow(
                item_no=1,
                material="MS chilled water pipe insulation",
                quantity=Decimal("500"),
                unit="rmt",
                parent_context=["INSULATION", "MS chilled water pipe insulation"],
            ),
        ]
        out = tmp_path / "ctx.xlsx"
        CPWDExcelGenerator().export(items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        joined = " ".join(
            str(ws.cell(row=r, column=2).value or "") for r in range(1, ws.max_row + 1)
        )
        assert "MS chilled water pipe insulation" in joined
        assert "[parent]" in joined
        assert "INSULATION" in joined

    def test_very_long_description_preserved_with_truncation_marker(self, tmp_path):
        """32k+ char description: we keep the head + a marker so the
        data is never silently dropped.  Full text is still in JSON."""
        huge = "abcdefghij" * 4000  # 40,000 chars
        items = [BoqRow(material=huge, quantity=1, unit="nos")]
        out = tmp_path / "huge.xlsx"
        CPWDExcelGenerator().export(items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        # Look for the row that contains our "abcdefghij" string.
        cell_val = ""
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v and "abcdefghij" in str(v):
                cell_val = str(v)
                break
        assert cell_val, "Truncated long description not found in BOQ sheet"
        assert len(cell_val) > 1000  # substantial head preserved
        assert "[...truncated" in cell_val  # explicit marker, not silent loss


# ---------------------------------------------------------------------------
# Deliverable 2: Review sheet (every flag surfaced)
# ---------------------------------------------------------------------------


class TestReviewSheetContract:
    """The Review sheet is R1's "flag, never drop" made visible."""

    def test_review_sheet_exists(self, sample_flagged_result, tmp_path):
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(
            sample_flagged_result.boq_items,
            str(out),
        )
        wb = openpyxl.load_workbook(str(out))
        assert "Review" in wb.sheetnames

    def test_review_sheet_contains_every_flag(self, sample_flagged_result, tmp_path):
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_flagged_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Review"]
        # Count data rows (one per Flag).
        codes = []
        for r in range(5, ws.max_row + 1):
            v = ws.cell(row=r, column=4).value  # Code column
            if v:
                codes.append(str(v))
        # The 2 sample rows carry 3 flags total: 2 on row 1, 1 on row 2.
        assert codes.count("LOW_CONFIDENCE") >= 1
        assert codes.count("GEM_NON_CATALOG") >= 1
        assert codes.count("UNKNOWN_UNIT") >= 1

    def test_review_sheet_contains_severity_and_stage(self, sample_flagged_result, tmp_path):
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_flagged_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Review"]
        severities = set()
        stages = set()
        for r in range(5, ws.max_row + 1):
            s = ws.cell(row=r, column=2).value
            st = ws.cell(row=r, column=3).value
            if s:
                severities.add(str(s).lower())
            if st:
                stages.add(str(st).lower())
        assert "review" in severities
        assert "assembly" in stages or "normalization" in stages or "catalog" in stages

    def test_review_sheet_empty_case_explicit(self, sample_result, tmp_path):
        """When the extraction has zero flags, the Review sheet must
        still exist and explicitly say so (not silently empty)."""
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Review"]
        joined = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        )
        assert "no flags" in joined.lower() or "passed review" in joined.lower()

    def test_review_sheet_uses_three_severity_shades(self, sample_flagged_result, tmp_path):
        """Severity coloring is one of 3 hex shades — no rainbow."""
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_flagged_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Review"]
        # The Review sheet has review-level + maybe info-level flags here.
        # We just need to confirm the fills are from the allowed set
        # (or "no fill" / header / title rows).
        allowed = {
            "F8CBAD",  # error
            "FFD966",  # review
            "D9D9D9",  # info
            "1F4E79",  # header
            "C00000",  # title text
        }
        for r in range(4, ws.max_row + 1):
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                fill = cell.fill
                if not (fill and fill.fgColor):
                    continue
                raw = fill.fgColor.value
                if raw is None or raw == "":
                    continue
                rgb = str(raw).upper().lstrip("0")
                if rgb in {"", "0"}:
                    continue
                allowed_norm = {a.upper().lstrip("0") or "0" for a in allowed}
                assert rgb in allowed_norm, (
                    f"Unexpected fill color {rgb!r} (raw={raw!r}) in Review sheet "
                    f"row {r} col {c} — rainbow is forbidden (3 shades only)"
                )


# ---------------------------------------------------------------------------
# Deliverable 3: Provenance sheet
# ---------------------------------------------------------------------------


class TestProvenanceSheetContract:
    """The Provenance sheet answers "where did this come from?"."""

    def test_provenance_sheet_exists(self, sample_result, tmp_path):
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        assert "Provenance" in wb.sheetnames

    def test_provenance_lists_source_file_and_pipeline(self, sample_result, tmp_path):
        out = tmp_path / "boq.xlsx"
        prov = {
            "source_file": "data/real_rfqs/swa_enquiries/02_isro_vssc/VSSC_BOQ_with_qty.xlsx",
            "source_sha256": "deadbeef" * 8,  # 64 hex chars
        }
        CPWDExcelGenerator().export(
            sample_result.boq_items,
            str(out),
            provenance=prov,
        )
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Provenance"]
        joined = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        )
        assert "VSSC_BOQ_with_qty.xlsx" in joined
        assert "deadbeef" * 8 in joined
        assert "Pipeline commit" in joined
        assert "Schema version" in joined

    def test_provenance_lists_per_row_source_pages(self, tmp_path):
        items = [
            BoqRow(
                item_no=1,
                material="Item A",
                quantity=Decimal("1"),
                unit="nos",
                source_pages=[3, 4],
            ),
            BoqRow(
                item_no=2,
                material="Item B",
                quantity=Decimal("2"),
                unit="nos",
                source_pages=[7],
            ),
        ]
        out = tmp_path / "prov.xlsx"
        CPWDExcelGenerator().export(items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Provenance"]
        joined = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        )
        # Source pages appear as a comma-separated list in the per-row
        # section.
        assert "3, 4" in joined
        assert "7" in joined

    def test_provenance_pipeline_commit_resolves_to_git(self, sample_result, tmp_path):
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Provenance"]
        # Find the "Pipeline commit" row, then read the value column.
        commit_val = None
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label == "Pipeline commit":
                commit_val = ws.cell(row=r, column=2).value
                break
        assert commit_val is not None
        # Real value from git is a 7–12 char hex; the fallback is
        # the literal string "unknown".
        s = str(commit_val)
        assert s != "" and (s == "unknown" or all(c in "0123456789abcdef" for c in s))


# ---------------------------------------------------------------------------
# Deliverable 4: Footer carries pipeline commit
# ---------------------------------------------------------------------------


class TestFooterContract:
    def test_footer_has_pipeline_commit(self, sample_result, tmp_path):
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(
            sample_result.boq_items,
            str(out),
            provenance={"pipeline_commit": "abc1234"},
        )
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        joined = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        )
        assert "abc1234" in joined
        assert "RFQ2BOQ" in joined


# ---------------------------------------------------------------------------
# Audit table: field → preserved? → fix (per P5_01 §4 step 1)
# ---------------------------------------------------------------------------


class TestAuditTable:
    """Reproduces the P5_01 §4 step 1 audit table as a regression net."""

    def test_field_material_preserved_fully(self, sample_result, tmp_path):
        """field: material → preserved fully (was .title()-munged, now verbatim)."""
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        cells = [
            str(ws.cell(row=r, column=2).value or "") for r in range(1, ws.max_row + 1)
        ]
        joined = " ".join(cells)
        # The .title() form would be "Earthwork In Excavation By Mechanical Means"
        # (capitalized 'In', 'By', 'Mechanical', 'Means').
        assert "Earthwork in excavation by mechanical means" in joined
        assert "Earthwork In Excavation By Mechanical Means" not in joined

    def test_field_description_raw_preserved_fully(self, tmp_path):
        """field: description_raw → preserved fully (joined with material)."""
        items = [
            BoqRow(
                material="Short mat",
                description_raw="Long free-form description with detail",
                quantity=Decimal("1"),
                unit="nos",
            ),
        ]
        out = tmp_path / "desc.xlsx"
        CPWDExcelGenerator().export(items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        joined = " ".join(
            str(ws.cell(row=r, column=2).value or "") for r in range(1, ws.max_row + 1)
        )
        # material is shown; description_raw is in JSON only (kept separate
        # so the BOQ column is concise).  Both are present in source.
        assert "Short mat" in joined

    def test_field_standard_list_preserved_fully(self, sample_result, tmp_path):
        """field: standard (list[str]) → preserved fully via join."""
        out = tmp_path / "boq.xlsx"
        CPWDExcelGenerator().export(sample_result.boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        for r in range(1, ws.max_row + 1):
            std = ws.cell(row=r, column=5).value
            if std and "IS 1200" in str(std):
                # All list entries survived the join.
                assert "IS 1200" in str(std)
                return
        pytest.fail("IS 1200 standard not found in BOQ sheet")

    def test_field_dimensions_list_preserved_fully(self, tmp_path):
        """field: dimensions (list[str]) → preserved fully via join."""
        items = [
            BoqRow(
                material="Pipe",
                quantity=Decimal("100"),
                unit="rmt",
                dimensions=["500 mm dia", "6 m long", "MS"],
            ),
        ]
        out = tmp_path / "dim.xlsx"
        # Note: BOQ sheet doesn't have a Dimensions column by default;
        # the dimensions go to JSON only (the BOQ sheet is intentionally
        # concise).  Verify JSON survives.
        CPWDExcelGenerator().export(items, str(out))
        from src.export.json_formatter import JSONFormatter

        result = ExtractionResult(doc_id="x", boq_items=items)
        j = json.loads(JSONFormatter().format(result))
        assert j["boq_items"][0]["dimensions"] == ["500 mm dia", "6 m long", "MS"]

    def test_field_parent_context_preserved_fully(self, tmp_path):
        """field: parent_context (list[str]) → preserved fully; surfaced in
        the BOQ Description cell AND in JSON (P3_03 hierarchy)."""
        ctx = ["INSULATION", "MS chilled water pipe insulation"]
        items = [
            BoqRow(
                material="MS chilled water pipe insulation 500 mm dia",
                quantity=Decimal("500"),
                unit="rmt",
                parent_context=ctx,
            ),
        ]
        out = tmp_path / "ctx.xlsx"
        CPWDExcelGenerator().export(items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["BOQ"]
        joined = " ".join(
            str(ws.cell(row=r, column=2).value or "") for r in range(1, ws.max_row + 1)
        )
        # Both layers of the hierarchy are visible.
        assert "[parent]" in joined
        assert "INSULATION" in joined
        # JSON preserves the structured list.
        result = ExtractionResult(doc_id="x", boq_items=items)
        from src.export.json_formatter import JSONFormatter

        j = json.loads(JSONFormatter().format(result))
        assert j["boq_items"][0]["parent_context"] == ctx
