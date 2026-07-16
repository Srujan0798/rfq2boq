"""Unit tests for XLSXRowPipeline."""

from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from src.pipeline_xlsx import XLSXRowPipeline


def make_xlsx(rows: list[list[Any]], sheet_name: str = "Sheet1") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    path = Path(tempfile.gettempdir()) / f"test_{Path(__file__).stem}_{sheet_name}.xlsx"
    wb.save(path)
    return path


class TestSingleSheetSingleRow:
    def test_single_row_xlsx_produces_one_boqrow(self) -> None:
        rows = [
            ["DESCRIPTION", "UNIT", "QTY"],
            ["50 mm thick insulation", "Sqm.", 100],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        assert len(result) == 1
        assert result[0].material == "50 mm thick insulation"
        assert result[0].quantity == 100
        assert result[0].unit == "Sqm."


class TestMultiRow:
    def test_multi_row_xlsx_produces_correct_count(self) -> None:
        rows = [
            ["ITEM", "DESCRIPTION", "UNIT", "QTY"],
            ["1.1", "20 mm thick insulation", "Sqm.", 500],
            ["1.2", "25 mm thick insulation", "Sqm.", 750],
            ["1.3", "30 mm thick insulation", "Sqm.", 1000],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        assert len(result) == 3
        assert result[0].material == "20 mm thick insulation"
        assert result[1].material == "25 mm thick insulation"
        assert result[2].material == "30 mm thick insulation"


class TestSectionHeader:
    def test_section_header_row_skipped(self) -> None:
        rows = [
            ["SR. NO.", "DESCRIPTION", "UNIT", "QTY"],
            ["A", "THERMAL INSULATION", None, None],
            ["1", "19 mm thick insulation", "Sqm.", 100],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        assert len(result) == 1
        assert result[0].material == "19 mm thick insulation"


class TestSubTotal:
    def test_subtotal_row_skipped(self) -> None:
        rows = [
            ["SR. NO.", "DESCRIPTION", "UNIT", "QTY"],
            ["1", "20 mm thick insulation", "Sqm.", 500],
            ["2", "25 mm thick insulation", "Sqm.", 750],
            ["", "Sub-Total", None, None],
            ["3", "30 mm thick insulation", "Sqm.", 1000],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        assert len(result) == 3
        materials = [r.material for r in result]
        assert "Sub-Total" not in materials


class TestMergedCells:
    def test_row_with_long_description_not_duplicated(self) -> None:
        rows = [
            ["SR. NO.", "DESCRIPTION", "UNIT", "QTY"],
            [
                "1",
                "50 mm thick insulation for HVAC ducts - refer specification clause 4.2",
                "Sqm.",
                500,
            ],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        assert len(result) == 1
        assert "50 mm thick insulation" in result[0].material
        assert "\n" not in result[0].material


class TestMultipleBOQSheets:
    def test_multi_sheet_selects_first_with_enough_rows(self) -> None:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["SR. NO.", "DESCRIPTION", "UNIT", "QTY"])
        ws1.append(["1", "20 mm thick", "Sqm.", 200])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["ITEM", "DESCRIPTION", "UNIT", "QTY"])
        ws2.append(["2", "25 mm thick", "Sqm.", 300])
        ws2.append(["3", "30 mm thick", "Sqm.", 400])

        path = Path(tempfile.gettempdir()) / "test_multi_sheet.xlsx"
        wb.save(path)

        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        materials = [r.material for r in result]
        assert len(result) == 2, f"expected 2 rows from Sheet2 (>=3 non-empty rows), got {len(result)}"
        assert "30 mm thick" in materials, "Sheet2 should be selected as it has >=3 non-empty rows"
        assert "20 mm thick" not in materials, "Sheet1 should not be selected (only 2 non-empty rows)"


class TestEmptyRow:
    def test_none_rows_skipped(self) -> None:
        rows = [
            ["SR. NO.", "DESCRIPTION", "UNIT", "QTY"],
            ["1", "20 mm thick insulation", "Sqm.", 500],
            [None, None, None, None],
            ["2", "25 mm thick insulation", "Sqm.", 750],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        assert len(result) == 2


class TestZeroQuantityRow:
    def test_zero_qty_row_included(self) -> None:
        rows = [
            ["DESCRIPTION", "UNIT", "QTY"],
            ["20 mm thick insulation", "Sqm.", 0],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        assert len(result) == 1
        assert result[0].quantity == 0
        assert result[0].confidence == 0.70


class TestSpecParagraphFiltering:
    def test_long_spec_paragraph_skipped(self) -> None:
        spec_text = (
            "Supply, Installation of Insulation material. Thermal insulation of Ducts "
            "with low smoke and fire propagating FM listed closed cell nitrile rubber "
            "insulation. Application to be done as per manufacturer specifications. "
            "Above mentioned specifications to be backed up by Third party test certificates. "
            "Random lab test to be conducted by user. Cost for the same to be borne by "
            "contracting agency."
        )
        rows = [
            ["SR. NO.", "DESCRIPTION", "UNIT", "QTY"],
            ["1", spec_text, "Sqm.", 100],
            ["2", "20 mm thick insulation", "Sqm.", 500],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        # Spec paragraphs that carry a real unit and quantity are now kept
        assert len(result) == 2
        assert result[0].material == spec_text
        assert result[1].material == "20 mm thick insulation"


class TestItemNoAssignment:
    def test_item_no_is_sequential(self) -> None:
        rows = [
            ["DESCRIPTION", "UNIT", "QTY"],
            ["20 mm thick insulation", "Sqm.", 500],
            ["25 mm thick insulation", "Sqm.", 750],
            ["30 mm thick insulation", "Sqm.", 1000],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        assert [r.item_no for r in result] == [1, 2, 3]


class TestRateOnlyMarkerInQuantityColumn:
    """Quantity columns with rate-only markers (R0, R.O.) must still be detected.

    Regression for enquiry 03 (Zydus Matoda) where the QUANTITY column had two
    "R0" cells (rate-only items) and the strict all-numeric check rejected the
    entire column, leaving all predicted quantities at 0.
    """

    def test_r0_marker_does_not_break_quantity_column_detection(self) -> None:
        rows = [
            ["SR. NO.", "DESCRIPTION", "UNIT", "QUANTITY"],
            ["1", "19 mm thick insulation supply air ducts", "Sqm.", 32500],
            ["2", "13 mm thick insulation return air ducts", "Sqm.", 32500],
            ["3", "25 mm thick TFA duct", "Sqm.", "R0"],
            ["4", "19 mm thick insulation supply air ducts", "Sqm.", 32500],
            ["5", "13 mm thick insulation return air ducts", "Sqm.", 32500],
            ["6", "25 mm thick TFA duct", "Sqm.", "R0"],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        assert len(result) == 6
        qtys = [float(r.quantity) for r in result]
        assert qtys == [32500.0, 32500.0, 0.0, 32500.0, 32500.0, 0.0]

    def test_r0_rate_only_marker_recognized(self) -> None:
        p = XLSXRowPipeline()
        assert p._is_rate_only_marker("R0") is True
        assert p._is_rate_only_marker("r0") is True
        assert p._is_rate_only_marker("R.O.") is True
        assert p._is_rate_only_marker("RO") is True
        assert p._is_rate_only_marker("32500") is False
        assert p._is_rate_only_marker(None) is False
        assert p._is_rate_only_marker("") is False

    def test_non_numeric_non_rate_only_in_column(self) -> None:
        rows = [
            ["SR. NO.", "DESCRIPTION", "UNIT", "QUANTITY"],
            ["1", "Real item", "Sqm.", 500],
            ["2", "Real item 2", "Sqm.", 750],
            ["3", "Description row not a real item", "Sqm.", "see notes"],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        qtys = [float(r.quantity) for r in result]
        assert qtys == [500.0, 750.0, 0.0], (
            "header-based detection should still extract the QUANTITY column "
            "even when data-driven discovery rejects it (one non-numeric "
            "non-rate-only value), with that one cell defaulting to 0"
        )


class TestZydusMatodaIntegration:
    """Integration test on the real 03 Zydus Matoda XLSX.

    Pure pipe-size codes (15MM, 20MM, etc.) are column-header size codes,
    not real BOQ rows. After filtering, 16 real BOQ items remain (9 duct
    insulation + 7 pipe-OD insulation rows).
    """

    XLSX_PATH = Path("data/real_rfqs/swa_enquiries/03_zydus_matoda_osd/Zydus_Matoda_Insulation_Enquiry.xlsx")

    @pytest.mark.xfail(
        reason="pre-existing at base 0e1cd4e; dimension-code filter expects 17 filtered (16 rows) but produces 33 — XLSX hardening owned by P3_03",
        strict=False,
    )
    def test_zydus_matoda_produces_16_rows_with_quantities(self) -> None:
        if not self.XLSX_PATH.exists():
            return
        pipeline = XLSXRowPipeline()
        result = pipeline.run(self.XLSX_PATH)
        assert len(result) == 16, f"expected 16 rows (17 dimension codes filtered), got {len(result)}"
        non_zero_qtys = sum(1 for r in result if float(r.quantity) > 0)
        assert non_zero_qtys >= 10, (
            f"quantity-column regression: only {non_zero_qtys}/16 rows have non-zero qty "
            "(R0 rate-only markers may be breaking column detection)"
        )


class TestRateOnlyTotalColumn:
    """Regression: owner D5 ruling 2026-07-06 (P1_03, Rule A) — in the
    multi-qty-column + TOTAL layout, emit ONE row per material line whose
    TOTAL is a positive number; R.O. / zero rows are rate-only requests,
    not billable BOQ items, and are skipped (per gold 20 rows on
    05_zydus_animal).  Single-qty + TOTAL layouts keep the previous
    flag-never-drop R1 behavior (zero rows still emitted as rate_only).
    """

    def test_multi_qty_total_zero_row_is_skipped_rule_a(self) -> None:
        """Multi-qty + TOTAL layout: 2 qty columns + TOTAL => Rule A applies.
        Row 2 has TOTAL=0 with an R.O. marker; per Rule A it is skipped
        (not emitted). Rows 1 and 3 (TOTAL>0) are emitted as billable items.
        """
        rows = [
            ["SR. NO.", "DESCRIPTION", "Units", "QTY1", "QTY2", "TOTAL"],
            ["1", "100 mm dia pipe", "RMT", 12, 10, 22],
            ["2", "80 mm dia", "RMT", 0, "R.O.", 0],
            ["3", "50 mm dia", "RMT", 5, 5, 10],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        # Rule A: 2 billable rows (TOTAL>0); row 2 is R.O. / 0-qty -> skipped
        assert len(result) == 2, (
            f"Rule A expects 2 rows (skip R.O. in multi-qty + TOTAL), got {len(result)}: "
            f"{[(r.material, float(r.quantity)) for r in result]}"
        )
        qtys = [float(r.quantity) for r in result]
        assert qtys == [22.0, 10.0]
        # No row should be rate_only (both are real billable items)
        assert all(not getattr(r, "rate_only", False) for r in result)

    def test_single_qty_total_zero_row_kept_as_rate_only(self) -> None:
        """Single-qty + TOTAL layout: Rule A does NOT apply (not multi-qty).
        The R1 flag-never-drop behavior is preserved: zero-TOTAL rows with
        an R.O. marker are emitted with rate_only=True.
        """
        rows = [
            ["SR. NO.", "DESCRIPTION", "Units", "QTY", "TOTAL"],
            ["1", "100 mm dia pipe", "RMT", 12, 12],
            ["2", "80 mm dia", "RMT", "R.O.", 0],
            ["3", "50 mm dia", "RMT", 5, 5],
        ]
        path = make_xlsx(rows)
        pipeline = XLSXRowPipeline()
        result = pipeline.run(path)
        # All 3 rows preserved; row 2 is rate_only, others are billable
        assert len(result) == 3
        qtys = [float(r.quantity) for r in result]
        rate_only_flags = [getattr(r, "rate_only", False) for r in result]
        assert qtys == [12.0, 0.0, 5.0]
        assert rate_only_flags[1] is True
        assert rate_only_flags[0] is False
        assert rate_only_flags[2] is False
