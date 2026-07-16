"""Tests for R1 'flag, never drop' — confidence scoring and low-confidence row flagging.

Verifies:
- Low-confidence rows are NOT dropped — they appear in BOQ output.
- Confidence is set correctly based on presence/quality of {material, quantity, unit}.
- Excel export marks low-confidence rows with yellow fill + REVIEW note.
- JSON output includes confidence field on every row.
"""

import json
from datetime import datetime
from decimal import Decimal

import openpyxl
from config.constants import EntityType
from src.domain.boq_assembler import BOQAssembler
from src.domain.models import BoqRow, EntitySpan, ExtractionMetadata, ExtractionResult
from src.export.excel_generator import CPWDExcelGenerator
from src.export.json_formatter import JSONFormatter

# ---------------------------------------------------------------------------
# BOQAssembler confidence scoring
# ---------------------------------------------------------------------------


class TestComputeRowConfidence:
    def _mat(self, text: str, conf: float = 0.9) -> EntitySpan:
        return EntitySpan(text=text, type=EntityType.MATERIAL, start=0, end=len(text), page=1, conf=conf)

    def _qty(self, text: str, conf: float = 0.9) -> EntitySpan:
        return EntitySpan(text=text, type=EntityType.QUANTITY, start=20, end=25, page=1, conf=conf)

    def _unit(self, text: str, conf: float = 0.9) -> EntitySpan:
        return EntitySpan(text=text, type=EntityType.UNIT, start=26, end=30, page=1, conf=conf)

    def test_all_three_present_gives_high_confidence(self):
        mat = self._mat("cement", conf=0.95)
        qty = self._qty("500", conf=0.95)
        unit = self._unit("bags", conf=0.9)
        conf = BOQAssembler._compute_row_confidence(mat, qty, unit)
        assert conf >= 0.7, f"Expected >= 0.7, got {conf}"

    def test_missing_quantity_gives_low_confidence(self):
        mat = self._mat("steel", conf=0.9)
        conf = BOQAssembler._compute_row_confidence(mat, None, None)
        assert conf < 0.7, f"Expected < 0.7, got {conf}"

    def test_quantity_present_no_unit_mid_confidence(self):
        mat = self._mat("cement", conf=0.85)
        qty = self._qty("100", conf=0.85)
        conf = BOQAssembler._compute_row_confidence(mat, qty, None)
        assert conf < 0.7, f"Expected < 0.7 (unit guessed), got {conf}"

    def test_very_short_material_no_qty_gives_low_confidence(self):
        mat = self._mat("15mm", conf=0.8)
        conf = BOQAssembler._compute_row_confidence(mat, None, None)
        assert conf <= 0.3, f"Expected <= 0.3 for short label, got {conf}"

    def test_confidence_bounded_0_to_1(self):
        mat = self._mat("concrete", conf=0.99)
        qty = self._qty("1000", conf=0.99)
        unit = self._unit("cum", conf=0.99)
        conf = BOQAssembler._compute_row_confidence(mat, qty, unit)
        assert 0.0 <= conf <= 1.0


# ---------------------------------------------------------------------------
# BOQAssembler: low-confidence rows are NOT dropped
# ---------------------------------------------------------------------------


class TestLowConfidenceRowsNotDropped:
    def test_material_only_row_appears_in_output(self):
        """A row with no quantity is low-confidence but must NOT be dropped."""
        assembler = BOQAssembler()
        entities = [
            EntitySpan(text="reinforcement steel", type=EntityType.MATERIAL, start=0, end=19, page=1, conf=0.6),
        ]
        rows = assembler.assemble(entities, [], "reinforcement steel")
        assert len(rows) == 1, "Low-confidence row must not be dropped"
        assert rows[0].material == "reinforcement steel"
        assert rows[0].confidence < 0.7

    def test_low_confidence_row_has_warning(self):
        assembler = BOQAssembler()
        entities = [
            EntitySpan(text="pipe insulation", type=EntityType.MATERIAL, start=0, end=15, page=1, conf=0.5),
        ]
        rows = assembler.assemble(entities, [], "pipe insulation")
        assert len(rows) == 1
        assert "LOW_CONFIDENCE" in rows[0].warnings

    def test_high_confidence_row_no_warning(self):
        assembler = BOQAssembler()
        entities = [
            EntitySpan(text="cement", type=EntityType.MATERIAL, start=0, end=6, page=1, conf=0.95),
            EntitySpan(text="500", type=EntityType.QUANTITY, start=7, end=10, page=1, conf=0.95),
            EntitySpan(text="bags", type=EntityType.UNIT, start=11, end=15, page=1, conf=0.9),
        ]
        rows = assembler.assemble(entities, [], "cement 500 bags")
        assert len(rows) == 1
        assert "LOW_CONFIDENCE" not in rows[0].warnings
        assert rows[0].confidence >= 0.7


# ---------------------------------------------------------------------------
# Excel export: low-confidence rows flagged with yellow + REVIEW
# ---------------------------------------------------------------------------


class TestExcelLowConfidenceFlagging:
    def _make_items(self):
        return [
            {"material": "Cement M20", "quantity": 500, "unit": "bags", "confidence": 0.9},
            {"material": "Steel bars", "quantity": 0, "unit": "kg", "confidence": 0.4},
            {"material": "PVC pipe", "quantity": 100, "unit": "rmt", "confidence": 0.65},
        ]

    def test_low_confidence_rows_have_review_note(self, tmp_path):
        gen = CPWDExcelGenerator()
        out = tmp_path / "test_conf.xlsx"
        gen.export(self._make_items(), str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active

        # Column 8 = Remarks (0-indexed: 7)
        review_notes = []
        for row in ws.iter_rows(min_row=10):
            remarks_cell = row[7]  # column 8 = Remarks
            if remarks_cell.value and "REVIEW" in str(remarks_cell.value):
                review_notes.append(remarks_cell.value)

        assert len(review_notes) >= 2, (
            f"Expected at least 2 REVIEW notes for conf=0.4 and conf=0.65, got: {review_notes}"
        )

    def test_low_confidence_rows_have_yellow_fill(self, tmp_path):
        gen = CPWDExcelGenerator()
        out = tmp_path / "test_yellow.xlsx"
        gen.export(self._make_items(), str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active

        # Low-confidence highlight is on column 8 = Remarks (0-indexed: 7)
        # _LOW_CONF_BG = "FFF2CC" (light yellow)
        yellow_rows = []
        for row in ws.iter_rows(min_row=10):
            remarks_cell = row[7]  # column 8 = Remarks
            fill = remarks_cell.fill
            if fill and fill.fgColor:
                rgb = fill.fgColor.rgb
                # FFF2CC is the low-confidence fill colour; openpyxl prepends 00 for alpha
                if rgb in ("FFF2CC", "00FFF2CC", "FFFFF2CC"):
                    yellow_rows.append(remarks_cell.value)

        assert len(yellow_rows) >= 2, (
            f"Expected yellow fill on Remarks cell of low-conf rows, got: {yellow_rows}"
        )

    def test_high_confidence_row_not_flagged(self, tmp_path):
        gen = CPWDExcelGenerator()
        out = tmp_path / "test_no_flag.xlsx"
        items = [{"material": "Cement M20", "quantity": 500, "unit": "bags", "confidence": 0.95}]
        gen.export(items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active

        for row in ws.iter_rows(min_row=10):
            remarks_cell = row[7]  # column 8 = Remarks
            if remarks_cell.value:
                assert "REVIEW" not in str(remarks_cell.value), (
                    f"High-confidence row should not be flagged, got: {remarks_cell.value}"
                )


# ---------------------------------------------------------------------------
# JSON export: confidence field present on every row
# ---------------------------------------------------------------------------


class TestJsonConfidenceField:
    def _make_result(self, items: list[BoqRow]) -> ExtractionResult:
        return ExtractionResult(
            doc_id="test-r1",
            project_name="R1 Test",
            extraction_date=datetime.now(),
            source_file="test.pdf",
            boq_items=items,
            metadata=ExtractionMetadata(total_items=len(items)),
        )

    def test_confidence_in_json_output(self):
        items = [
            BoqRow(item_no=1, material="cement", quantity=Decimal("100"), unit="bags", confidence=0.9),
            BoqRow(item_no=2, material="steel", quantity=Decimal("0"), unit="kg", confidence=0.4),
        ]
        formatter = JSONFormatter()
        output = json.loads(formatter.format(self._make_result(items)))
        for boq_item in output["boq_items"]:
            assert "confidence" in boq_item, "confidence field missing from JSON output"

    def test_low_confidence_rows_not_dropped_in_json(self):
        """R1: low-confidence rows must appear in JSON output, not be silently dropped."""
        items = [
            BoqRow(item_no=1, material="cement", quantity=Decimal("100"), unit="bags", confidence=0.9),
            BoqRow(item_no=2, material="steel", quantity=Decimal("0"), unit="kg", confidence=0.35),
        ]
        formatter = JSONFormatter()
        output = json.loads(formatter.format(self._make_result(items)))
        assert len(output["boq_items"]) == 2, (
            "Low-confidence rows must NOT be dropped from JSON (R1: flag, never drop)"
        )
        confidences = [item["confidence"] for item in output["boq_items"]]
        assert any(c < 0.7 for c in confidences), "Low-confidence row missing from JSON"
