"""Tests for CPWD Excel BOQ generator (unpriced scope)."""

import openpyxl
import pytest
from src.export.excel_generator import CPWDExcelGenerator


@pytest.fixture
def sample_boq_items():
    return [
        {
            "material": "Earthwork in excavation by mechanical means",
            "quantity": 100,
            "unit": "cum",
            "description": "Earthwork in excavation by mechanical means",
        },
        {
            "material": "M20 grade concrete in foundation",
            "quantity": 50,
            "unit": "cum",
            "description": "M20 grade concrete in foundation",
        },
        {
            "material": "Brickwork in cement mortar 1:6",
            "quantity": 200,
            "unit": "cum",
            "description": "Brickwork in cement mortar 1:6",
        },
    ]


class TestCPWDExcelGenerator:
    def test_export_creates_file(self, sample_boq_items, tmp_path):
        gen = CPWDExcelGenerator()
        out = tmp_path / "test_boq.xlsx"
        gen.export(sample_boq_items, str(out), {"project": "Test Project", "location": "Delhi"})
        assert out.exists()
        assert out.stat().st_size > 1024

    def test_column_headers_correct(self, sample_boq_items, tmp_path):
        gen = CPWDExcelGenerator()
        out = tmp_path / "test_boq.xlsx"
        gen.export(sample_boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        headers = [ws.cell(row=9, column=c).value for c in range(1, 9)]
        expected = ["S.No", "Description", "Quantity", "Unit", "Standard", "Grade", "Confidence", "Remarks"]
        assert headers == expected

    def test_trade_grouping(self, sample_boq_items, tmp_path):
        gen = CPWDExcelGenerator()
        out = tmp_path / "test_boq.xlsx"
        gen.export(sample_boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        cell_values = [ws.cell(row=r, column=1).value for r in range(1, 60)]
        assert any("excavation" in str(v).lower() for v in cell_values if v)
        assert any("concrete" in str(v).lower() for v in cell_values if v)
        assert any("brickwork" in str(v).lower() for v in cell_values if v)

    def test_subtotal_label_present(self, sample_boq_items, tmp_path):
        gen = CPWDExcelGenerator()
        out = tmp_path / "test_boq.xlsx"
        gen.export(sample_boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        labels = []
        for row in range(1, 60):
            val = ws.cell(row=row, column=1).value
            if val and isinstance(val, str) and val.startswith("Subtotal"):
                labels.append(val)
        assert len(labels) >= 1, "No subtotal labels found"

    def test_grand_total_label_present(self, sample_boq_items, tmp_path):
        gen = CPWDExcelGenerator()
        out = tmp_path / "test_boq.xlsx"
        gen.export(sample_boq_items, str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        found = False
        for row in range(1, 60):
            cell = ws.cell(row=row, column=1)
            if cell.value and "GRAND TOTAL" in str(cell.value):
                found = True
                break
        assert found, "GRAND TOTAL row not found"

    def test_project_metadata_in_header(self, sample_boq_items, tmp_path):
        gen = CPWDExcelGenerator()
        out = tmp_path / "test_boq.xlsx"
        gen.export(
            sample_boq_items,
            str(out),
            {"project": "Test Project", "location": "Delhi", "reference": "REF/001", "contractor": "ABC Corp"},
        )
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "Test Project"
        assert ws.cell(row=3, column=2).value == "Delhi"
        assert ws.cell(row=5, column=2).value == "REF/001"

    def test_boq_row_model_compatible(self):
        from src.domain.models import BoqRow

        row = BoqRow(material="Test", quantity=10, unit="nos", rate=100)
        gen = CPWDExcelGenerator()
        d = gen._as_dict(row)
        assert d["material"] == "Test"
        assert d["quantity"] == 10

    def test_empty_boq_export(self, tmp_path):
        gen = CPWDExcelGenerator()
        out = tmp_path / "empty_boq.xlsx"
        gen.export([], str(out), {"project": "Empty Test"})
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "BILL OF QUANTITIES (CPWD FORMAT)"

    def test_trade_detection(self, sample_boq_items):
        gen = CPWDExcelGenerator()
        assert gen._detect_trade("earthwork in excavation") == "excavation"
        assert gen._detect_trade("M20 concrete") == "concrete"
        assert gen._detect_trade("random material") == "general"

    def test_amount_is_zero_stub(self, sample_boq_items):
        gen = CPWDExcelGenerator()
        item = {"quantity": 10, "rate": 500}
        assert gen._get_amount(item) == 0.0
